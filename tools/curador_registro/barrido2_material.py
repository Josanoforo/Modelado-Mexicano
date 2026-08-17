#!/usr/bin/env python3
"""Motor material offline de BARRIDO-2.

Este módulo no conoce necesidades, relaciones ni adjudicaciones. Separa
declaración (payload), representación física y contenido; prepara tareas
cegadas y caracteriza E0/E1/E2 sin persistir observaciones individuales.
"""

from __future__ import annotations

import csv
import hashlib
import io
import itertools
import json
import os
import re
import shutil
import socket
import struct
import subprocess
import tempfile
import zipfile
from collections import Counter, defaultdict
from datetime import date
from html.parser import HTMLParser
from pathlib import Path, PurePosixPath
from typing import Any, BinaryIO, Iterable

import olefile
import openpyxl
import yaml
from jsonschema import Draft202012Validator


AUTHORIZED_ROOTS = ("data_raw", "descargas_mx")
MATERIAL_BUILD_VERSION = "BARRIDO2-MATERIAL-1.0"
MATERIAL_BUILD_SHA256 = hashlib.sha256(Path(__file__).read_bytes()).hexdigest()
PRIVACY_CONTRACT = "BARRIDO2-PRIVACY-1.0"
HASH_RE = re.compile(r"^[0-9a-f]{64}$")
PAYLOAD_ID_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,159}$")
W2_SUFFIXES = {".pdf", ".xls", ".xlsx", ".csv", ".tsv", ".dta", ".sav", ".docx"}
WAVE_MAX_WORKERS = {"W1": 3, "W2": 3, "W3": 2, "W4": 1, "W5": 3}
TEXT_SUFFIXES = {".txt", ".php", ".md", ".log", ".dat"}
E0_TERMINALS = {
    "PRESENTE-INTEGRO", "PRESENTE-HASH-DIVERGENTE",
    "PRESENTE-TAMANO-DIVERGENTE", "FUERA-DE-DISCO",
    "RAIZ-NO-CONFIGURADA", "RUTA-INVALIDA", "CORRUPTO", "CIFRADO",
    "FORMATO-NO-SOPORTADO", "NO-DETERMINADO",
}
TASK_ALLOWED = {
    "tarea_id", "representacion_id", "payload_id", "root_id",
    "ruta_relativa", "sha256", "formato", "profundidad", "presupuesto",
    "contrato_sha256", "wave_initial", "network_habilitada",
}
PII_PATTERNS = (
    re.compile(r"\b[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3}\b", re.I),
    re.compile(r"\b[A-Z]{4}\d{6}[HM][A-Z]{5}[A-Z0-9]\d\b", re.I),
    re.compile(r"\b[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}\b"),
    re.compile(r"(?<!\d)(?:\+?52[ -]?)?(?:\d[ -]?){10}(?!\d)"),
    re.compile(r"\b[0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}\b", re.I),
    re.compile(r"(?<!\d)\d{11,18}(?!\d)"),
    re.compile(r"\b(?:calle|avenida|av\.?|domicilio|direcci[oó]n)\b.{0,80}\d", re.I),
    re.compile(
        r"\b[A-ZÁÉÍÓÚÑ][a-záéíóúñ]{2,}(?:[ _-]+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]{2,}){1,3}\b"
    ),
    re.compile(r"^[A-ZÁÉÍÓÚÑ]{3,}(?:[ _-]+[A-ZÁÉÍÓÚÑ]{3,}){1,3}$"),
    re.compile(r"^[a-záéíóúñ]{3,}(?:[ _-]+[a-záéíóúñ]{3,}){1,3}$"),
    re.compile(r"^(?:/|[A-Za-z]:[/\\])"),
)


class MaterialDriftError(RuntimeError):
    """Los bytes o la geometría material cambiaron durante la lectura."""


class NetworkIsolationError(RuntimeError):
    """El proceso material no está en un namespace de red aislado."""


def canonical_bytes(value: object) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")


def canonical_sha(value: object) -> str:
    return hashlib.sha256(canonical_bytes(value)).hexdigest()


def sha256_file(path: Path) -> str:
    before = path.stat()
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    after = path.stat()
    if (before.st_dev, before.st_ino, before.st_size, before.st_mtime_ns) != (
        after.st_dev, after.st_ino, after.st_size, after.st_mtime_ns
    ):
        raise MaterialDriftError(f"MATERIAL_CAMBIO_DURANTE_HASH:{path}")
    return digest.hexdigest()


def material_build_sha256() -> str:
    return MATERIAL_BUILD_SHA256


def _atomic_write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def _atomic_write_lines(path: Path, lines: Iterable[str]) -> None:
    """Escribe un artefacto grande sin duplicarlo completo en memoria."""
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as handle:
            for line in lines:
                handle.write(line)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def _atomic_write_tsv_rows(
    path: Path, fieldnames: list[str], rows: Iterable[dict[str, str]]
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle, fieldnames=fieldnames, delimiter="\t", lineterminator="\n"
            )
            writer.writeheader()
            writer.writerows(rows)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def normalize_relative(value: str) -> str:
    raw = str(value or "").strip().replace("\\", "/")
    path = PurePosixPath(raw)
    if not raw or path.is_absolute() or any(part in {"", ".", ".."} for part in path.parts):
        raise ValueError(f"RUTA_INVALIDA:{value}")
    return path.as_posix()


def representation_id(root_id: str, relative: str, digest: str) -> str:
    relative = normalize_relative(relative)
    if root_id not in AUTHORIZED_ROOTS or not HASH_RE.fullmatch(digest):
        raise ValueError("IDENTIDAD_REPRESENTACION_INVALIDA")
    token = hashlib.sha256(
        f"{root_id}\x00{relative}\x00{digest}".encode("utf-8")
    ).hexdigest()
    return "REP-" + token


def logical_object_id(digest: str, locator: str) -> str:
    locator = "/".join(part for part in str(locator).replace("\\", "/").split("/") if part not in {"", "."})
    return "OBJ-B2-" + hashlib.sha256(
        f"{digest}\x00{locator}".encode("utf-8")
    ).hexdigest()


def safe_text(value: object, *, durable: bool = False) -> tuple[str, bool]:
    text = " ".join(str(value or "").replace("\t", " ").replace("\r", " ").replace("\n", " ").split())
    redacted = any(pattern.search(text) for pattern in PII_PATTERNS)
    if redacted:
        text = "[REDACTADO-PRIVACIDAD]"
    if not text:
        text = "NO-APLICA"
    if durable:
        text = text[:160].rstrip()
    return text, redacted


def valid_payload_id(value: object, *, allow_no_aplica: bool = False) -> bool:
    """Valida la identidad administrativa; no la interpreta como prosa.

    Los ids del manifiesto son *slugs* controlados, no texto extraído.  Aplicar
    las heurísticas de nombres propios a ellos rechazaría ids históricos
    legítimos como ``nota_metodologica_rotulo_pareada``. La gramática cerrada
    excluye espacios, correos y otros valores narrativos; el marcador reservado
    sólo se admite expresamente para físicos no declarados.
    """
    token = str(value or "").strip()
    if token == "NO-APLICA":
        return allow_no_aplica
    return bool(PAYLOAD_ID_RE.fullmatch(token))


def load_roots(config_path: Path) -> tuple[dict[str, Path], str]:
    payload = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
    if not isinstance(payload, dict):
        raise ValueError("RAICES_CONFIG_NO_MAPA")
    roots: dict[str, Path] = {}
    for root_id in AUTHORIZED_ROOTS:
        raw = str(payload.get(root_id, "") or "").strip()
        if not raw:
            raise ValueError(f"RAIZ_NO_CONFIGURADA:{root_id}")
        root = Path(raw).resolve()
        if not root.is_dir():
            raise ValueError(f"RAIZ_NO_EXISTE:{root_id}")
        roots[root_id] = root
    left, right = (roots[root_id] for root_id in AUTHORIZED_ROOTS)
    try:
        left.relative_to(right)
        overlap = True
    except ValueError:
        try:
            right.relative_to(left)
            overlap = True
        except ValueError:
            overlap = False
    if overlap:
        raise ValueError("RAICES_SOLAPADAS")
    # Hash durable: IDs y valores textuales exactos; ninguna salida durable
    # incluye esos valores ni una ruta absoluta.
    selected = {root_id: str(payload[root_id]) for root_id in AUTHORIZED_ROOTS}
    return roots, canonical_sha(selected)


def assert_network_disabled() -> None:
    self_net = os.stat("/proc/self/ns/net").st_ino
    outer_declared = os.environ.get("BARRIDO2_OUTER_NET_NS_INODE", "")
    try:
        comparison_net = int(outer_declared) if outer_declared else os.stat("/proc/1/ns/net").st_ino
    except PermissionError:
        # En `unshare -Urn`, el user namespace puede impedir leer ns/net de
        # PID 1. El mapeo unitario demuestra que unshare -Ur sí ocurrió; la
        # sonda negativa de abajo demuestra el aislamiento efectivo de salida.
        uid_map = Path("/proc/self/uid_map").read_text(encoding="ascii").split()
        if len(uid_map) != 3 or uid_map[0] != "0" or uid_map[2] != "1":
            raise NetworkIsolationError("NAMESPACE_RED_NO_COMPARABLE_Y_USER_NS_NO_UNITARIO")
        comparison_net = -1
    except (OSError, ValueError) as exc:
        raise NetworkIsolationError(f"NAMESPACE_NO_VERIFICABLE:{type(exc).__name__}") from exc
    if self_net == comparison_net:
        raise NetworkIsolationError("NAMESPACE_RED_NO_AISLADO")
    try:
        interfaces = {name for _, name in socket.if_nameindex()}
    except OSError:
        interfaces = {
            line.split(":", 1)[0].strip()
            for line in Path("/proc/net/dev").read_text(encoding="ascii", errors="replace").splitlines()[2:]
            if ":" in line
        }
    if interfaces - {"lo"}:
        raise NetworkIsolationError(f"NAMESPACE_RED_INTERFAZ_NO_LOOPBACK:{','.join(sorted(interfaces))}")
    routes = Path("/proc/net/route").read_text(encoding="ascii", errors="replace").splitlines()[1:]
    if any(line.split()[1] == "00000000" for line in routes if len(line.split()) >= 2):
        raise NetworkIsolationError("NAMESPACE_RED_RUTA_DEFAULT")
    for host, port in (("1.1.1.1", 53), ("8.8.8.8", 53), ("93.184.216.34", 80)):
        try:
            with socket.create_connection((host, port), timeout=0.5):
                raise NetworkIsolationError(f"NAMESPACE_RED_CON_EGRESO:{host}:{port}")
        except NetworkIsolationError:
            raise
        except OSError:
            pass


def magic_label(head: bytes) -> str:
    signatures = (
        (b"PK\x03\x04", "ZIP"), (b"%PDF-", "PDF"),
        (b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1", "OLE"),
        (b"\x7fELF", "ELF"), (b"SQLite format 3\x00", "SQLITE"),
        (b"\x89PNG\r\n\x1a\n", "PNG"), (b"\xff\xd8\xff", "JPEG"),
    )
    for signature, label in signatures:
        if head.startswith(signature):
            return label
    if head and b"\x00" not in head and sum(byte in b"\t\n\r" or 32 <= byte < 127 for byte in head) / len(head) > 0.8:
        return "TEXTO-PROBABLE"
    return "NO-DETERMINADO"


def zip_geometry(path: Path) -> dict[str, Any]:
    result: dict[str, Any] = {
        "container_integrity": "NO-APLICA", "encrypted": False,
        "zip_slip": False, "uncompressed": 0, "max_member": 0,
        "max_ratio": 0.0, "members": 0,
    }
    if path.suffix.casefold() not in {".zip", ".xlsx", ".docx"}:
        return result
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            result["members"] = len(infos)
            for info in infos:
                normalized = PurePosixPath(info.filename.replace("\\", "/"))
                if normalized.is_absolute() or ".." in normalized.parts:
                    result["zip_slip"] = True
                result["encrypted"] = result["encrypted"] or bool(info.flag_bits & 0x1)
                result["uncompressed"] += info.file_size
                result["max_member"] = max(result["max_member"], info.file_size)
                ratio = info.file_size / max(info.compress_size, 1)
                result["max_ratio"] = max(result["max_ratio"], ratio)
            result["container_integrity"] = "CENTRAL-DIRECTORY-INTEGRO"
    except zipfile.BadZipFile:
        result["container_integrity"] = "CORRUPTO"
    return result


def assign_wave(relative: str, size: int, geometry: dict[str, Any] | None = None) -> str:
    suffix = Path(relative).suffix.casefold()
    geometry = geometry or {}
    risky = (
        size >= 512 * 1024**2
        or int(geometry.get("max_member", 0)) > 8 * 1024**3
        or float(geometry.get("max_ratio", 0.0)) > 200
        or int(geometry.get("uncompressed", 0)) >= 2 * 1024**3
    )
    if risky:
        return "W4"
    if suffix == ".zip":
        return "W3"
    if suffix in W2_SUFFIXES:
        return "W2"
    return "W1"


def wave_concurrency_limit(wave: str, formats: Iterable[str] = ()) -> int:
    if wave not in WAVE_MAX_WORKERS:
        raise ValueError(f"OLA_INVALIDA:{wave}")
    normalized = {str(value).casefold().lstrip(".") for value in formats}
    if wave == "W2" and normalized.intersection({"pdf", "xls", "xlsx"}):
        return 2
    return WAVE_MAX_WORKERS[wave]


def _manifest_rows(path: Path) -> list[dict[str, Any]]:
    payload = yaml.safe_load(path.read_text(encoding="utf-8")) or []
    if not isinstance(payload, list) or any(not isinstance(row, dict) for row in payload):
        raise ValueError("MANIFIESTO_NO_LISTA_DE_MAPAS")
    return payload


def _physical_files(root: Path) -> list[tuple[Path, Path, tuple[Any, ...]]]:
    """Enumera archivos sin permitir que un symlink saque la lectura de raíz."""
    files: list[tuple[Path, Path, tuple[Any, ...]]] = []
    for path in root.rglob("*"):
        if path.is_symlink():
            try:
                resolved = path.resolve(strict=True)
                resolved_relative = resolved.relative_to(root).as_posix()
            except (OSError, ValueError) as exc:
                raise MaterialDriftError(
                    f"SYMLINK_FUERA_DE_RAIZ:{path.relative_to(root).as_posix()}"
                ) from exc
            if resolved.is_dir():
                # El corpus histórico contiene ``data/raw -> <raíz>`` como
                # alias autorreferente. No es una representación física y
                # seguirlo duplicaría (o ciclaría) el universo. Sólo este caso
                # inode-equivalente se ignora; cualquier otro enlace de
                # directorio permanece fail-closed.
                if resolved == root:
                    continue
                raise MaterialDriftError(
                    f"SYMLINK_DIRECTORIO_NO_ENUMERABLE:{path.relative_to(root).as_posix()}"
                )
        else:
            resolved = path
            resolved_relative = path.relative_to(root).as_posix()
        if not resolved.is_file():
            continue
        stat = resolved.stat()
        link_stat = path.lstat()
        guard = (
            path.relative_to(root).as_posix(), resolved_relative,
            stat.st_size, stat.st_mtime_ns, stat.st_ino,
            link_stat.st_mtime_ns, link_stat.st_ino,
        )
        files.append((path, resolved, guard))
    return sorted(files, key=lambda item: item[0].relative_to(root).as_posix())


def build_material_snapshot(manifest_path: Path, roots_config: Path, output_path: Path) -> dict[str, Any]:
    """Censa las dos raíces sin colapsar declaración, representación y SHA."""
    roots, roots_hash = load_roots(roots_config)
    resolved_output = output_path.resolve()
    for root_id, root in roots.items():
        try:
            resolved_output.relative_to(root)
        except ValueError:
            continue
        raise ValueError(f"SNAPSHOT_OUTPUT_DENTRO_DE_RAIZ:{root_id}")
    manifest_hash = sha256_file(manifest_path)
    manifest = _manifest_rows(manifest_path)
    if sha256_file(manifest_path) != manifest_hash:
        raise MaterialDriftError("MANIFIESTO_CAMBIO_DURANTE_SNAPSHOT")
    payload_ids = [str(row.get("id", "") or "").strip() for row in manifest]
    if "" in payload_ids or len(payload_ids) != len(set(payload_ids)):
        raise ValueError("PAYLOAD_IDS_VACIOS_O_DUPLICADOS")
    if any(not valid_payload_id(payload_id) for payload_id in payload_ids):
        raise ValueError("PAYLOAD_ID_RESERVADO_LARGO_O_PRIVADO")

    physical: dict[tuple[str, str], dict[str, Any]] = {}
    inventory_guards: dict[str, list[tuple[Any, ...]]] = {}
    for root_id in AUTHORIZED_ROOTS:
        root = roots[root_id]
        paths = _physical_files(root)
        inventory_guards[root_id] = [guard for _, _, guard in paths]
        for path, resolved, guard in paths:
            relative = normalize_relative(path.relative_to(root).as_posix())
            digest = sha256_file(resolved)
            size = guard[2]
            with resolved.open("rb") as handle:
                head = handle.read(32)
            geometry = zip_geometry(resolved)
            state = "PRESENTE-INTEGRO"
            if geometry["container_integrity"] == "CORRUPTO":
                state = "CORRUPTO"
            elif geometry["encrypted"]:
                state = "CIFRADO"
            physical[(root_id, relative)] = {
                "representacion_id": representation_id(root_id, relative, digest),
                "root_id": root_id,
                "ruta_relativa": relative,
                "sha256": digest,
                "tamano_observado": size,
                "extension": Path(relative).suffix.casefold() or "SIN-EXTENSION",
                "magic": magic_label(head),
                "legible": True,
                "integridad_contenedor": geometry["container_integrity"],
                "zip_slip": geometry["zip_slip"],
                "estado_e0": state,
                "wave_initial": assign_wave(relative, size, geometry),
                "payload_ids": [],
                "duplicate_content_count": 0,
                "zip_geometry": geometry,
            }
    for root_id in AUTHORIZED_ROOTS:
        root = roots[root_id]
        after = [guard for _, _, guard in _physical_files(root)]
        if inventory_guards[root_id] != after:
            raise MaterialDriftError(f"INVENTARIO_RAIZ_CAMBIO_DURANTE_SNAPSHOT:{root_id}")

    declarations: list[dict[str, Any]] = []
    declaration_by_key: dict[tuple[str, str], list[str]] = defaultdict(list)
    for source in manifest:
        payload_id = str(source.get("id", "") or "").strip()
        root_id = str(source.get("raiz", "") or "data_raw").strip()
        raw_path = str(source.get("archivo", "") or "").strip()
        declared_hash = str(source.get("sha256", "") or "").strip().casefold()
        raw_size = source.get("tamano_bytes", "")
        declared_size = int(raw_size) if str(raw_size).strip().isdigit() else None
        row: dict[str, Any] = {
            "payload_id": payload_id, "root_id": root_id,
            "ruta_relativa": raw_path or "NO-APLICA",
            "sha256_declarado": declared_hash or "NO-APLICA",
            "tamano_declarado": declared_size if declared_size is not None else "NO-APLICA",
            "representacion_id": "NO-APLICA", "sha256_observado": "NO-APLICA",
            "estado_e0": "NO-DETERMINADO", "estado_administrativo": "NO-DETERMINADO",
        }
        if not raw_path and not declared_hash:
            row["estado_administrativo"] = "DECLARACION-SIN-ARCHIVO-SHA"
            declarations.append(row)
            continue
        if root_id not in roots:
            row.update(estado_e0="RAIZ-NO-CONFIGURADA", estado_administrativo="TERMINAL")
            declarations.append(row)
            continue
        try:
            relative = normalize_relative(raw_path)
        except ValueError:
            row.update(estado_e0="RUTA-INVALIDA", estado_administrativo="TERMINAL")
            declarations.append(row)
            continue
        row["ruta_relativa"] = relative
        representation = physical.get((root_id, relative))
        if representation is None:
            row.update(estado_e0="FUERA-DE-DISCO", estado_administrativo="TERMINAL")
            declarations.append(row)
            continue
        observed_hash = representation["sha256"]
        observed_size = representation["tamano_observado"]
        state = representation["estado_e0"]
        hash_valid = bool(HASH_RE.fullmatch(declared_hash))
        row["hash_coincide_manifiesto"] = "SI" if hash_valid and declared_hash == observed_hash else "NO"
        row["tamano_coincide_manifiesto"] = "SI" if declared_size is not None and declared_size == observed_size else "NO"
        if declared_hash and not hash_valid:
            if state not in {"CORRUPTO", "CIFRADO"}:
                state = "NO-DETERMINADO"
            row["razon_e0"] = "SHA-DECLARADO-INVALIDO"
        elif not declared_hash:
            if state not in {"CORRUPTO", "CIFRADO"}:
                state = "NO-DETERMINADO"
            row["razon_e0"] = "SHA-DECLARADO-AUSENTE"
        elif state not in {"CORRUPTO", "CIFRADO"} and declared_hash != observed_hash:
            state = "PRESENTE-HASH-DIVERGENTE"
        elif state not in {"CORRUPTO", "CIFRADO"} and declared_size is not None and declared_size != observed_size:
            state = "PRESENTE-TAMANO-DIVERGENTE"
        row.update(
            representacion_id=representation["representacion_id"],
            sha256_observado=observed_hash, estado_e0=state,
            estado_administrativo="TERMINAL",
        )
        declaration_by_key[(root_id, relative)].append(payload_id)
        declarations.append(row)

    content_counts = Counter(row["sha256"] for row in physical.values())
    for key, representation in physical.items():
        payloads = sorted(declaration_by_key.get(key, []))
        representation["payload_ids"] = payloads or ["NO-APLICA"]
        representation["coincidencia_manifiesto"] = "DECLARADA" if payloads else "NO-DECLARADA"
        representation["duplicate_content_count"] = content_counts[representation["sha256"]]

    representations = sorted(physical.values(), key=lambda row: row["representacion_id"])
    contents = [
        {"sha256": digest, "representaciones": count}
        for digest, count in sorted(content_counts.items())
    ]
    core = {
        "schema_version": "BARRIDO2-T0-1.0",
        "manifest_sha": manifest_hash,
        "roots_config_sha256": roots_hash,
        "authorized_roots": list(AUTHORIZED_ROOTS),
        "network_habilitada": False,
        "declarations": declarations,
        "representations": representations,
        "contents": contents,
    }
    core["counts"] = {
        "declaraciones_totales": len(declarations),
        "declaraciones_con_archivo_sha": sum(
            row["ruta_relativa"] != "NO-APLICA" and row["sha256_declarado"] != "NO-APLICA"
            for row in declarations
        ),
        "declaraciones_sin_archivo_sha": sum(
            row["estado_administrativo"] == "DECLARACION-SIN-ARCHIVO-SHA" for row in declarations
        ),
        "representaciones_fisicas": len(representations),
        "contenidos_sha_unicos": len(contents),
        "representaciones_declaradas": sum(row["coincidencia_manifiesto"] == "DECLARADA" for row in representations),
        "representaciones_no_declaradas": sum(row["coincidencia_manifiesto"] == "NO-DECLARADA" for row in representations),
        "fuera_de_disco": sum(row["estado_e0"] == "FUERA-DE-DISCO" for row in declarations),
    }
    core["snapshot_sha256"] = canonical_sha(core)
    _atomic_write_text(output_path, json.dumps(core, ensure_ascii=False, indent=2, sort_keys=True) + "\n")
    return core


def materialize_tasks(
    snapshot_path: Path,
    contract_path: Path,
    task_root: Path,
    ledger_path: Path,
    staging_root: Path | None = None,
) -> dict[str, int]:
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    snapshot_errors = validate_material_snapshot(snapshot)
    if snapshot_errors:
        raise ValueError("SNAPSHOT_MATERIAL_INVALIDO:" + ";".join(snapshot_errors))
    contract_hash = sha256_file(contract_path)
    task_root.mkdir(parents=True, exist_ok=True)
    existing: dict[str, dict[str, str]] = {}
    if ledger_path.is_file():
        with ledger_path.open(encoding="utf-8-sig", newline="") as handle:
            existing = {row["representacion_id"]: row for row in csv.DictReader(handle, delimiter="\t")}
    summaries: dict[str, tuple[dict[str, Any], Path]] = {}
    if staging_root is not None and staging_root.is_dir():
        for summary_path in staging_root.rglob("resumen.json"):
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
            rep_id = str(summary.get("representacion_id", ""))
            if rep_id in summaries:
                raise ValueError(f"EXPEDIENTE_REPRESENTACION_NO_UNICO:{rep_id}")
            summaries[rep_id] = (summary, summary_path.parent)
    rows: list[dict[str, str]] = []
    current_task_ids: set[str] = set()
    for representation in snapshot["representations"]:
        rep_id = representation["representacion_id"]
        task_id = "TASK-B2-" + hashlib.sha256(
            f"{snapshot['snapshot_sha256']}\x00{rep_id}\x00{contract_hash}".encode("utf-8")
        ).hexdigest()
        task = {
            "tarea_id": task_id,
            "representacion_id": rep_id,
            "payload_id": sorted(representation["payload_ids"])[0],
            "root_id": representation["root_id"],
            "ruta_relativa": representation["ruta_relativa"],
            "sha256": representation["sha256"],
            "formato": representation["extension"],
            "profundidad": "E2-COMPLETO",
            "presupuesto": {
                "timeout_segundos": 1800,
                "memoria_mib": 4096,
                "miembro_max_bytes": 8 * 1024**3,
                "temp_max_bytes": "MIN-50GIB-10PORCIENTO-LIBRE",
            },
            "contrato_sha256": contract_hash,
            "wave_initial": representation["wave_initial"],
            "network_habilitada": False,
        }
        current_task_ids.add(task_id)
        if set(task) != TASK_ALLOWED:
            raise AssertionError("CONTRATO_TAREA_CAMBIADO")
        task_path = task_root / f"{task_id}.json"
        _atomic_write_text(task_path, json.dumps(task, ensure_ascii=False, indent=2, sort_keys=True) + "\n")
        old = existing.get(rep_id, {})
        summary_entry = summaries.get(rep_id)
        summary, directory = summary_entry if summary_entry else ({}, Path("."))
        index_path = directory / "e2-neutral-index.jsonl"
        report_path = directory / "reportes-durables.tsv"
        compatible = bool(summary_entry) and (
            summary.get("tarea_id") == task_id
            and summary.get("tarea_sha256") == sha256_file(task_path)
            and summary.get("payload_id") == task["payload_id"]
            and summary.get("sha256") == representation["sha256"]
            and summary.get("formato") == task["formato"]
            and summary.get("contrato_sha256") == contract_hash
            and summary.get("profundidad") == "E2-COMPLETO"
            and summary.get("parser_version") == MATERIAL_BUILD_VERSION
            and summary.get("build_sha256") == material_build_sha256()
            and summary.get("privacidad") == PRIVACY_CONTRACT
            and summary.get("network_habilitada") is False
            and bool(summary.get("frontera_inspeccion"))
            and index_path.is_file() and report_path.is_file()
            and summary.get("index_sha256") == sha256_file(index_path)
            and summary.get("report_sha256") == sha256_file(report_path)
            and _completed_expediente_matches_task(summary, directory, task, task_path)
            and (
                old.get("estado_terminal") != "SI"
                or old.get("frontera_inspeccion") == summary.get("frontera_inspeccion")
            )
            and (
                old.get("estado_terminal") != "SI"
                or old.get("parser_version") == MATERIAL_BUILD_VERSION
            )
        )
        rows.append({
            "tarea_id": task_id, "representacion_id": rep_id,
            "payload_id": task["payload_id"], "root_id": representation["root_id"],
            "payload_ids_json": json.dumps(sorted(representation["payload_ids"]), ensure_ascii=False, separators=(",", ":")),
            "ruta_relativa": representation["ruta_relativa"], "sha256": representation["sha256"],
            "wave_initial": representation["wave_initial"], "wave_retry_ref": "NO-APLICA",
            "contrato_sha256": contract_hash,
            "parser": str(summary.get("parser", "NO-APLICA")) if compatible else "NO-APLICA",
            "parser_version": str(summary.get("parser_version", "NO-APLICA")) if compatible else "NO-APLICA",
            "build_sha256": str(summary.get("build_sha256", "NO-APLICA")) if compatible else "NO-APLICA",
            "profundidad": "E2-COMPLETO",
            "frontera_inspeccion": str(summary.get("frontera_inspeccion", "NO-APLICA")) if compatible else "NO-APLICA",
            "privacidad": PRIVACY_CONTRACT,
            "estado": "COMPLETADO" if compatible else "PENDIENTE",
            "estado_terminal": "SI" if compatible else "NO",
            "reporte_sha256": str(summary.get("report_sha256", "NO-APLICA")) if compatible else "NO-APLICA",
            "fecha": date.today().isoformat(),
        })
    fields = [
        "tarea_id", "representacion_id", "payload_id", "payload_ids_json", "root_id", "ruta_relativa",
        "sha256", "wave_initial", "wave_retry_ref", "contrato_sha256", "estado",
        "parser", "parser_version", "build_sha256", "profundidad", "frontera_inspeccion", "privacidad",
        "estado_terminal", "reporte_sha256", "fecha",
    ]
    ledger_buffer = io.StringIO(newline="")
    writer = csv.DictWriter(ledger_buffer, fieldnames=fields, delimiter="\t", lineterminator="\n")
    writer.writeheader(); writer.writerows(sorted(rows, key=lambda row: row["representacion_id"]))
    _atomic_write_text(ledger_path, ledger_buffer.getvalue())
    for stale_task in task_root.glob("*.json"):
        if stale_task.stem not in current_task_ids:
            stale_task.unlink()
    waves = Counter(row["wave_initial"] for row in rows)
    return {"tasks": len(rows), **{wave: waves[wave] for wave in ("W1", "W2", "W3", "W4")}}


class _NeutralHTML(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.stack: list[str] = []
        self.records: list[tuple[str, str, str]] = []
        self.capture: list[str] = []
        self.capture_tag = ""
        self.ordinal = 0

    def _locator(self) -> str:
        self.ordinal += 1
        return f"html#nodo={self.ordinal}/{'/'.join(self.stack)}"

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        self.stack.append(tag)
        if re.fullmatch(r"h[1-6]|th|label|legend|option", tag):
            self.capture_tag, self.capture = tag, []
        if tag in {"table", "form", "section"}:
            self.records.append((self._locator(), tag.upper(), ""))
        if tag in {"input", "select", "textarea", "button"}:
            attributes = dict(attrs)
            name, _ = safe_text(attributes.get("name") or attributes.get("id") or tag)
            input_type, _ = safe_text(attributes.get("type") or tag)
            self.records.append((self._locator(), f"CONTROL-{input_type.upper()}", name))

    def handle_endtag(self, tag: str) -> None:
        if tag == self.capture_tag:
            text, _ = safe_text("".join(self.capture))
            self.records.append((self._locator(), tag.upper(), text))
            self.capture_tag, self.capture = "", []
        if self.stack:
            self.stack.pop()

    def handle_data(self, data: str) -> None:
        if self.capture_tag:
            self.capture.append(data)


def _decode_text(payload: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return payload.decode(encoding), encoding
        except UnicodeDecodeError:
            pass
    return payload.decode("utf-8", errors="replace"), "utf-8-replace"


def _csv_objects(stream: BinaryIO, locator: str, suffix: str, *, dictionary_hint: str = "") -> list[dict[str, Any]]:
    prefix = stream.read(65536)
    _, encoding = _decode_text(prefix)
    stream.seek(0)
    wrapper = io.TextIOWrapper(stream, encoding=encoding.replace("-replace", ""), errors="replace", newline="")
    try:
        sample = wrapper.read(65536)
        wrapper.seek(0)
        delimiter = "\t" if suffix == ".tsv" else max((",", "\t", ";", "|"), key=sample.splitlines()[0].count) if sample.splitlines() else ","
        reader = csv.reader(wrapper, delimiter=delimiter)
        try:
            first_row = next(reader)
        except StopIteration:
            first_row = []
        declared_dictionary = bool(re.search(
            r"(?:^|[/_.\s-])(diccionario|dictionary|codebook|catalogo|cat[aá]logo|variables?)(?:[/_.#\s-]|$)",
            f"{locator} {dictionary_hint}", re.I,
        ))
        variable_aliases = {"variable", "var", "nombre_variable", "campo"}
        label_aliases = {"etiqueta", "label", "definicion", "definition", "reactivo", "pregunta"}
        first_normalized = {
            str(value or "").strip().casefold().replace(" ", "_"): index
            for index, value in enumerate(first_row)
        }
        dictionary_schema = bool(
            declared_dictionary
            and variable_aliases.intersection(first_normalized)
            and label_aliases.intersection(first_normalized)
        )
        structural_header = dictionary_schema
        header = first_row if structural_header else [f"COLUMNA-{index}" for index in range(1, len(first_row) + 1)]
        normalized = {str(value or "").strip().casefold().replace(" ", "_"): index for index, value in enumerate(header)}
        dictionary = bool(
            dictionary_schema
            and variable_aliases.intersection(normalized)
            and label_aliases.intersection(normalized)
        )
        variable_column = normalized[sorted(variable_aliases.intersection(normalized))[0]] if dictionary else -1
        label_columns = [normalized[name] for name in sorted(label_aliases.intersection(normalized))]
        dictionary_objects: list[dict[str, Any]] = []
        count = 0
        data_rows = reader if structural_header else itertools.chain([first_row], reader)
        for values in data_rows:
            count += 1
            if not dictionary or variable_column >= len(values) or not values[variable_column].strip():
                continue
            variable, _ = safe_text(values[variable_column])
            labels = [safe_text(values[index])[0] for index in label_columns if index < len(values) and values[index].strip()]
            dictionary_objects.append({
                "locator": f"{locator}#diccionario-fila={count}:variable={variable}",
                "parent_locator": locator,
                "type": "VARIABLE-DICCIONARIO", "name": variable,
                "label": labels[0] if labels else "NO-APLICA",
                "definition": " · ".join(labels) if labels else "metadato de diccionario",
            })
    finally:
        wrapper.detach()
    objects: list[dict[str, Any]] = [{
        "locator": locator, "type": "TABLA", "name": "tabla-raiz" if locator == "tabla=raiz" else Path(locator).name,
        "definition": f"filas={count};columnas={len(header)};delimitador={repr(delimiter)};encabezado={'SI' if structural_header else 'NO-DETERMINADO'}",
    }]
    for index, name in enumerate(header, 1):
        clean, _ = safe_text(name)
        objects.append({
            "locator": f"{locator}#columna={index}", "type": "COLUMNA",
            "parent_locator": locator,
            "name": clean, "definition": "encabezado de tabla delimitada",
        })
    objects.extend(dictionary_objects)
    return objects


def _json_schema_objects(payload: Any, locator: str = "$") -> list[dict[str, Any]]:
    objects: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()

    def walk(value: Any, path: str, parent: str | None = None) -> None:
        type_name = type(value).__name__.upper()
        key = (path, type_name)
        typed_locator = f"{path}#tipo={type_name.casefold()}"
        first = key not in seen
        if first:
            seen.add(key)
        if isinstance(value, dict):
            if first:
                objects.append({"locator": typed_locator, "parent_locator": parent, "type": "OBJETO-JSON", "name": path, "definition": "unión estructural de claves; valores no persistidos"})
            map_like = len(value) >= 2 and all(isinstance(child, (dict, list)) for child in value.values())
            for name, child in value.items():
                clean, redacted_key = safe_text(name)
                human_name_like = bool(re.fullmatch(
                    r"[A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚáéíóúÑñ'-]+(?:\s+[A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚáéíóúÑñ'-]+)+",
                    str(name), re.I,
                ))
                child_path = f"{path}[*]" if map_like or redacted_key or human_name_like else f"{path}.{clean}"
                walk(child, child_path, typed_locator)
        elif isinstance(value, list):
            if first:
                objects.append({"locator": typed_locator, "parent_locator": parent, "type": "ARREGLO-JSON", "name": path, "definition": f"elementos={len(value)};filas-no-persistidas"})
            # Unión estructural completa, nunca una fila durable por individuo.
            for child in value:
                walk(child, f"{path}[*]", typed_locator)
        elif first:
            objects.append({"locator": typed_locator, "parent_locator": parent, "type": f"VALOR-{type_name}", "name": path, "definition": "valor no persistido"})

    walk(payload, locator)
    return objects


def _xlsx_objects(path: Path) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    objects: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            locator = f"hoja={sheet.title}"
            # En modo read-only openpyxl puede devolver ``None`` para las
            # dimensiones de una hoja completamente vacía.  La ausencia de
            # celdas es una frontera estructural válida, no una excepción del
            # contenedor ni una razón para perder las hojas posteriores.
            max_row = int(sheet.max_row or 0)
            max_column = int(sheet.max_column or 0)
            declared_dictionary = bool(re.search(
                r"diccionario|dictionary|codebook|catalogo|cat[aá]logo|variables?",
                f"{path.name} {sheet.title}", re.I,
            ))
            first_two = list(
                sheet.iter_rows(min_row=1, max_row=min(max_row, 2), values_only=True)
            ) if max_row else []
            first = first_two[0] if first_two else ()
            variable_aliases = {"variable", "var", "nombre_variable", "campo"}
            label_aliases = {"etiqueta", "label", "definicion", "definition", "reactivo", "pregunta"}
            first_normalized = {
                str(value or "").strip().casefold().replace(" ", "_"): index
                for index, value in enumerate(first) if str(value or "").strip()
            }
            dictionary_schema = bool(
                declared_dictionary
                and variable_aliases.intersection(first_normalized)
                and label_aliases.intersection(first_normalized)
            )
            declared_table = bool(getattr(sheet, "tables", {}))
            structural_header = dictionary_schema or declared_table
            objects.append({
                "locator": locator, "type": "HOJA-XLSX", "name": sheet.title,
                "definition": f"filas={max_row};columnas={max_column};encabezado={'SI' if structural_header else 'NO-DETERMINADO'}", "sheet": sheet.title,
            })
            for index, value in enumerate(first, 1):
                name, _ = safe_text(value if structural_header else f"COLUMNA-{index}")
                objects.append({
                    "locator": f"{locator}#columna={index}", "type": "COLUMNA",
                    "parent_locator": locator,
                    "name": name,
                    "definition": "encabezado estructural" if structural_header else "encabezado no determinado; valor no persistido",
                    "sheet": sheet.title,
                })
            # Solo las hojas que se autodeclaran diccionario se recorren como
            # metadatos fila a fila. Las hojas de observaciones nunca exportan
            # sus celdas. Se exige columna variable + etiqueta/definición.
            header_row = 0
            header_map: dict[str, int] = {}
            category_aliases = {"categoria", "categorias", "category", "categories", "value_label", "value_labels"}
            candidate_rows = (
                sheet.iter_rows(min_row=1, max_row=min(max_row, 20), values_only=True)
                if max_row else ()
            )
            for candidate_number, candidate in enumerate(candidate_rows, 1):
                normalized = {
                    str(value or "").strip().casefold().replace(" ", "_"): index
                    for index, value in enumerate(candidate)
                    if str(value or "").strip()
                }
                if dictionary_schema and variable_aliases.intersection(normalized) and label_aliases.intersection(normalized):
                    header_row, header_map = candidate_number, normalized
                    break
            if header_row:
                variable_column = header_map[sorted(variable_aliases.intersection(header_map))[0]]
                label_columns = [header_map[name] for name in sorted(label_aliases.intersection(header_map))]
                category_columns = [header_map[name] for name in sorted(category_aliases.intersection(header_map))]
                for row_number, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
                    if variable_column >= len(values) or not str(values[variable_column] or "").strip():
                        continue
                    variable, _ = safe_text(values[variable_column])
                    labels = [safe_text(values[index])[0] for index in label_columns if index < len(values) and str(values[index] or "").strip()]
                    categories = [safe_text(values[index])[0] for index in category_columns if index < len(values) and str(values[index] or "").strip()]
                    objects.append({
                        "locator": f"{locator}#diccionario-fila={row_number}:variable={variable}",
                        "parent_locator": locator,
                        "type": "VARIABLE-DICCIONARIO-XLSX", "name": variable,
                        "label": labels[0] if labels else "NO-APLICA",
                        "definition": " · ".join(labels) if labels else "metadato de diccionario",
                        "categories": categories, "sheet": sheet.title,
                    })
            for table_name in sorted(getattr(sheet, "tables", {})):
                objects.append({
                    "locator": f"{locator}#tabla={table_name}", "type": "TABLA-XLSX",
                    "parent_locator": locator,
                    "name": table_name, "definition": "tabla declarada en workbook", "sheet": sheet.title,
                })
    finally:
        workbook.close()
    return objects


def _xls_objects(path: Path) -> list[dict[str, Any]]:
    objects: list[dict[str, Any]] = []
    with olefile.OleFileIO(path) as document:
        stream = "Workbook" if document.exists("Workbook") else "Book" if document.exists("Book") else ""
        names: list[str] = []
        if stream:
            payload = document.openstream(stream).read()
            cursor = 0
            while cursor + 4 <= len(payload):
                record_type, length = struct.unpack("<HH", payload[cursor:cursor + 4])
                data = payload[cursor + 4:cursor + 4 + length]
                if record_type == 0x0085 and len(data) >= 8:
                    count, flags = data[6], data[7]
                    raw = data[8:8 + count * (2 if flags & 1 else 1)]
                    names.append(raw.decode("utf-16le" if flags & 1 else "latin-1", errors="replace"))
                cursor += 4 + length
        for sheet_ordinal, name in enumerate(names, 1):
            clean, _ = safe_text(name)
            name_sha256 = hashlib.sha256(name.encode("utf-8", errors="replace")).hexdigest()
            objects.append({
                "locator": f"hoja={sheet_ordinal}:nombre_sha256={name_sha256}", "type": "HOJA-XLS",
                "name": clean, "definition": "EXCEPCION-ESPECIFICA:celdas-BIFF-no-decodificadas",
                "state": "EXCEPCION-ESPECIFICA", "sheet": clean,
            })
    return objects


def _pdf_objects(path: Path) -> list[dict[str, Any]]:
    info = subprocess.run(["pdfinfo", str(path)], capture_output=True, text=True, timeout=60)
    if info.returncode:
        raise ValueError(f"PDFINFO:{info.stderr[:300]}")
    metadata = dict(line.split(":", 1) for line in info.stdout.splitlines() if ":" in line)
    pages = int(metadata.get("Pages", "0").strip() or 0)
    encrypted = metadata.get("Encrypted", "no").strip().casefold()
    if encrypted.startswith("yes"):
        raise PermissionError("PDF_CIFRADO")
    objects: list[dict[str, Any]] = []
    for page in range(1, pages + 1):
        result = subprocess.run(
            ["pdftotext", "-f", str(page), "-l", str(page), str(path), "-"],
            capture_output=True, timeout=120,
        )
        if result.returncode != 0:
            objects.append({
                "locator": f"pagina={page}", "type": "PAGINA-PDF",
                "name": f"Página {page}",
                "definition": f"EXCEPCION-ESPECIFICA:PDFTOTEXT;stderr_sha256={hashlib.sha256(result.stderr).hexdigest()}",
                "page": page, "state": "EXCEPCION-ESPECIFICA",
            })
            continue
        text, _ = _decode_text(result.stdout)
        lines = [" ".join(line.split()) for line in text.splitlines() if line.strip()]
        objects.append({
            "locator": f"pagina={page}", "type": "PAGINA-PDF", "name": f"Página {page}",
            "definition": f"lineas_texto={len(lines)};texto_extraible={'SI' if lines else 'NO'}", "page": page,
        })
        for ordinal, line in enumerate(lines):
            if not (line.endswith("?") or (len(line) <= 180 and (line.isupper() or re.match(r"^(?:\d+\.?|[IVX]+\.)\s", line)))):
                continue
            clean, _ = safe_text(line)
            objects.append({
                "locator": f"pagina={page}#linea-estructural={ordinal}",
                "parent_locator": f"pagina={page}",
                "type": "REACTIVO-PDF" if line.endswith("?") else "SECCION-PDF",
                "name": clean if not line.endswith("?") else "reactivo",
                "question": clean if line.endswith("?") else "NO-APLICA",
                "definition": "texto estructural extraído", "page": page,
            })
    return objects


def _docx_objects(path: Path) -> list[dict[str, Any]]:
    import xml.etree.ElementTree as ET
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    objects: list[dict[str, Any]] = []
    with zipfile.ZipFile(path) as archive:
        story_parts = sorted(
            name for name in archive.namelist()
            if re.fullmatch(r"word/(?:document|header\d+|footer\d+|footnotes|endnotes|comments)\.xml", name)
        )
        for part_name in story_parts:
            root = ET.fromstring(archive.read(part_name))
            part_locator = f"parte={part_name}"
            objects.append({
                "locator": part_locator, "type": "PARTE-DOCX", "name": part_name,
                "definition": "parte narrativa enumerada; texto ordinario no persistido",
            })
            table_paragraphs = {
                paragraph
                for cell in root.findall(".//w:tc", namespace)
                for paragraph in cell.findall(".//w:p", namespace)
            }
            for index, paragraph in enumerate(root.findall(".//w:p", namespace), 1):
                text = "".join(node.text or "" for node in paragraph.findall(".//w:t", namespace))
                style_node = paragraph.find("./w:pPr/w:pStyle", namespace)
                style = style_node.get(f"{{{namespace['w']}}}val", "") if style_node is not None else ""
                stripped = text.strip()
                inside_table = paragraph in table_paragraphs
                is_question = stripped.endswith("?") and not inside_table
                is_heading = bool(re.search(r"heading|t[ií]tulo", style, re.I)) and not inside_table
                clean, _ = safe_text(text)
                objects.append({
                    "locator": f"{part_locator}#parrafo={index}", "parent_locator": part_locator,
                    "type": "PARRAFO-DOCX", "name": clean if is_heading else "párrafo",
                    "definition": f"caracteres={len(text)};estilo={safe_text(style)[0]};texto-ordinario-no-persistido",
                    "question": clean if is_question else "NO-APLICA",
                })
            for index, table in enumerate(root.findall(".//w:tbl", namespace), 1):
                rows = table.findall("./w:tr", namespace)
                max_cells = max((len(row.findall("./w:tc", namespace)) for row in rows), default=0)
                objects.append({
                    "locator": f"{part_locator}#tabla={index}", "parent_locator": part_locator,
                    "type": "TABLA-DOCX", "name": f"Tabla {index}",
                    "definition": f"filas={len(rows)};columnas_max={max_cells};celdas-no-persistidas",
                    "table": str(index),
                })
            for index, _section in enumerate(root.findall(".//w:sectPr", namespace), 1):
                objects.append({
                    "locator": f"{part_locator}#seccion={index}", "parent_locator": part_locator,
                    "type": "SECCION-DOCX", "name": f"Sección {index}",
                    "definition": "límite estructural de sección",
                })
    return objects


def _sav_objects(path: Path) -> list[dict[str, Any]]:
    """Lee el diccionario SAV clásico sin tocar el bloque de casos."""
    file_size = path.stat().st_size
    with path.open("rb") as stream:
        header = stream.read(176)
        if len(header) < 176 or header[:4] != b"$FL2":
            raise ValueError("SAV_CABECERA_INVALIDA")
        layout_raw = header[64:68]
        if int.from_bytes(layout_raw, "little", signed=True) in {2, 3}:
            endian = "<"
        elif int.from_bytes(layout_raw, "big", signed=True) in {2, 3}:
            endian = ">"
        else:
            raise ValueError("SAV_LAYOUT_NO_SOPORTADO")

        def read_exact(size: int) -> bytes:
            value = stream.read(size)
            if len(value) != size:
                raise ValueError("SAV_DICCIONARIO_TRUNCADO")
            return value

        def read_int() -> int:
            return struct.unpack(endian + "i", read_exact(4))[0]

        variables: list[dict[str, Any]] = []
        collections: list[dict[str, Any]] = []
        extensions: list[dict[str, Any]] = []
        while stream.tell() + 4 <= file_size:
            record_type = read_int()
            if record_type == 2:
                variable_type, has_label, missing_count, print_format, write_format = struct.unpack(
                    endian + "iiiii", read_exact(20)
                )
                name = read_exact(8).rstrip(b" \x00").decode("cp1252", errors="replace")
                label = ""
                if has_label:
                    label_size = read_int()
                    if label_size < 0 or label_size > 1_048_576:
                        raise ValueError("SAV_ETIQUETA_TAMANO_INVALIDO")
                    label = read_exact(label_size).decode("cp1252", errors="replace")
                    read_exact((-label_size) % 4)
                missing_slots = abs(missing_count)
                if missing_slots > 3:
                    raise ValueError("SAV_MISSING_VALUES_INVALIDOS")
                read_exact(missing_slots * 8)
                if variable_type == -1:
                    continue
                clean_name, _ = safe_text(name)
                clean_label, _ = safe_text(label)
                variable_ordinal = len(variables) + 1
                name_sha256 = hashlib.sha256(name.encode("utf-8", errors="replace")).hexdigest()
                variables.append({
                    "locator": f"tabla=sav#variable={variable_ordinal}:nombre_sha256={name_sha256}",
                    "parent_locator": "tabla=sav",
                    "type": "VARIABLE-SAV", "name": clean_name, "label": clean_label,
                    "definition": (
                        f"tipo={'NUMERICO' if variable_type == 0 else f'STRING-{variable_type}'};"
                        f"formato_impresion={print_format:#x};formato_escritura={write_format:#x}"
                    ),
                })
            elif record_type == 3:
                label_count = read_int()
                if label_count < 0 or label_count > 1_000_000:
                    raise ValueError("SAV_VALUE_LABEL_COUNT_INVALIDO")
                labels: list[str] = []
                for _ in range(label_count):
                    category_code = read_exact(8)
                    label_size = read_exact(1)[0]
                    label = read_exact(label_size).decode("cp1252", errors="replace")
                    read_exact((-(label_size + 1)) % 8)
                    labels.append(f"codigo_hex={category_code.hex()};label={safe_text(label)[0]}")
                if read_int() != 4:
                    raise ValueError("SAV_VALUE_LABEL_SIN_VARIABLES")
                variable_count = read_int()
                if variable_count < 0 or variable_count > len(variables) + 4096:
                    raise ValueError("SAV_VALUE_LABEL_VARIABLE_COUNT_INVALIDO")
                variable_indexes = [read_int() for _ in range(variable_count)]
                associated_names = [
                    str(variables[index - 1]["name"])
                    for index in variable_indexes if 1 <= index <= len(variables)
                ]
                collections.append({
                    "locator": f"value-label-collection={len(collections) + 1}",
                    "parent_locator": "tabla=sav",
                    "type": "VALUE-LABEL-COLLECTION-SAV",
                    "name": f"Colección {len(collections) + 1}",
                    "definition": (
                        f"variables={','.join(associated_names) if associated_names else 'NO-DETERMINADO'};"
                        f"categorias={len(labels)}"
                    ),
                    "value_labels": labels,
                })
            elif record_type == 6:
                line_count = read_int()
                if line_count < 0 or line_count > 1_000_000:
                    raise ValueError("SAV_DOCUMENT_COUNT_INVALIDO")
                read_exact(line_count * 80)
            elif record_type == 7:
                subtype, element_size, element_count = struct.unpack(endian + "iii", read_exact(12))
                total = element_size * element_count
                if element_size < 0 or element_count < 0 or total > file_size - stream.tell():
                    raise ValueError("SAV_EXTENSION_TAMANO_INVALIDO")
                read_exact(total)
                extensions.append({
                    "locator": f"tabla=sav#extension={len(extensions) + 1}:subtipo={subtype}",
                    "parent_locator": "tabla=sav",
                    "type": "EXTENSION-DICCIONARIO-SAV",
                    "name": f"Subtipo {subtype}",
                    "definition": f"EXCEPCION-ESPECIFICA:subtipo={subtype};tamano={element_size};elementos={element_count}",
                    "state": "EXCEPCION-ESPECIFICA",
                })
            elif record_type == 999:
                read_exact(4)
                break
            else:
                raise ValueError(f"SAV_RECORD_TYPE_NO_SOPORTADO:{record_type}")
    if not variables:
        raise ValueError("SAV_SIN_DICCIONARIO_VARIABLES")
    return [{
        "locator": "tabla=sav", "type": "TABLA-SAV", "name": "tabla-sav",
        "definition": f"variables={len(variables)};observaciones-no-leidas",
    }, *variables, *collections, *extensions]


def _dta_objects(path: Path) -> list[dict[str, Any]]:
    from pandas.io.stata import StataReader
    objects: list[dict[str, Any]] = []
    with StataReader(path, convert_categoricals=False) as reader:
        labels = reader.variable_labels()
        value_tables = reader.value_labels()
        variables = list(reader._varlist)
        formats = dict(zip(variables, reader._fmtlist))
        label_tables = dict(zip(variables, reader._lbllist))
        objects.append({
            "locator": "tabla=stata", "type": "TABLA-DTA", "name": "tabla-stata",
            "definition": f"observaciones={reader._nobs};variables={reader._nvar};observaciones-no-persistidas",
        })
        for variable_ordinal, name in enumerate(variables, 1):
            clean_name, _ = safe_text(name)
            clean_label, _ = safe_text(labels.get(name, ""))
            mapping = value_tables.get(label_tables.get(name, ""), {})
            label_names = sorted({safe_text(value)[0] for value in mapping.values()})
            name_sha256 = hashlib.sha256(name.encode("utf-8", errors="replace")).hexdigest()
            objects.append({
                "locator": f"tabla=stata#variable={variable_ordinal}:nombre_sha256={name_sha256}", "type": "VARIABLE-DTA",
                "parent_locator": "tabla=stata",
                "name": clean_name, "label": clean_label,
                "definition": f"formato={formats.get(name, 'NO-DETERMINADO')}",
                "value_labels": label_names,
            })
    return objects


def _nest_objects(raw_objects: list[dict[str, Any]], member_locator: str) -> list[dict[str, Any]]:
    locator_map = {
        str(raw.get("locator", "objeto")): f"{member_locator}!/contenido/{raw.get('locator', 'objeto')}"
        for raw in raw_objects
    }
    nested: list[dict[str, Any]] = []
    for raw in raw_objects:
        projected = dict(raw)
        original = str(raw.get("locator", "objeto"))
        original_parent = raw.get("parent_locator")
        projected["locator"] = locator_map[original]
        projected["parent_locator"] = locator_map.get(str(original_parent), member_locator)
        nested.append(projected)
    return nested


def _zip_objects(source: Path | BinaryIO, *, depth: int = 0, parent: str = "zip") -> list[dict[str, Any]]:
    if depth > 4:
        return [{
            "locator": f"{parent}!/contenido-profundidad", "parent_locator": parent,
            "type": "ZIP-ANIDADO", "name": "zip-anidado",
            "definition": "EXCEPCION-ESPECIFICA:profundidad>4", "state": "EXCEPCION-ESPECIFICA",
        }]
    objects: list[dict[str, Any]] = []
    with zipfile.ZipFile(source) as archive:
        for ordinal, info in enumerate(archive.infolist(), 1):
            safe_filename, _ = safe_text(info.filename)
            locator = f"{parent}!/miembro={ordinal}:{safe_filename}"
            normalized = PurePosixPath(info.filename.replace("\\", "/"))
            slip = normalized.is_absolute() or ".." in normalized.parts
            temp_limit = min(50 * 1024**3, shutil.disk_usage(tempfile.gettempdir()).free // 10)
            nested_over_memory_guard = Path(info.filename).suffix.casefold() == ".zip" and info.file_size > 512 * 1024**2
            temp_over_budget = info.file_size > temp_limit
            state = "EXCEPCION-ESPECIFICA" if (
                slip or bool(info.flag_bits & 1) or info.file_size > 8 * 1024**3
                or nested_over_memory_guard or temp_over_budget
            ) else "E2-COMPLETO"
            definition = (
                f"bytes={info.file_size};comprimidos={info.compress_size};crc={info.CRC};"
                f"zip_slip={'SI' if slip else 'NO'};cifrado={'SI' if info.flag_bits & 1 else 'NO'};"
                f"guard_memoria_anidado={'SI' if nested_over_memory_guard else 'NO'};"
                f"guard_temp={'SI' if temp_over_budget else 'NO'}"
            )
            objects.append({
                "locator": locator, "type": "MIEMBRO-ZIP", "name": info.filename,
                "parent_locator": parent if parent != "zip" else None,
                "definition": definition, "state": state,
            })
            if info.is_dir() or state != "E2-COMPLETO" or info.file_size > 8 * 1024**3:
                continue
            suffix = Path(info.filename).suffix.casefold()
            try:
                with archive.open(info) as probe:
                    member_magic = magic_label(probe.read(32))
            except (OSError, RuntimeError, NotImplementedError, zipfile.BadZipFile) as exc:
                objects.append({
                    "locator": f"{locator}!/contenido", "parent_locator": locator,
                    "type": "EXCEPCION-MIEMBRO-ZIP", "name": safe_filename,
                    "definition": f"EXCEPCION-ESPECIFICA:{type(exc).__name__};detalle_sha256={hashlib.sha256(str(exc).encode()).hexdigest()}",
                    "state": "EXCEPCION-ESPECIFICA",
                })
                continue
            if member_magic == "ZIP" and suffix not in {".zip", ".xlsx", ".docx"}:
                suffix = ".zip"
            elif member_magic == "PDF" and suffix != ".pdf":
                suffix = ".pdf"
            elif member_magic == "OLE" and suffix != ".xls":
                suffix = ".xls"
            if suffix in {".csv", ".tsv"}:
                try:
                    with archive.open(info) as stream:
                        children = _csv_objects(stream, f"{locator}#/contenido-tabla", suffix)
                except (OSError, RuntimeError, NotImplementedError, zipfile.BadZipFile) as exc:
                    objects.append({
                        "locator": f"{locator}!/contenido", "parent_locator": locator,
                        "type": "EXCEPCION-MIEMBRO-ZIP", "name": safe_filename,
                        "definition": f"EXCEPCION-ESPECIFICA:{type(exc).__name__};detalle_sha256={hashlib.sha256(str(exc).encode()).hexdigest()}",
                        "state": "EXCEPCION-ESPECIFICA",
                    })
                    continue
                for child in children:
                    child.setdefault("parent_locator", locator)
                objects.extend(children)
            elif suffix == ".zip" and info.file_size <= 512 * 1024**2:
                try:
                    with archive.open(info) as stream:
                        nested = io.BytesIO(stream.read())
                    objects.extend(_zip_objects(nested, depth=depth + 1, parent=locator))
                except (OSError, RuntimeError, NotImplementedError, zipfile.BadZipFile) as exc:
                    objects.append({
                        "locator": f"{locator}!/contenido", "parent_locator": locator,
                        "type": "EXCEPCION-MIEMBRO-ZIP", "name": safe_filename,
                        "definition": f"EXCEPCION-ESPECIFICA:{type(exc).__name__};detalle_sha256={hashlib.sha256(str(exc).encode()).hexdigest()}",
                        "state": "EXCEPCION-ESPECIFICA",
                    })
            elif suffix in {
                ".pdf", ".xls", ".xlsx", ".dta", ".sav", ".docx",
                ".json", ".xml", ".html", ".htm", ".txt", ".php", ".md", ".log", ".dat",
            } or not suffix:
                temporary_path: Path | None = None
                try:
                    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temporary:
                        temporary_path = Path(temporary.name)
                        with archive.open(info) as stream:
                            shutil.copyfileobj(stream, temporary, length=1024 * 1024)
                        temporary.flush()
                        os.fsync(temporary.fileno())
                    child_objects, _, _ = inspect_e2(temporary_path)
                    for child in child_objects:
                        child["locator"] = str(child.get("locator", "objeto")).replace(temporary_path.name, safe_filename)
                        if child.get("parent_locator"):
                            child["parent_locator"] = str(child["parent_locator"]).replace(temporary_path.name, safe_filename)
                        if child.get("name") == temporary_path.name:
                            child["name"] = safe_filename
                    objects.extend(_nest_objects(child_objects, locator))
                except (OSError, ValueError, EOFError, zipfile.BadZipFile, subprocess.SubprocessError, NotImplementedError) as exc:
                    objects.append({
                        "locator": f"{locator}!/contenido",
                        "parent_locator": locator,
                        "type": "EXCEPCION-MIEMBRO-ZIP",
                        "name": safe_filename,
                        "definition": f"EXCEPCION-ESPECIFICA:{type(exc).__name__};detalle_sha256={hashlib.sha256(str(exc).encode()).hexdigest()}",
                        "state": "EXCEPCION-ESPECIFICA",
                    })
                finally:
                    if temporary_path is not None:
                        temporary_path.unlink(missing_ok=True)
            else:
                # Lee hasta EOF para verificar CRC, sin conservar los bytes.
                with archive.open(info) as stream:
                    for _ in iter(lambda: stream.read(1024 * 1024), b""):
                        pass
                objects.append({
                    "locator": f"{locator}!/contenido",
                    "parent_locator": locator,
                    "type": "FORMATO-NO-SOPORTADO",
                    "name": safe_filename,
                    "definition": f"EXCEPCION-ESPECIFICA:extension={suffix or 'sin-extension'}",
                    "state": "EXCEPCION-ESPECIFICA",
                })
    return objects


def inspect_e2(path: Path) -> tuple[list[dict[str, Any]], str, str]:
    suffix = path.suffix.casefold()
    with path.open("rb") as handle:
        observed_magic = magic_label(handle.read(32))
    if observed_magic == "ZIP" and suffix not in {".zip", ".xlsx", ".docx"}:
        suffix = ".zip"
    elif observed_magic == "PDF" and suffix != ".pdf":
        suffix = ".pdf"
    elif observed_magic == "OLE" and suffix != ".xls":
        suffix = ".xls"
    parser = "barrido2-stdlib-1"
    if suffix == ".zip":
        return _zip_objects(path), parser + "+zipfile", "contenedor y miembros completos; anidados hasta profundidad 4"
    if suffix == ".pdf":
        return _pdf_objects(path), parser + "+poppler", "todas las páginas; texto no extraíble se declara por página"
    if suffix == ".xlsx":
        return _xlsx_objects(path), parser + "+openpyxl", "todas las hojas/tablas y encabezados; celdas de observación no persistidas"
    if suffix == ".xls":
        return _xls_objects(path), parser + "+olefile", "todas las hojas BIFF; celdas como excepción específica"
    if suffix in {".csv", ".tsv"}:
        with path.open("rb") as handle:
            return _csv_objects(handle, "tabla=raiz", suffix, dictionary_hint=path.name), parser + "+csv", "encabezado/esquema y conteo completo; filas no persistidas"
    if suffix == ".dta":
        return _dta_objects(path), parser + "+pandas-stata", "metadatos completos; observaciones no persistidas"
    if suffix == ".sav":
        return _sav_objects(path), parser + "+sav-dictionary", "diccionario base, variables, formatos y value labels; extensiones no decodificadas quedan como excepción; observaciones no leídas"
    if suffix == ".docx":
        return _docx_objects(path), parser + "+xml", "todos los párrafos y tablas; texto depurado"
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        return _json_schema_objects(payload), parser + "+json", "estructura recursiva completa; valores no persistidos"
    if suffix == ".xml":
        import xml.etree.ElementTree as ET
        root = ET.parse(path).getroot()
        objects = []
        parent_by_element = {child: parent for parent in root.iter() for child in parent}
        locator_by_element: dict[Any, str] = {}
        for index, element in enumerate(root.iter()):
            tag, _ = safe_text(element.tag)
            attribute_names = [safe_text(name)[0] for name in sorted(element.attrib)]
            locator = f"elemento={index}"
            parent_element = parent_by_element.get(element)
            objects.append({
                "locator": locator,
                "parent_locator": locator_by_element.get(parent_element),
                "type": "ELEMENTO-XML", "name": tag,
                "definition": f"atributos={','.join(attribute_names) if attribute_names else 'NINGUNO'};texto-no-persistido",
            })
            locator_by_element[element] = locator
        return objects, parser + "+xml", "todos los elementos; valores de texto/atributos no persistidos"
    if suffix in {".html", ".htm"}:
        parser_html = _NeutralHTML(); parser_html.feed(path.read_text(encoding="utf-8", errors="replace"))
        document_locator = "documento=html"
        objects = [{"locator": document_locator, "type": "DOCUMENTO-HTML", "name": "documento-html", "definition": "raíz estructural HTML"}]
        objects.extend({"locator": loc, "parent_locator": document_locator, "type": typ, "name": text or typ, "definition": "estructura HTML"} for loc, typ, text in parser_html.records)
        return objects, parser + "+html", "documento completo; enlaces no seguidos y scripts no ejecutados"
    if suffix in TEXT_SUFFIXES or not suffix:
        text, encoding = _decode_text(path.read_bytes())
        lines = text.splitlines()
        objects = [{"locator": "documento=texto", "type": "TEXTO", "name": "documento-texto", "definition": f"lineas={len(lines)};encoding={encoding}"}]
        for index, line in enumerate(lines):
            stripped = line.strip()
            if stripped.startswith("#") or re.match(r"^(?:\d+\.?|[IVX]+\.)\s", stripped):
                clean, _ = safe_text(stripped)
                objects.append({"locator": f"linea-estructural={index}", "parent_locator": "documento=texto", "type": "SECCION-TEXTO", "name": clean, "definition": "encabezado estructural"})
        return objects, parser, "clasificación segura y estructura; líneas ordinarias no persistidas"
    return [{"locator": "contenido=raiz", "type": "FORMATO-NO-SOPORTADO", "name": "contenido-raiz", "definition": f"EXCEPCION-ESPECIFICA:extension={suffix or 'sin-extension'}", "state": "EXCEPCION-ESPECIFICA"}], parser, "formato identificado; contenido no decodificado"


def _e2_record(task: dict[str, Any], raw: dict[str, Any], parser: str, boundary: str) -> dict[str, Any]:
    def contains_privacy_redaction(value: Any) -> bool:
        if value == "[REDACTADO-PRIVACIDAD]":
            return True
        if isinstance(value, dict):
            return any(contains_privacy_redaction(item) for item in value.values())
        if isinstance(value, (list, tuple, set)):
            return any(contains_privacy_redaction(item) for item in value)
        return False

    raw_locator = str(raw.get("locator", "objeto"))
    locator, red_locator = safe_text(raw_locator)
    name, red_name = safe_text(raw.get("name", ""))
    label, red_label = safe_text(raw.get("label", ""))
    question, red_question = safe_text(raw.get("question", ""))
    definition, red_definition = safe_text(raw.get("definition", ""))
    categories_with_flags = [safe_text(value) for value in raw.get("categories", [])]
    labels_with_flags = [safe_text(value) for value in raw.get("value_labels", [])]
    unit, red_unit = safe_text(raw.get("unit", ""))
    period, red_period = safe_text(raw.get("period", ""))
    population, red_population = safe_text(raw.get("population", ""))
    sheet, red_sheet = safe_text(raw.get("sheet", ""))
    table, red_table = safe_text(raw.get("table", ""))
    route, red_route = safe_text(task["ruta_relativa"])
    parent_locator = raw.get("parent_locator")
    parent_id = (
        str(raw.get("source_parent_id"))
        if raw.get("source_parent_id")
        else logical_object_id(task["sha256"], str(parent_locator)) if parent_locator
        else "NO-APLICA"
    )
    object_id = str(raw.get("source_object_id")) if raw.get("source_object_id") else logical_object_id(task["sha256"], raw_locator)
    redacted = any((
        red_locator, red_name, red_label, red_question, red_definition,
        red_unit, red_period, red_population, red_sheet, red_table, red_route,
        *(flag for _, flag in categories_with_flags),
        *(flag for _, flag in labels_with_flags),
    ))
    base = {
        "schema_version": "BARRIDO2-E2-1.0",
        "payload_id": task["payload_id"], "representacion_id": task["representacion_id"],
        "sha256": task["sha256"], "objeto_logico_id": object_id,
        "root_id": task["root_id"], "ruta_relativa": route,
        "format": task["formato"], "depth": int(raw.get("depth", 1 if parent_locator else 0)),
        "localizador": locator, "objeto_tipo": raw.get("type", "OBJETO"),
        "nombre": name, "etiqueta": label, "texto_reactivo": question,
        "definicion": definition,
        "categorias": [value for value, _ in categories_with_flags],
        "value_labels": [value for value, _ in labels_with_flags],
        "unidad": unit, "periodo": period, "poblacion": population,
        "pagina": raw.get("page", "NO-APLICA"), "hoja": sheet,
        "tabla": table,
        "objeto_padre_id": parent_id,
        "relacion_estructural": str(raw.get("source_relation")) if raw.get("source_relation") else (
            "CONTENIDO-EN-OBJETO" if parent_locator else "CONTENIDO-EN-REPRESENTACION"
        ),
        "frontera_inspeccion": safe_text(boundary)[0], "parser": parser,
        "parser_version": MATERIAL_BUILD_VERSION, "estado": raw.get("state", "E2-COMPLETO"),
        "privacidad": (
            "[REDACTADO-PRIVACIDAD]"
            if raw.get("source_privacy") == "[REDACTADO-PRIVACIDAD]"
            or contains_privacy_redaction(raw)
            or redacted
            else "DEPURADO"
        ),
        "fecha": date.today().isoformat(),
    }
    record_seed = canonical_sha(base)
    base["record_id"] = "E2R-" + record_seed
    base["record_sha256"] = canonical_sha(base)
    return base


REPORT_FIELDS = [
    "reporte_id", "record_id", "record_sha256", "batch_id", "batch_sha256",
    "payload_id", "representacion_id", "sha256", "objeto_logico_id",
    "grado_inspeccion", "afirmacion_tipo", "objeto_tipo", "localizador",
    "descripcion_neutral", "frontera_inspeccion", "estado", "privacidad", "fecha",
]


def _audit_e2_files_streaming(
    index_path: Path,
    report_path: Path,
    summary: dict[str, Any],
    task: dict[str, Any],
    schema_validator: Draft202012Validator,
) -> dict[str, Any]:
    """Audita un expediente E2 en O(n), sin cargarlo completo en RAM."""
    errors: list[str] = []

    def fail(code: str) -> None:
        # Un artefacto corrupto no debe poder agotar memoria sólo generando
        # millones de diagnósticos repetidos. El primer centenar conserva
        # evidencia suficiente y el gate permanece cerrado.
        if len(errors) < 200:
            errors.append(code)
        elif len(errors) == 200:
            errors.append("E2_ERRORES_TRUNCADOS")

    object_ids: set[str] = set()
    parent_ids: set[str] = set()
    record_ids: set[str] = set()
    record_hashes: list[str] = []
    durable_expected: dict[str, str] = {}
    records_count = e2_count = exception_count = 0
    summary_batch = str(summary.get("batch_sha256", ""))
    try:
        with index_path.open(encoding="utf-8") as handle:
            for line_number, line in enumerate(handle, 1):
                if not line.strip():
                    continue
                try:
                    record = json.loads(line)
                except (json.JSONDecodeError, TypeError):
                    fail(f"E2_INDEX_JSON_INVALIDO:{line_number}")
                    continue
                records_count += 1
                record_id = str(record.get("record_id", ""))
                if list(schema_validator.iter_errors(record)):
                    fail(f"E2_SCHEMA_INVALIDO:{record_id or line_number}")
                seed = {
                    key: value for key, value in record.items()
                    if key not in {"record_id", "record_sha256", "batch_id", "batch_sha256"}
                }
                expected_id = "E2R-" + canonical_sha(seed)
                expected_hash = canonical_sha({**seed, "record_id": expected_id})
                if record_id != expected_id or record.get("record_sha256") != expected_hash:
                    fail(f"E2_RECORD_HASH_INVALIDO:{record_id or line_number}")
                if record_id in record_ids:
                    fail(f"E2_RECORD_ID_DUPLICADO:{record_id}")
                record_ids.add(record_id)
                record_hashes.append(str(record.get("record_sha256", "")))
                if (
                    record.get("representacion_id") != task.get("representacion_id")
                    or record.get("payload_id") != task.get("payload_id")
                    or record.get("sha256") != task.get("sha256")
                ):
                    fail(f"E2_JOIN_INVALIDO:{record_id}")
                if (
                    record.get("parser") != summary.get("parser")
                    or record.get("parser_version") != MATERIAL_BUILD_VERSION
                ):
                    fail(f"E2_PARSER_BUILD_INVALIDO:{record_id}")
                if (
                    record.get("batch_id") != "E2B-" + summary_batch
                    or record.get("batch_sha256") != summary_batch
                ):
                    fail(f"E2_BATCH_HASH_INVALIDO:{record_id}")
                if any(
                    key in record
                    for key in ("necesidad_id", "relacion_id", "valor_individual", "fila_microdato", "pii")
                ):
                    fail(f"E2_CONTAMINACION_SEMANTICA_PRIVACIDAD:{record_id}")
                privacy_values = [
                    # payload_id es un slug administrativo validado por
                    # ``valid_payload_id``. No es prosa extraída: aplicarle
                    # heurísticas de nombres/teléfonos produce falsos PII en
                    # ids legítimos con guiones bajos o timestamps.
                    record.get("ruta_relativa", ""), record.get("localizador", ""), record.get("nombre", ""),
                    record.get("etiqueta", ""), record.get("texto_reactivo", ""),
                    record.get("definicion", ""), *(record.get("categorias", []) or []),
                    *(record.get("value_labels", []) or []), record.get("unidad", ""),
                    record.get("periodo", ""), record.get("poblacion", ""),
                    record.get("hoja", ""), record.get("tabla", ""),
                ]
                if any(
                    pattern.search(str(value))
                    for pattern in PII_PATTERNS for value in privacy_values
                ):
                    fail(f"E2_PII_NO_REDACTADA:{record_id}")
                if (
                    "[REDACTADO-PRIVACIDAD]" in privacy_values
                    and record.get("privacidad") != "[REDACTADO-PRIVACIDAD]"
                ):
                    fail(f"E2_MARCA_PRIVACIDAD_INCONSISTENTE:{record_id}")
                object_ids.add(str(record.get("objeto_logico_id", "")))
                parent_id = str(record.get("objeto_padre_id", ""))
                if parent_id != "NO-APLICA":
                    parent_ids.add(parent_id)
                if record.get("estado") == "E2-COMPLETO":
                    e2_count += 1
                if record.get("estado") == "EXCEPCION-ESPECIFICA":
                    exception_count += 1
                try:
                    durable = _durable_row(record)
                    durable_hash = canonical_sha(durable)
                except (KeyError, TypeError, ValueError):
                    fail(f"REPORTE_DURABLE_PROYECCION_INVALIDA:{record_id}")
                else:
                    if record_id in durable_expected:
                        fail(f"REPORTE_DURABLE_RECORD_ID_DUPLICADO:{record_id}")
                    durable_expected[record_id] = durable_hash
    except (OSError, UnicodeError):
        fail("E2_INDEX_ILEGIBLE")

    if not records_count:
        fail("E2_SIN_OBJETOS")
    if not parent_ids.issubset(object_ids):
        for missing in sorted(parent_ids - object_ids)[:20]:
            fail(f"E2_PADRE_NO_DEREFERENCIABLE:{missing}")
    expected_batch = canonical_sha(sorted(record_hashes))
    if summary_batch != expected_batch:
        fail("E2_SUMMARY_BATCH_HASH_INVALIDO")
    if (
        summary.get("objetos_e1") != records_count
        or summary.get("objetos_e2") != e2_count
        or summary.get("excepciones") != exception_count
    ):
        fail("E2_CONTEOS_NO_RECONCILIAN")

    seen_report_ids: set[str] = set()
    durable_count = 0
    try:
        with report_path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            if reader.fieldnames != REPORT_FIELDS:
                fail("REPORTE_DURABLE_CABECERA_INVALIDA")
            for durable_row in reader:
                durable_count += 1
                report_id = str(durable_row.get("reporte_id", ""))
                record_id = str(durable_row.get("record_id", ""))
                if report_id in seen_report_ids:
                    fail(f"REPORTE_DURABLE_ID_DUPLICADO:{report_id}")
                seen_report_ids.add(report_id)
                if any(value is None or str(value) == "" for value in durable_row.values()):
                    fail(f"REPORTE_DURABLE_CELDA_VACIA:{report_id}")
                if any(len(str(value or "")) > 160 for value in durable_row.values()):
                    fail(f"REPORTE_DURABLE_TEXTO_LARGO:{report_id}")
                if any(
                    str(value).startswith("/")
                    or re.match(r"^[A-Za-z]:[/\\]", str(value))
                    for value in durable_row.values()
                ):
                    fail(f"REPORTE_DURABLE_RUTA_ABSOLUTA:{report_id}")
                expected_durable_hash = durable_expected.pop(record_id, None)
                if (
                    expected_durable_hash is None
                    or canonical_sha(durable_row) != expected_durable_hash
                ):
                    fail(f"REPORTE_DURABLE_NO_DEREFERENCIABLE:{report_id}")
    except (OSError, UnicodeError, csv.Error):
        fail("REPORTE_DURABLE_ILEGIBLE")
    if durable_expected or durable_count != records_count:
        fail("REPORTE_DURABLE_NO_CUBRE_INDICE_1A1")
    return {
        "errors": errors,
        "records": records_count,
        "e2": e2_count,
        "exceptions": exception_count,
        "batch_sha256": expected_batch,
    }


def _completed_expediente_matches_task(
    summary: dict[str, Any], directory: Path, task: dict[str, Any], task_path: Path
) -> bool:
    try:
        index_path = directory / "e2-neutral-index.jsonl"
        report_path = directory / "reportes-durables.tsv"
        if not index_path.is_file() or not report_path.is_file():
            return False
        if not (
            summary.get("tarea_id") == task["tarea_id"]
            and summary.get("tarea_sha256") == sha256_file(task_path)
            and summary.get("payload_id") == task["payload_id"]
            and summary.get("representacion_id") == task["representacion_id"]
            and summary.get("sha256") == task["sha256"]
            and summary.get("formato") == task["formato"]
            and summary.get("contrato_sha256") == task["contrato_sha256"]
            and summary.get("profundidad") == task["profundidad"]
            and isinstance(summary.get("parser"), str) and bool(summary.get("parser"))
            and summary.get("parser_version") == MATERIAL_BUILD_VERSION
            and summary.get("build_sha256") == material_build_sha256()
            and summary.get("privacidad") == PRIVACY_CONTRACT
            and summary.get("network_habilitada") is False
        ):
            return False
        if summary.get("index_sha256") != sha256_file(index_path) or summary.get("report_sha256") != sha256_file(report_path):
            return False
        schema_path = Path(__file__).with_name("schemas") / "barrido2-e2-neutral-record.schema.json"
        schema_validator = Draft202012Validator(json.loads(schema_path.read_text(encoding="utf-8")))
        audit = _audit_e2_files_streaming(
            index_path, report_path, summary, task, schema_validator
        )
        return not audit["errors"]
    except (OSError, ValueError, KeyError, TypeError, json.JSONDecodeError):
        return False


def _durable_row(record: dict[str, Any]) -> dict[str, str]:
    description = " · ".join(
        value for value in (
            record["nombre"], record["etiqueta"], record["texto_reactivo"], record["definicion"]
        ) if value != "NO-APLICA"
    )
    row = {
        "reporte_id": "RPT-B2-" + hashlib.sha256(
            f"{record['record_id']}\x00{record['representacion_id']}".encode()
        ).hexdigest(),
        "record_id": record["record_id"], "record_sha256": record["record_sha256"],
        "batch_id": record["batch_id"], "batch_sha256": record["batch_sha256"],
        "payload_id": record["payload_id"], "representacion_id": record["representacion_id"],
        "sha256": record["sha256"], "objeto_logico_id": record["objeto_logico_id"],
        "grado_inspeccion": "E2", "afirmacion_tipo": "HECHO_OBSERVADO",
        "objeto_tipo": record["objeto_tipo"], "localizador": record["localizador"],
        "descripcion_neutral": description,
        "frontera_inspeccion": record["frontera_inspeccion"],
        "estado": record["estado"], "privacidad": record["privacidad"], "fecha": record["fecha"],
    }
    compact = {
        field: safe_text(value, durable=True)[0]
        if field in {"objeto_tipo", "localizador", "descripcion_neutral", "frontera_inspeccion"}
        else str(value)
        for field, value in row.items()
    }
    if any(value == "" or len(value) > 160 for value in compact.values()):
        raise ValueError("REPORTE_DURABLE_CELDA_INVALIDA")
    return compact


def _inspect_task_inner(
    task_path: Path,
    roots_config: Path,
    contract_path: Path,
    staging_dir: Path,
    *,
    verify_network: bool = True,
    reuse_source_dir: Path | None = None,
) -> dict[str, Any]:
    # `verify_network` se conserva sólo por compatibilidad de llamada: nunca
    # desactiva la atestación. No existe una ruta de apertura sin esta prueba.
    assert_network_disabled()
    task = json.loads(task_path.read_text(encoding="utf-8"))
    if set(task) != TASK_ALLOWED or task.get("network_habilitada") is not False:
        raise ValueError("TAREA_CIEGA_INVALIDA")
    payload_id = str(task.get("payload_id", ""))
    if not valid_payload_id(payload_id, allow_no_aplica=True):
        raise ValueError("TAREA_PAYLOAD_ID_PRIVADO_O_INVALIDO")
    if task.get("wave_initial") not in {"W1", "W2", "W3", "W4"}:
        raise ValueError("TAREA_OLA_INICIAL_INVALIDA")
    if task.get("contrato_sha256") != sha256_file(contract_path):
        raise ValueError("CONTRATO_OBSOLETO")
    roots, _ = load_roots(roots_config)
    relative = normalize_relative(task["ruta_relativa"])
    path = roots[task["root_id"]] / relative
    route_lstat = path.lstat()
    resolved = path.resolve(strict=True)
    route_guard = (route_lstat.st_dev, route_lstat.st_ino, route_lstat.st_mtime_ns, str(resolved))
    resolved.relative_to(roots[task["root_id"]])
    if not resolved.is_file() or sha256_file(resolved) != task["sha256"]:
        raise MaterialDriftError("TAREA_RUTA_HASH_NO_RECONCILIA")
    before_stat = resolved.stat()
    before_guard = (before_stat.st_size, before_stat.st_mtime_ns, before_stat.st_ino)
    reused_from = "NO-APLICA"
    source_records: list[dict[str, Any]] = []
    if reuse_source_dir is not None:
        source_summary_path = reuse_source_dir / "resumen.json"
        source_index_path = reuse_source_dir / "e2-neutral-index.jsonl"
        source_report_path = reuse_source_dir / "reportes-durables.tsv"
        if not all(path.is_file() for path in (source_summary_path, source_index_path, source_report_path)):
            raise ValueError("REUSE_EXPEDIENTE_INCOMPLETO")
        source_summary = json.loads(source_summary_path.read_text(encoding="utf-8"))
        if not (
            source_summary.get("sha256") == task["sha256"]
            and source_summary.get("formato") == task["formato"]
            and source_summary.get("contrato_sha256") == task["contrato_sha256"]
            and source_summary.get("profundidad") == task["profundidad"]
            and source_summary.get("parser_version") == MATERIAL_BUILD_VERSION
            and source_summary.get("build_sha256") == material_build_sha256()
            and source_summary.get("privacidad") == PRIVACY_CONTRACT
            and source_summary.get("network_habilitada") is False
            and source_summary.get("index_sha256") == sha256_file(source_index_path)
            and source_summary.get("report_sha256") == sha256_file(source_report_path)
        ):
            raise ValueError("REUSE_EXPEDIENTE_NO_EXACTO")
        source_records = [json.loads(line) for line in source_index_path.read_text(encoding="utf-8").splitlines() if line]
        if not source_records:
            raise ValueError("REUSE_EXPEDIENTE_SIN_OBJETOS")
        source_batch = canonical_sha(sorted(str(record.get("record_sha256", "")) for record in source_records))
        schema_path = Path(__file__).with_name("schemas") / "barrido2-e2-neutral-record.schema.json"
        schema_validator = Draft202012Validator(json.loads(schema_path.read_text(encoding="utf-8")))
        for record in source_records:
            seed_payload = {
                key: value for key, value in record.items()
                if key not in {"record_id", "record_sha256", "batch_id", "batch_sha256"}
            }
            expected_record_id = "E2R-" + canonical_sha(seed_payload)
            expected_record_sha = canonical_sha({**seed_payload, "record_id": expected_record_id})
            if (
                list(schema_validator.iter_errors(record))
                or record.get("record_id") != expected_record_id
                or record.get("record_sha256") != expected_record_sha
                or record.get("batch_id") != "E2B-" + source_batch
                or record.get("batch_sha256") != source_batch
                or record.get("representacion_id") != source_summary.get("representacion_id")
                or record.get("sha256") != task["sha256"]
                or record.get("parser") != source_summary.get("parser")
                or record.get("parser_version") != MATERIAL_BUILD_VERSION
            ):
                raise ValueError("REUSE_EXPEDIENTE_REGISTRO_INVALIDO")
        if source_summary.get("batch_sha256") != source_batch:
            raise ValueError("REUSE_EXPEDIENTE_LOTE_INVALIDO")
        with source_report_path.open(encoding="utf-8-sig", newline="") as handle:
            source_reader = csv.DictReader(handle, delimiter="\t")
            source_reports = list(source_reader)
        if (
            source_reader.fieldnames != REPORT_FIELDS
            or sorted(source_reports, key=lambda row: row["record_id"])
            != sorted((_durable_row(record) for record in source_records), key=lambda row: row["record_id"])
        ):
            raise ValueError("REUSE_EXPEDIENTE_REPORTE_INVALIDO")
        parser = str(source_summary["parser"])
        boundary = str(source_summary["frontera_inspeccion"])
        reused_from = str(source_summary["representacion_id"])
        raw_objects = [{
            "locator": record["localizador"], "type": record["objeto_tipo"],
            "name": record["nombre"], "label": record["etiqueta"],
            "question": record["texto_reactivo"], "definition": record["definicion"],
            "categories": record["categorias"], "value_labels": record["value_labels"],
            "unit": record["unidad"], "period": record["periodo"],
            "population": record["poblacion"], "page": record["pagina"],
            "sheet": record["hoja"], "table": record["tabla"], "state": record["estado"],
            "source_privacy": record["privacidad"],
            "source_parent_id": record["objeto_padre_id"],
            "source_object_id": record["objeto_logico_id"],
            "source_relation": record["relacion_estructural"],
            "depth": record["depth"],
        } for record in source_records]
    else:
        try:
            raw_objects, parser, boundary = inspect_e2(resolved)
        except (
            zipfile.BadZipFile, PermissionError, EOFError, ValueError,
            SyntaxError, OSError, subprocess.SubprocessError, NotImplementedError,
        ) as exc:
            parser = "barrido2-excepcion-material-1"
            boundary = f"apertura detenida en parser de formato: {type(exc).__name__}"
            raw_objects = [{
                "locator": "contenido=raiz",
                "type": "EXCEPCION-MATERIAL",
                "name": "contenido-raiz",
                "definition": f"EXCEPCION-ESPECIFICA:{type(exc).__name__};detalle_sha256={hashlib.sha256(str(exc).encode()).hexdigest()}",
                "state": "EXCEPCION-ESPECIFICA",
            }]
    after_stat = resolved.stat()
    after_guard = (after_stat.st_size, after_stat.st_mtime_ns, after_stat.st_ino)
    try:
        after_route_lstat = path.lstat()
        after_resolved = path.resolve(strict=True)
        after_route_guard = (
            after_route_lstat.st_dev, after_route_lstat.st_ino,
            after_route_lstat.st_mtime_ns, str(after_resolved),
        )
    except OSError as exc:
        raise MaterialDriftError("RUTA_CAMBIO_DURANTE_APERTURA") from exc
    if (
        route_guard != after_route_guard
        or before_guard != after_guard
        or sha256_file(resolved) != task["sha256"]
    ):
        raise MaterialDriftError("MATERIAL_CAMBIO_DURANTE_APERTURA")
    if reuse_source_dir is None:
        raw_by_locator = {str(raw.get("locator", "objeto")): raw for raw in raw_objects}
        if len(raw_by_locator) != len(raw_objects):
            raise ValueError("LOCALIZADORES_E2_NO_UNICOS")
        depth_cache: dict[str, int] = {}

        def locator_depth(locator: str, trail: set[str] | None = None) -> int:
            if locator in depth_cache:
                return depth_cache[locator]
            trail = set(trail or ())
            if locator in trail:
                raise ValueError("CICLO_JERARQUIA_E2")
            trail.add(locator)
            parent = raw_by_locator[locator].get("parent_locator")
            if not parent:
                depth = 0
            elif str(parent) not in raw_by_locator:
                raise ValueError(f"PADRE_E2_INEXISTENTE:{parent}")
            else:
                depth = locator_depth(str(parent), trail) + 1
            depth_cache[locator] = depth
            return depth

        for locator, raw in raw_by_locator.items():
            raw["depth"] = locator_depth(locator)
    records = [_e2_record(task, raw, parser, boundary) for raw in raw_objects]
    # Los contenedores grandes pueden producir cientos de miles de objetos.
    # Los registros ya contienen la proyección completa; conservar además el
    # árbol crudo duplica varios GiB sin aportar información al expediente.
    raw_objects.clear()
    source_records.clear()
    if reuse_source_dir is None:
        raw_by_locator.clear()
        depth_cache.clear()
    batch_hash = canonical_sha(sorted(row["record_sha256"] for row in records))
    batch_id = "E2B-" + batch_hash
    for record in records:
        record["batch_id"] = batch_id
        record["batch_sha256"] = batch_hash
        # record_sha256 excluye hashes de lote por contrato.
    index_path = staging_dir / "e2-neutral-index.jsonl"
    ordered_records = sorted(
        records, key=lambda row: (row["objeto_logico_id"], row["record_id"])
    )
    _atomic_write_lines(index_path, (
        json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(",", ":")) + "\n"
        for record in ordered_records
    ))
    report_path = staging_dir / "reportes-durables.tsv"
    ordered_for_report = sorted(
        records,
        key=lambda record: hashlib.sha256(
            f"{record['record_id']}\x00{record['representacion_id']}".encode()
        ).hexdigest(),
    )
    _atomic_write_tsv_rows(
        report_path, REPORT_FIELDS,
        (_durable_row(record) for record in ordered_for_report),
    )
    summary = {
        "tarea_id": task["tarea_id"], "representacion_id": task["representacion_id"],
        "tarea_sha256": sha256_file(task_path),
        "payload_id": task["payload_id"], "formato": task["formato"],
        "sha256": task["sha256"], "objetos_e1": len(records),
        "objetos_e2": sum(row["estado"] == "E2-COMPLETO" for row in records),
        "excepciones": sum(row["estado"] == "EXCEPCION-ESPECIFICA" for row in records),
        "network_habilitada": False, "index_sha256": sha256_file(index_path),
        "report_sha256": sha256_file(report_path), "batch_sha256": batch_hash,
        "contrato_sha256": task["contrato_sha256"],
        "profundidad": task["profundidad"],
        "parser": parser,
        "parser_version": MATERIAL_BUILD_VERSION,
        "build_sha256": material_build_sha256(),
        "frontera_inspeccion": boundary,
        "privacidad": PRIVACY_CONTRACT,
        "reutilizada_desde_representacion_id": reused_from,
    }
    _atomic_write_text(
        staging_dir / "resumen.json",
        json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
    )
    return summary


def inspect_task(
    task_path: Path,
    roots_config: Path,
    contract_path: Path,
    staging_dir: Path,
    *,
    verify_network: bool = True,
    reuse_source_dir: Path | None = None,
) -> dict[str, Any]:
    """Ejecuta una tarea con límites efectivos de tiempo y memoria.

    La tarea sigue siendo validada de nuevo por ``_inspect_task_inner`` antes
    de abrir bytes. Este envoltorio sólo convierte el presupuesto cegado en
    límites del proceso; siempre restaura el estado para que las pruebas
    sintéticas puedan compartir intérprete.
    """
    import resource
    import signal

    task = json.loads(task_path.read_text(encoding="utf-8"))
    budget = task.get("presupuesto")
    expected_budget = {
        "timeout_segundos": 1800,
        "memoria_mib": 4096,
        "miembro_max_bytes": 8 * 1024**3,
        "temp_max_bytes": "MIN-50GIB-10PORCIENTO-LIBRE",
    }
    if budget != expected_budget:
        raise ValueError("TAREA_PRESUPUESTO_INVALIDO")
    timeout_seconds = int(budget["timeout_segundos"])
    memory_bytes = int(budget["memoria_mib"]) * 1024**2
    old_limit = resource.getrlimit(resource.RLIMIT_AS)
    old_handler = signal.getsignal(signal.SIGALRM)

    def _timeout_handler(_signum: int, _frame: object) -> None:
        raise TimeoutError("TAREA_TIMEOUT_PRESUPUESTO")

    hard_limit = old_limit[1]
    bounded_soft = memory_bytes if hard_limit == resource.RLIM_INFINITY else min(memory_bytes, hard_limit)
    if old_limit[0] != resource.RLIM_INFINITY:
        bounded_soft = min(bounded_soft, old_limit[0])
    try:
        resource.setrlimit(resource.RLIMIT_AS, (bounded_soft, hard_limit))
        signal.signal(signal.SIGALRM, _timeout_handler)
        signal.alarm(timeout_seconds)
        return _inspect_task_inner(
            task_path, roots_config, contract_path, staging_dir,
            verify_network=verify_network, reuse_source_dir=reuse_source_dir,
        )
    finally:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, old_handler)
        resource.setrlimit(resource.RLIMIT_AS, old_limit)


def validate_material_snapshot(snapshot: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if snapshot.get("network_habilitada") is not False:
        errors.append("NETWORK_HABILITADA_NO_FALSE")
    if snapshot.get("authorized_roots") != list(AUTHORIZED_ROOTS):
        errors.append("RAICES_AUTORIZADAS_INVALIDAS")
    if not HASH_RE.fullmatch(str(snapshot.get("manifest_sha", ""))) or not HASH_RE.fullmatch(str(snapshot.get("roots_config_sha256", ""))):
        errors.append("HASHES_BASE_SNAPSHOT_INVALIDOS")
    declarations = snapshot.get("declarations", [])
    representations = snapshot.get("representations", [])
    contents = snapshot.get("contents", [])
    counts = snapshot.get("counts", {})
    expected_counts = {
        "declaraciones_totales": len(declarations),
        "declaraciones_con_archivo_sha": sum(
            row.get("ruta_relativa") != "NO-APLICA" and row.get("sha256_declarado") != "NO-APLICA"
            for row in declarations
        ),
        "declaraciones_sin_archivo_sha": sum(
            row.get("estado_administrativo") == "DECLARACION-SIN-ARCHIVO-SHA" for row in declarations
        ),
        "representaciones_fisicas": len(representations),
        "contenidos_sha_unicos": len(contents),
        "representaciones_declaradas": sum(row.get("coincidencia_manifiesto") == "DECLARADA" for row in representations),
        "representaciones_no_declaradas": sum(row.get("coincidencia_manifiesto") == "NO-DECLARADA" for row in representations),
        "fuera_de_disco": sum(row.get("estado_e0") == "FUERA-DE-DISCO" for row in declarations),
    }
    if counts != expected_counts:
        errors.append("CONTEOS_SNAPSHOT_NO_RECONCILIAN")
    payload_ids = [str(row.get("payload_id", "")) for row in declarations]
    if len(payload_ids) != len(set(payload_ids)) or any(not valid_payload_id(value) for value in payload_ids):
        errors.append("PAYLOAD_IDS_SNAPSHOT_INVALIDOS")
    for declaration in declarations:
        if declaration.get("estado_administrativo") not in {"TERMINAL", "DECLARACION-SIN-ARCHIVO-SHA"}:
            errors.append(f"DECLARACION_NO_TERMINAL:{declaration.get('payload_id', '')}")
        if declaration.get("estado_e0") not in E0_TERMINALS:
            errors.append(f"DECLARACION_E0_INVALIDO:{declaration.get('payload_id', '')}")
    rep_ids = [row.get("representacion_id", "") for row in representations]
    if len(rep_ids) != len(set(rep_ids)) or any(not re.fullmatch(r"REP-[0-9a-f]{64}", value) for value in rep_ids):
        errors.append("REPRESENTACION_IDS_INVALIDOS")
    for row in representations:
        try:
            expected = representation_id(row["root_id"], row["ruta_relativa"], row["sha256"])
        except (KeyError, ValueError):
            expected = ""
        if row.get("representacion_id") != expected:
            errors.append(f"REPRESENTACION_IDENTIDAD_INVALIDA:{row.get('representacion_id', '')}")
        if row.get("wave_initial") not in {"W1", "W2", "W3", "W4"}:
            errors.append(f"OLA_INVALIDA:{row.get('representacion_id', '')}")
        elif row.get("wave_initial") != assign_wave(
            str(row.get("ruta_relativa", "")), int(row.get("tamano_observado", -1)), row.get("zip_geometry", {})
        ):
            errors.append(f"OLA_NO_REDERIVABLE:{row.get('representacion_id', '')}")
        if row.get("estado_e0") not in E0_TERMINALS:
            errors.append(f"E0_NO_TERMINAL:{row.get('representacion_id', '')}")
        row_payloads = row.get("payload_ids", [])
        if (
            not isinstance(row_payloads, list)
            or not row_payloads
            or len(row_payloads) != len(set(row_payloads))
            or ("NO-APLICA" in row_payloads and len(row_payloads) != 1)
            or any(value != "NO-APLICA" and value not in payload_ids for value in row_payloads)
        ):
            errors.append(f"REPRESENTACION_PAYLOADS_INVALIDOS:{row.get('representacion_id', '')}")
    representation_by_id = {str(row.get("representacion_id", "")): row for row in representations}
    for declaration in declarations:
        rep_id = str(declaration.get("representacion_id", ""))
        if rep_id == "NO-APLICA":
            continue
        representation = representation_by_id.get(rep_id)
        if (
            representation is None
            or declaration.get("payload_id") not in representation.get("payload_ids", [])
            or declaration.get("sha256_observado") != representation.get("sha256")
        ):
            errors.append(f"DECLARACION_REPRESENTACION_JOIN_INVALIDO:{declaration.get('payload_id', '')}")
    content_counts = Counter(str(row.get("sha256", "")) for row in representations)
    for representation in representations:
        if representation.get("duplicate_content_count") != content_counts[str(representation.get("sha256", ""))]:
            errors.append(f"DUPLICACION_CONTENIDO_INVALIDA:{representation.get('representacion_id', '')}")
        expected_declared = representation.get("payload_ids") != ["NO-APLICA"]
        if representation.get("coincidencia_manifiesto") != ("DECLARADA" if expected_declared else "NO-DECLARADA"):
            errors.append(f"COINCIDENCIA_MANIFIESTO_INVALIDA:{representation.get('representacion_id', '')}")
    declared_contents = {str(row.get("sha256", "")): row.get("representaciones") for row in contents}
    if dict(content_counts) != declared_contents:
        errors.append("CONTENIDOS_NO_RECONCILIAN")
    if snapshot.get("snapshot_sha256") != canonical_sha({k: v for k, v in snapshot.items() if k != "snapshot_sha256"}):
        errors.append("SNAPSHOT_HASH_INVALIDO")
    return errors


def validate_material_files(
    snapshot_path: Path,
    contract_path: Path,
    task_root: Path,
    ledger_path: Path,
    staging_root: Path | None = None,
    *,
    require_complete: bool = False,
) -> dict[str, Any]:
    """Valida offline snapshot, partición, contratos, reanudación y E2."""
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    errors = validate_material_snapshot(snapshot)
    contract_hash = sha256_file(contract_path)
    with ledger_path.open(encoding="utf-8-sig", newline="") as handle:
        ledger = list(csv.DictReader(handle, delimiter="\t"))
    expected = {row["representacion_id"]: row for row in snapshot["representations"]}
    observed_ids = [row.get("representacion_id", "") for row in ledger]
    if len(observed_ids) != len(set(observed_ids)) or set(observed_ids) != set(expected):
        errors.append("LEDGER_NO_CUBRE_REPRESENTACIONES_1A1")
    expected_task_ids = {str(row.get("tarea_id", "")) for row in ledger}
    actual_task_ids = {path.stem for path in task_root.glob("*.json")}
    if actual_task_ids != expected_task_ids:
        errors.append("TASK_ROOT_NO_CUBRE_LEDGER_1A1")
    wave_sets: dict[str, set[str]] = {wave: set() for wave in ("W1", "W2", "W3", "W4")}
    for row in ledger:
        rep_id = row.get("representacion_id", "")
        expected_representation = expected.get(rep_id, {})
        expected_task_id = "TASK-B2-" + hashlib.sha256(
            f"{snapshot['snapshot_sha256']}\x00{rep_id}\x00{contract_hash}".encode("utf-8")
        ).hexdigest()
        expected_payloads = sorted(expected_representation.get("payload_ids", []))
        expected_ledger_values = {
            "tarea_id": expected_task_id,
            "payload_id": expected_payloads[0] if expected_payloads else "",
            "payload_ids_json": json.dumps(expected_payloads, ensure_ascii=False, separators=(",", ":")),
            "root_id": str(expected_representation.get("root_id", "")),
            "ruta_relativa": str(expected_representation.get("ruta_relativa", "")),
            "sha256": str(expected_representation.get("sha256", "")),
            "wave_initial": str(expected_representation.get("wave_initial", "")),
            "contrato_sha256": contract_hash,
        }
        for field, expected_value in expected_ledger_values.items():
            if str(row.get(field, "")) != expected_value:
                errors.append(f"LEDGER_SNAPSHOT_JOIN_INVALIDO:{rep_id}:{field}")
        wave = row.get("wave_initial", "")
        if wave in wave_sets:
            wave_sets[wave].add(rep_id)
        else:
            errors.append(f"LEDGER_OLA_INVALIDA:{rep_id}")
        if row.get("wave_retry_ref") != "NO-APLICA" and row.get("wave_retry_ref") not in expected:
            errors.append(f"W5_REFERENCIA_AJENA:{rep_id}")
        if row.get("contrato_sha256") != contract_hash:
            errors.append(f"LEDGER_CONTRATO_OBSOLETO:{rep_id}")
        task_path = task_root / f"{row.get('tarea_id', '')}.json"
        if not task_path.is_file():
            errors.append(f"TAREA_INEXISTENTE:{rep_id}")
            continue
        task = json.loads(task_path.read_text(encoding="utf-8"))
        if set(task) != TASK_ALLOWED or task.get("network_habilitada") is not False:
            errors.append(f"TAREA_NO_CIEGA:{rep_id}")
        for field in ("tarea_id", "representacion_id", "payload_id", "root_id", "ruta_relativa", "sha256", "wave_initial", "contrato_sha256"):
            if str(task.get(field, "")) != str(row.get(field, "")):
                errors.append(f"TAREA_LEDGER_JOIN_INVALIDO:{rep_id}:{field}")
        if str(task.get("formato", "")) != str(expected_representation.get("extension", "")):
            errors.append(f"TAREA_SNAPSHOT_JOIN_INVALIDO:{rep_id}:formato")
    wave_names = tuple(wave_sets)
    for index, left in enumerate(wave_names):
        for right in wave_names[index + 1:]:
            if wave_sets[left].intersection(wave_sets[right]):
                errors.append(f"OLAS_SOLAPADAS:{left}:{right}")
    if set().union(*wave_sets.values()) != set(expected):
        errors.append("OLAS_NO_EXHAUSTIVAS")

    completed = 0
    e2_records = 0
    if staging_root is not None and staging_root.is_dir():
        schema_path = Path(__file__).with_name("schemas") / "barrido2-e2-neutral-record.schema.json"
        schema_validator = Draft202012Validator(json.loads(schema_path.read_text(encoding="utf-8")))
        summaries: dict[str, Path] = {}
        for path in staging_root.rglob("resumen.json"):
            summary_rep = str(json.loads(path.read_text(encoding="utf-8")).get("representacion_id", ""))
            if summary_rep not in expected:
                errors.append(f"EXPEDIENTE_REPRESENTACION_AJENA:{summary_rep}")
            if summary_rep in summaries:
                errors.append(f"EXPEDIENTE_REPRESENTACION_NO_UNICO:{summary_rep}")
            else:
                summaries[summary_rep] = path.parent
        for row in ledger:
            rep_id = row.get("representacion_id", "")
            if require_complete and row.get("estado_terminal") != "SI":
                errors.append(f"LEDGER_NO_TERMINAL:{rep_id}")
            directory = summaries.get(rep_id)
            if directory is None:
                if require_complete:
                    errors.append(f"REPRESENTACION_SIN_E2:{rep_id}")
                continue
            summary = json.loads((directory / "resumen.json").read_text(encoding="utf-8"))
            current_task_path = task_root / f"{row.get('tarea_id', '')}.json"
            if not current_task_path.is_file():
                errors.append(f"TAREA_INEXISTENTE_EN_GATE_E2:{rep_id}")
                continue
            current_task = json.loads(current_task_path.read_text(encoding="utf-8"))
            index_path = directory / "e2-neutral-index.jsonl"
            report_path = directory / "reportes-durables.tsv"
            if not index_path.is_file() or not report_path.is_file():
                errors.append(f"EXPEDIENTE_E2_INCOMPLETO:{rep_id}")
                continue
            if summary.get("network_habilitada") is not False:
                errors.append(f"E2_NETWORK_INVALIDA:{rep_id}")
            if (
                summary.get("tarea_id") != row.get("tarea_id")
                or summary.get("tarea_sha256") != sha256_file(current_task_path)
                or summary.get("payload_id") != row.get("payload_id")
                or summary.get("formato") != current_task.get("formato")
                or summary.get("contrato_sha256") != contract_hash
                or summary.get("profundidad") != "E2-COMPLETO"
                or summary.get("parser_version") != MATERIAL_BUILD_VERSION
                or summary.get("build_sha256") != material_build_sha256()
                or summary.get("privacidad") != PRIVACY_CONTRACT
                or not summary.get("frontera_inspeccion")
            ):
                errors.append(f"E2_REUSE_CONTRATO_INEXACTO:{rep_id}")
            reuse_rep = summary.get("reutilizada_desde_representacion_id", "NO-APLICA")
            if reuse_rep != "NO-APLICA" and (
                reuse_rep == rep_id
                or reuse_rep not in expected
                or expected[reuse_rep].get("sha256") != row.get("sha256")
            ):
                errors.append(f"E2_REUSE_CONTENIDO_INVALIDO:{rep_id}")
            if summary.get("index_sha256") != sha256_file(index_path) or summary.get("report_sha256") != sha256_file(report_path):
                errors.append(f"E2_HASH_INVALIDO:{rep_id}")
            if row.get("estado_terminal") == "SI" and (
                row.get("reporte_sha256") != summary.get("report_sha256")
                or row.get("parser_version") != summary.get("parser_version")
                or row.get("build_sha256") != summary.get("build_sha256")
                or row.get("profundidad") != summary.get("profundidad")
                or row.get("frontera_inspeccion") != summary.get("frontera_inspeccion")
                or row.get("privacidad") != summary.get("privacidad")
            ):
                errors.append(f"LEDGER_E2_REUSE_INEXACTO:{rep_id}")
            audit = _audit_e2_files_streaming(
                index_path, report_path, summary, current_task, schema_validator
            )
            errors.extend(f"{code}:{rep_id}" for code in audit["errors"])
            completed += 1
            e2_records += int(audit["records"])
    if require_complete and completed != len(expected):
        errors.append("GATE_E2_INCOMPLETO")
    return {
        "ok": not errors,
        "errors": errors,
        "representations": len(expected),
        "completed": completed,
        "e2_records": e2_records,
        "waves": {wave: len(values) for wave, values in wave_sets.items()},
        "network_habilitada": False,
    }
