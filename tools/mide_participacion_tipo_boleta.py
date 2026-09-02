#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ACTO MAESTRA35-L3 · CIVICA-TIPO-DE-BOLETA — medidor por TIPO de boleta federal.

Hermano de tools/l6_estimador_concurrencia.py (ACTO MAESTRA34-L6): importa sus
lectores y NO lo modifica, de modo que la corrida de L6 se reproduce byte a byte
como control de regresion (--control-l6).

Sustituye la tendencia lineal de L6 (unica gamma) por un efecto fijo de TIPO de
anio federal, que es la firma c1 de mesa (2/sep/2026):

    Dy(m,k) = alpha*hueco(k) + beta_pres*DD_pres(k) + beta_int*DD_int(k) + e(m,k)

Modos:
    --tabla-identificacion   escribe la tabla de identificacion (P0), TSV a stdout
    --control-l6             re-corre el estimador de L6 y compara con su JSON
    --json <ruta>            corre el modelo de arriba y vuelca el resultado
"""
import argparse, json, os, re, sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

CALENDARIO = "data/p0-calendario-ayuntamientos-v1_0.tsv"
TRATAMIENTO = "data/p0-tratamiento-homologacion-v1_0.tsv"

# Anios federales de la ventana y su TIPO. Fuente: el propio calendario del INE
# ya cifrado en data/p0-calendario-ayuntamientos-v1_0.tsv (columna
# nota_concurrencia dice "anio federal" para 2018, 2021 y 2024) mas el hecho
# constitucional de que la eleccion presidencial cae cada seis anios (2018, 2024)
# y la intermedia en el punto medio (2015, 2021).
TIPO_FEDERAL = {2015: "intermedia", 2018: "presidencial",
                2021: "intermedia", 2024: "presidencial"}


# ───────────────────────────── calendario y tipos ─────────────────────────────
def lee_calendario(ruta=CALENDARIO):
    """{entidad: {anio: concurrente_bool}} leido por linea (nunca con csv)."""
    cal = {}
    with open(ruta, encoding="utf-8") as fh:
        lineas = [l.rstrip("\n") for l in fh]
    cab = lineas[1].split("\t")
    i_ent, i_anio = cab.index("entidad"), cab.index("anio_jornada")
    i_conc, i_ayto = cab.index("concurrente_con_federal"), cab.index("ayuntamientos")
    for l in lineas[2:]:
        if not l.strip():
            continue
        p = l.split("\t")
        if p[i_ayto] != "SI":
            continue
        cal.setdefault(p[i_ent], {})[int(p[i_anio])] = (p[i_conc] == "SI")
    return cal


def dpres_dint(anio, concurrente):
    """D_pres, D_int de una eleccion local, segun la spec §1.5 de este acto."""
    if not concurrente:
        return 0, 0
    t = TIPO_FEDERAL.get(anio)
    return (1 if t == "presidencial" else 0), (1 if t == "intermedia" else 0)


def clase_transicion(dp, di):
    if dp == 0 and di == 0:
        return "STAY"
    return "SWITCH"


def etiqueta_pata(anio, conc):
    if not conc:
        return "sin-federal"
    return TIPO_FEDERAL.get(anio, "federal-?")


# ────────────────────────── P0 · tabla de identificacion ──────────────────────
def tabla_identificacion(ventana=(2015, 2024)):
    """Una fila por entidad x transicion consecutiva, para TODAS las entidades
    del calendario. Se escribe ANTES de abrir un solo resultado."""
    cal = lee_calendario()
    trat = {}
    with open(TRATAMIENTO, encoding="utf-8") as fh:
        lineas = [l.rstrip("\n") for l in fh]
    cab = lineas[1].split("\t")
    for l in lineas[2:]:
        if not l.strip():
            continue
        p = l.split("\t")
        trat[p[cab.index("entidad")]] = (p[cab.index("estatus")], p[cab.index("cohorte")])

    filas = []
    for ent in sorted(cal):
        anios = sorted(a for a in cal[ent] if ventana[0] <= a <= ventana[1])
        for a, b in zip(anios, anios[1:]):
            dpa, dia = dpres_dint(a, cal[ent][a])
            dpb, dib = dpres_dint(b, cal[ent][b])
            ddp, ddi = dpb - dpa, dib - dia
            est, coh = trat.get(ent, ("?", ""))
            filas.append({
                "entidad": ent, "de": a, "a": b, "hueco": b - a,
                "tipo_de": etiqueta_pata(a, cal[ent][a]),
                "tipo_a": etiqueta_pata(b, cal[ent][b]),
                "dD_pres": ddp, "dD_int": ddi,
                "clase": clase_transicion(ddp, ddi),
                "estatus_entidad": est, "cohorte": coh,
                "identifica": identifica(ddp, ddi)})
    return filas


def identifica(ddp, ddi):
    if ddp == 0 and ddi == 0:
        return "alpha"
    if ddp != 0 and ddi == 0:
        return "beta_pres (+alpha)"
    if ddp == 0 and ddi != 0:
        return "beta_int (+alpha)"
    return "beta_pres - beta_int (+alpha)"


def imprime_tabla(filas):
    cols = ["entidad", "de", "a", "hueco", "tipo_de", "tipo_a", "dD_pres",
            "dD_int", "clase", "identifica", "estatus_entidad", "cohorte"]
    print("\t".join(cols))
    for f in filas:
        print("\t".join(str(f[c]) for c in cols))


def resumen_identificacion(filas, solo=None):
    """Conteo de transiciones por parametro identificado. `solo` = conjunto de
    entidades del panel; None = universo entero del calendario."""
    sub = [f for f in filas if solo is None or f["entidad"] in solo]
    c = {}
    for f in sub:
        c[f["identifica"]] = c.get(f["identifica"], 0) + 1
    return {"n_transiciones": len(sub), "por_parametro": c,
            "n_STAY": sum(1 for f in sub if f["clase"] == "STAY"),
            "n_SWITCH": sum(1 for f in sub if f["clase"] == "SWITCH"),
            "entidades": sorted({f["entidad"] for f in sub})}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tabla-identificacion", action="store_true")
    ap.add_argument("--resumen", action="store_true")
    ap.add_argument("--control-l6", action="store_true")
    ap.add_argument("--tipo-boleta", action="store_true")
    ap.add_argument("--json", default=None)
    a = ap.parse_args()
    if a.tabla_identificacion:
        filas = tabla_identificacion()
        imprime_tabla(filas)
        if a.resumen:
            print("\n# RESUMEN (universo del calendario):", file=sys.stderr)
            print(json.dumps(resumen_identificacion(filas), ensure_ascii=False,
                             indent=1), file=sys.stderr)
        return
    if a.control_l6:
        print(json.dumps(control_regresion_l6(), ensure_ascii=False, indent=1))
        return
    if a.tipo_boleta:
        salida = corre()
        texto = json.dumps(salida, ensure_ascii=False, indent=1, default=str)
        if a.json:
            open(a.json, "w", encoding="utf-8").write(texto + "\n")
            print(f"escrito: {a.json}")
        else:
            print(texto)
        return
    ap.error("elige un modo")



# ══════════════════════════════ P2 · lectores nuevos ══════════════════════════
# Un lector por fuente, con el mismo contrato que tools/l6_lectores.py:
#   devuelve ({municipio_normalizado: (votos_totales, lista_nominal)}, control)
# El control lleva las filas excluidas y la comprobacion de identidad de la
# columna de total, que es lo que distingue «TOTAL» de «Votacion Recibida»
# sin creerle al rotulo (precedente: la columna T VOTARON de Zacatecas 2024,
# que L6 midio con 31.7 % de «Sin Dato»).

import io as _io
import re as _re
import zipfile as _zip

_SERIE = "data/raw/electoral_local_municipal_serie"


_ALIAS_MUN = {
    # Baja California: el IEEBC alterna tres formas para el mismo municipio
    "P DE ROSARITO": "PLAYAS DE ROSARITO",
    "ROSARITO": "PLAYAS DE ROSARITO",
    "PLAYAS DE ROSARITO": "PLAYAS DE ROSARITO",
    # Chihuahua: el IEECH usa dos formas del mismo municipio entre anios.
    # Detectados por el diagnostico de municipios_perdidos de §1.3, que existe
    # justamente para distinguir una AUSENCIA de una discordancia de nombre:
    # Batopilas se llama oficialmente «Batopilas de Manuel Gomez Morin» desde
    # 2016 y el IEECH adopta el nombre largo a partir de 2021.
    "BATOPILAS": "BATOPILAS DE MANUEL GOMEZ MORIN",
    "BATOPILAS DE MANUEL GOMEZ MORIN": "BATOPILAS DE MANUEL GOMEZ MORIN",
    "NVO CASAS GRANDES": "NUEVO CASAS GRANDES",
    "NUEVO CASAS GRANDES": "NUEVO CASAS GRANDES",
}


def _norm_mun(s):
    import unicodedata
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = _re.sub(r"[^A-Za-z0-9 ]", " ", s)
    s = _re.sub(r"\s+", " ", s).strip().upper()
    s = _re.sub(r"^GRAL\.?\s+", "GENERAL ", s)
    return _ALIAS_MUN.get(s, s)


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("%", "").replace(" ", "")
    if s in ("", "-", "--", "N/A", "None", "SinDato", "Sin Dato"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


_NO_MUNICIPIO = _re.compile(r"^(TOTAL|TOTALES|SUMA|ESTATAL|GRAN TOTAL|VOTO EN EL EXTRANJERO|"
                            r"EXTRANJERO|ESPECIAL|RESUMEN|CONCENTRADO)", _re.I)


def _filas_xlsx(path_o_bytes, hoja=None):
    import openpyxl
    fuente = _io.BytesIO(path_o_bytes) if isinstance(path_o_bytes, bytes) else path_o_bytes
    wb = openpyxl.load_workbook(fuente, read_only=True, data_only=True)
    hojas = wb.sheetnames if hoja is None else [hoja]
    for h in hojas:
        ws = wb[h]
        yield h, [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()


def _filas_xls(path, hoja=None):
    import xlrd
    wb = xlrd.open_workbook(path)
    hojas = wb.sheet_names() if hoja is None else [hoja]
    for h in hojas:
        ws = wb.sheet_by_name(h)
        yield h, [ws.row_values(i) for i in range(ws.nrows)]


def _localiza_cabecera(filas, patrones, limite=20):
    """Devuelve (indice_de_fila, {clave: indice_de_columna}). La cabecera se
    localiza por REGEX sobre las primeras filas, nunca por indice fijo: el
    IEECH mueve la cabecera de la f7 a la f5 entre sus propios anios."""
    for i, fila in enumerate(filas[:limite]):
        celdas = ["" if c is None else str(c).replace("\n", " ").strip() for c in fila]
        idx = {}
        for clave, pats in patrones.items():
            for j, c in enumerate(celdas):
                if any(_re.fullmatch(p, c, _re.I) for p in pats):
                    idx[clave] = j
                    break
        if len(idx) == len(patrones):
            return i, idx
    return None, None


def _agrega(filas, i_cab, idx, etiqueta):
    """Suma por municipio y guarda las filas excluidas, una por una."""
    datos, excl, sin_dato = {}, [], 0
    for fila in filas[i_cab + 1:]:
        if max(idx.values()) >= len(fila):
            continue
        mun = fila[idx["mun"]]
        if mun is None or str(mun).strip() == "":
            continue
        k = _norm_mun(mun)
        if not k:
            continue
        tot, ln = _num(fila[idx["tot"]]), _num(fila[idx["ln"]])
        if _NO_MUNICIPIO.match(k):
            excl.append((k, tot, ln))
            continue
        if tot is None or ln is None:
            sin_dato += 1
            continue
        a = datos.setdefault(k, [0.0, 0.0, 0])
        a[0] += tot
        a[1] += ln
        a[2] += 1
    return ({k: (v[0], v[1]) for k, v in datos.items()},
            {"etiqueta": etiqueta, "excluidas": excl, "casillas_sin_dato": sin_dato,
             "casillas_por_municipio": {k: v[2] for k, v in datos.items()}})


# ─────────────────────────── Zacatecas 2016 (IEEZ) ───────────────────────────
def zacatecas2016_casilla():
    """Hoja 'Ayuntamientos' del .xls por casilla: Dtto|Municipio|Seccion|Casilla|
    Casilla|LN|…|NOREG|NULOS|TOTAL|% PARTICIPACIÓN."""
    p = os.path.join(_SERIE, "ieez_zacatecas_2016_eleccion_x_casilla.xls")
    _, filas = next(_filas_xls(p, "Ayuntamientos"))
    i, idx = _localiza_cabecera(filas, {"mun": [r"Municipio"], "ln": [r"LN"],
                                        "tot": [r"TOTAL"], "pct": [r"%\s*PARTICIPACI[OÓ]N"]})
    assert idx is not None, "cabecera de Zacatecas 2016 no localizada"
    idx_c = {k: idx[k] for k in ("mun", "ln", "tot")}
    datos, ctrl = _agrega(filas, i, idx_c, "zacatecas2016")
    # identidad de la columna de total: TOTAL / LN debe reproducir % PARTICIPACIÓN
    peor, n = 0.0, 0
    for fila in filas[i + 1:]:
        if max(idx.values()) >= len(fila):
            continue
        t, l, q = _num(fila[idx["tot"]]), _num(fila[idx["ln"]]), _num(fila[idx["pct"]])
        if None in (t, l, q) or l <= 0:
            continue
        peor = max(peor, abs(100.0 * t / l - q)); n += 1
    ctrl["identidad_total_sobre_ln_vs_pct_publicado"] = {"casillas": n, "max_abs_pp": peor}
    return datos, ctrl


def zacatecas2016_municipio_html():
    """Control independiente: la tabla HTML por municipio del propio IEEZ,
    cabecera 'Municipio | LN'. Devuelve {municipio: lista_nominal}."""
    import html as _html
    p = os.path.join(_SERIE, "ieez_zacatecas_2016_ayuntamientos_x_municipio.htm")
    s = open(p, encoding="utf-8", errors="replace").read()
    ln = {}
    for f in _re.findall(r"<TR.*?</TR>", s, _re.S | _re.I):
        celdas = [_html.unescape(_re.sub("<[^>]+>", "", c)).strip()
                  for c in _re.findall(r"<T[DH].*?</T[DH]>", f, _re.S | _re.I)]
        if len(celdas) < 2:
            continue
        k, v = _norm_mun(celdas[0]), _num(celdas[1])
        if not k or v is None or _NO_MUNICIPIO.match(k) or k == "MUNICIPIO":
            continue
        ln[k] = v
    return ln


# ────────────────────────── Baja California (IEEBC) ──────────────────────────
_BC = {2016: ("ieebc_bc_2016_municipes_x_casilla.xlsx", None),
       2019: ("ieebc_bc_2019_computo_x_casilla_mun.xls", "Todos_Ayuntamientos"),
       2021: ("ieebc_bc_2021_computo_x_casilla_mun.xls", "Todos_Ayuntamientos_Locales"),
       2024: ("ieebc_bc_2024_computo_x_casilla_mun.xls", "Todos_Ayuntamientos_Locales")}

_PAT_BC = {"mun": [r"MUNICIPIO"], "tot": [r"TOTAL\s*VOTOS"],
           "ln": [r"LISTA\s*NOMINAL"], "pct": [r"%\s*DE\s*PARTICIP\.?"]}


def bc_casilla(anio):
    arch, hoja = _BC[anio]
    p = os.path.join(_SERIE, arch)
    lector = _filas_xlsx if arch.endswith("x") else _filas_xls
    datos, excl, sin_dato, peor, n = {}, [], 0, 0.0, 0
    hojas_leidas = []
    for h, filas in lector(p, hoja):
        i, idx = _localiza_cabecera(filas, _PAT_BC)
        if idx is None:
            continue
        hojas_leidas.append(h)
        d, c = _agrega(filas, i, {k: idx[k] for k in ("mun", "ln", "tot")}, f"bc{anio}:{h}")
        for k, (t, l) in d.items():
            a = datos.setdefault(k, [0.0, 0.0]); a[0] += t; a[1] += l
        excl += c["excluidas"]; sin_dato += c["casillas_sin_dato"]
        for fila in filas[i + 1:]:
            if max(idx.values()) >= len(fila):
                continue
            t, l, q = _num(fila[idx["tot"]]), _num(fila[idx["ln"]]), _num(fila[idx["pct"]])
            if None in (t, l, q) or l <= 0:
                continue
            peor = max(peor, abs(100.0 * t / l - q)); n += 1
    return ({k: (v[0], v[1]) for k, v in datos.items()},
            {"etiqueta": f"bc{anio}", "hojas": hojas_leidas, "excluidas": excl,
             "casillas_sin_dato": sin_dato,
             "identidad_total_sobre_ln_vs_pct_publicado": {"casillas": n, "max_abs_pp": peor}})


# ──────────────────────────── Chihuahua (IEECH) ──────────────────────────────
# La cabecera y los rotulos CAMBIAN entre anios de la misma fuente:
#   2016/2018 -> f7, 'Municipio' / 'Listado Nominal' / 'Votación Total' / '% de Particip.'
#   2021/2024 -> f5-f6, 'MUNICIPIO' / 'LISTA NOMINAL' / 'TOTAL VOTOS'
_PAT_CHIH = {"mun": [r"Municipio", r"MUNICIPIO"],
             "tot": [r"Votaci[oó]n\s*Total", r"TOTAL\s*VOTOS"],
             "ln": [r"Listado\s*Nominal", r"LISTA\s*NOMINAL"]}
_PAT_CHIH_PCT = {"pct": [r"%\s*de\s*Particip\.?"]}


def chihuahua_casilla(anio):
    """El IEECH cambia de esquema entre sus propios anios y en 2016/2018 la
    cabecera es de DOS filas: 'Municipio' es un GRUPO sobre 'Clave'|'Nombre', y
    tomar la primera columna del grupo da la clave numerica en vez del nombre.
    Se detecta la subcabecera y se usa 'Nombre'."""
    p = os.path.join(_SERIE, f"ieech_chihuahua_{anio}_ayuntamientos_x_casilla.xlsx")
    h, filas = next(_filas_xlsx(p))          # la PRIMERA hoja es la de votacion por casilla
    i, idx = _localiza_cabecera(filas, _PAT_CHIH)
    assert idx is not None, f"cabecera de Chihuahua {anio} no localizada"
    inicio = i
    sub = sub_fila = None
    for k in (1, 2, 3):                       # la cabecera de 2016 ocupa TRES filas
        if i + k >= len(filas):
            break
        celdas = ["" if c is None else str(c).replace("\n", " ").strip() for c in filas[i + k]]
        for j, c in enumerate(celdas):
            if _re.fullmatch(r"Nombre", c, _re.I) and idx["mun"] <= j <= idx["mun"] + 3:
                sub, sub_fila = j, i + k
                break
        if sub is not None:
            break
    if sub is not None:
        idx = dict(idx, mun=sub)
        inicio = sub_fila                     # los datos empiezan tras la SUBcabecera
    datos, ctrl = _agrega(filas, inicio, idx, f"chihuahua{anio}")
    ctrl["hoja"] = h
    ctrl["cabecera_fila"] = i + 1
    ctrl["subcabecera_Nombre_col"] = sub
    ctrl["subcabecera_fila"] = None if sub_fila is None else sub_fila + 1
    j, ipct = _localiza_cabecera(filas, _PAT_CHIH_PCT)
    if ipct is not None:
        peor, n = 0.0, 0
        for fila in filas[inicio + 1:]:
            if max(list(idx.values()) + [ipct["pct"]]) >= len(fila):
                continue
            t, l, q = _num(fila[idx["tot"]]), _num(fila[idx["ln"]]), _num(fila[ipct["pct"]])
            if None in (t, l, q) or l <= 0:
                continue
            if q <= 1.5:                      # la fuente publica el % como fraccion en algun anio
                q *= 100.0
            peor = max(peor, abs(100.0 * t / l - q)); n += 1
        ctrl["identidad_total_sobre_ln_vs_pct_publicado"] = {"casillas": n, "max_abs_pp": peor}
    else:
        ctrl["identidad_total_sobre_ln_vs_pct_publicado"] = {
            "casillas": 0, "nota": "la fuente no publica % de participacion en esta hoja"}
    return datos, ctrl


# ═══════════════════════════ P2 · panel y estimador ══════════════════════════
SERIES_L3 = {
    "Coahuila":        [2017, 2018, 2021, 2024],
    "Nayarit":         [2017, 2021, 2024],
    "Zacatecas":       [2016, 2018, 2021, 2024],
    "Durango":         [2016, 2019],
    "Baja California": [2016, 2019, 2021, 2024],
    "Chihuahua":       [2016, 2018, 2021, 2024],
}


def panel_l3():
    """Panel de L6 (por sus lectores, sin tocarlos) mas las tres entidades que
    P1 de este acto adquirio."""
    import l6_estimador_concurrencia as L6
    P, ctrl = L6.panel()                       # Coahuila, Nayarit, Zacatecas 18/21/24, Durango
    ctrl = {"l6": ctrl}

    d, c = zacatecas2016_casilla()
    P[("Zacatecas", 2016)] = d
    ctrl["zacatecas2016"] = c
    ln_pub = zacatecas2016_municipio_html()
    comunes = set(d) & set(ln_pub)
    ctrl["zacatecas2016"]["control_reagregacion_vs_tabla_publicada"] = {
        "municipios_en_comun": len(comunes),
        "solo_en_casilla": sorted(set(d) - set(ln_pub)),
        "solo_en_tabla": sorted(set(ln_pub) - set(d)),
        "max_abs_delta_lista_nominal": max((abs(d[k][1] - ln_pub[k]) for k in comunes), default=None)}

    for a in SERIES_L3["Baja California"]:
        P[("Baja California", a)], ctrl[f"bc{a}"] = bc_casilla(a)
    for a in SERIES_L3["Chihuahua"]:
        P[("Chihuahua", a)], ctrl[f"chihuahua{a}"] = chihuahua_casilla(a)
    return P, ctrl


def tratamiento_l3():
    """D_pres y D_int por (entidad, anio), leidos del calendario del INE."""
    cal = lee_calendario()
    D = {}
    for ent, anios in SERIES_L3.items():
        for a in anios:
            conc = cal[ent][a]
            D[(ent, a)] = dpres_dint(a, conc)
    return D


def participacion(P):
    y, fuera = {}, []
    for (ent, anio), d in P.items():
        if ent not in SERIES_L3 or anio not in SERIES_L3[ent]:
            continue
        for m, (v, ln) in d.items():
            if ln is None or ln <= 0 or v is None:
                fuera.append((ent, anio, m, "sin dato")); continue
            p = 100.0 * v / ln
            if not (0 < p <= 100):
                fuera.append((ent, anio, m, f"fuera de (0,100]: {p:.4f}")); continue
            y[(ent, anio, m)] = p
    return y, fuera


def universo(P):
    univ, perdidos = {}, {}
    for ent, anios in SERIES_L3.items():
        conj = [set(P[(ent, a)]) for a in anios]
        comun = set.intersection(*conj)
        univ[ent] = comun
        perdidos[ent] = sorted(set.union(*conj) - comun)
    return univ, perdidos


def transiciones(y, univ, D):
    T = []
    for ent, anios in SERIES_L3.items():
        for a, b in zip(anios, anios[1:]):
            dpa, dia = D[(ent, a)]
            dpb, dib = D[(ent, b)]
            ddp, ddi = dpb - dpa, dib - dia
            for m in sorted(univ[ent]):
                if (ent, a, m) in y and (ent, b, m) in y:
                    T.append({"entidad": ent, "de": a, "a": b, "municipio": m,
                              "dy": y[(ent, b, m)] - y[(ent, a, m)],
                              "hueco": b - a, "ddp": ddp, "ddi": ddi,
                              "clase": clase_transicion(ddp, ddi),
                              "y0": y[(ent, a, m)], "y1": y[(ent, b, m)]})
    return T


# ── minimos cuadrados sin intercepto sobre 3 regresores (hueco, ddp, ddi) ─────
def _ols(T, regs=("hueco", "ddp", "ddi")):
    k = len(regs)
    A = [[0.0] * k for _ in range(k)]
    b = [0.0] * k
    for t in T:
        x = [t[r] for r in regs]
        for i in range(k):
            b[i] += x[i] * t["dy"]
            for j in range(k):
                A[i][j] += x[i] * x[j]
    return _resuelve(A, b, regs)


def _resuelve(A, b, regs):
    k = len(b)
    M = [row[:] + [b[i]] for i, row in enumerate(A)]
    for c in range(k):
        piv = max(range(c, k), key=lambda r: abs(M[r][c]))
        if abs(M[piv][c]) < 1e-12:
            return None
        M[c], M[piv] = M[piv], M[c]
        for r in range(k):
            if r == c:
                continue
            f = M[r][c] / M[c][c]
            for j in range(c, k + 1):
                M[r][j] -= f * M[c][j]
    return {regs[i]: M[i][k] / M[i][i] for i in range(k)}


# ─────────────────── bootstrap wild cluster (Rademacher) ─────────────────────
SEED, B_BOOT = 42, 10_000


def wild_cluster(T, coef, B=B_BOOT, seed=SEED):
    """IC95 e inversion del test para UN coeficiente, con los residuos del modelo
    RESTRINGIDO a que ese coeficiente sea 0 (Cameron-Gelbach-Miller), dejando los
    otros libres. Pesos de Rademacher sobre ENTIDADES."""
    import random
    rnd = random.Random(seed)
    full = _ols(T)
    if full is None:
        return None
    b0 = full[coef]
    libres = [r for r in ("hueco", "ddp", "ddi") if r != coef]
    h0 = _ols(T, tuple(libres))
    if h0 is None:
        return None
    res = [t["dy"] - sum(h0[r] * t[r] for r in libres) for t in T]
    ajuste = [sum(h0[r] * t[r] for r in libres) for t in T]
    ents = sorted({t["entidad"] for t in T})
    muestras = []
    for _ in range(B):
        s = {e: (1.0 if rnd.random() < 0.5 else -1.0) for e in ents}
        Tb = [dict(t, dy=a + s[t["entidad"]] * r) for t, a, r in zip(T, ajuste, res)]
        e = _ols(Tb)
        if e is not None:
            muestras.append(e[coef])
    muestras.sort()
    lo = muestras[int(0.025 * len(muestras))]
    hi = muestras[int(0.975 * len(muestras)) - 1]
    p = sum(1 for x in muestras if abs(x) >= abs(b0)) / len(muestras)
    return {"punto": b0, "ic95": [b0 - hi, b0 - lo], "p": p, "B": len(muestras),
            "n_conglomerados": len(ents), "p_minimo_alcanzable": 2.0 / (2 ** len(ents)),
            "valores_distintos_del_estadistico": len(set(round(x, 10) for x in muestras))}


def wild_cluster_contraste(T, B=B_BOOT, seed=SEED):
    """beta_int - beta_pres, con H0: beta_int = beta_pres. Bajo esa restriccion
    el modelo es  dy = alpha*hueco + beta*(ddp + ddi)."""
    import random
    rnd = random.Random(seed)
    full = _ols(T)
    if full is None:
        return None
    b0 = full["ddi"] - full["ddp"]
    Tr = [dict(t, suma=t["ddp"] + t["ddi"]) for t in T]
    h0 = _ols(Tr, ("hueco", "suma"))
    ajuste = [h0["hueco"] * t["hueco"] + h0["suma"] * (t["ddp"] + t["ddi"]) for t in T]
    res = [t["dy"] - a for t, a in zip(T, ajuste)]
    ents = sorted({t["entidad"] for t in T})
    muestras = []
    for _ in range(B):
        s = {e: (1.0 if rnd.random() < 0.5 else -1.0) for e in ents}
        Tb = [dict(t, dy=a + s[t["entidad"]] * r) for t, a, r in zip(T, ajuste, res)]
        e = _ols(Tb)
        if e is not None:
            muestras.append(e["ddi"] - e["ddp"])
    muestras.sort()
    lo = muestras[int(0.025 * len(muestras))]
    hi = muestras[int(0.975 * len(muestras)) - 1]
    return {"punto": b0, "ic95": [b0 - hi, b0 - lo],
            "p": sum(1 for x in muestras if abs(x) >= abs(b0)) / len(muestras),
            "B": len(muestras), "n_conglomerados": len(ents)}


def boot_municipio(T, B=B_BOOT, seed=SEED):
    """IC95 de contraste: remuestreo con reemplazo de las transiciones."""
    import random
    rnd = random.Random(seed)
    n = len(T)
    ac = {"hueco": [], "ddp": [], "ddi": [], "contraste": []}
    for _ in range(B):
        Tb = [T[rnd.randrange(n)] for _ in range(n)]
        e = _ols(Tb)
        if e is None:
            continue
        for k in ("hueco", "ddp", "ddi"):
            ac[k].append(e[k])
        ac["contraste"].append(e["ddi"] - e["ddp"])
    out = {}
    for k, v in ac.items():
        v.sort()
        out[k] = [v[int(0.025 * len(v))], v[int(0.975 * len(v)) - 1]]
    return out


# ═════════════════════ §1.7.7 control de regresion sobre L6 ══════════════════
def control_regresion_l6():
    """Re-corre el estimador de MAESTRA34-L6 sobre el panel de L6 y exige que
    reproduzca data/l6-resultados-concurrencia-v1_0.json BYTE A BYTE. PARO si no."""
    import hashlib, subprocess
    ref = "data/l6-resultados-concurrencia-v1_0.json"
    r = subprocess.run([sys.executable, "tools/l6_estimador_concurrencia.py"],
                       capture_output=True, text=True)
    if r.returncode != 0:
        return {"PARO": True, "motivo": "l6_estimador_concurrencia.py no corrio",
                "stderr": r.stderr[-1500:]}
    obtenido = r.stdout
    esperado = open(ref, encoding="utf-8").read()
    h_obt = hashlib.sha256(obtenido.encode("utf-8")).hexdigest()
    h_esp = hashlib.sha256(esperado.encode("utf-8")).hexdigest()
    return {"identico_byte_a_byte": obtenido == esperado,
            "sha256_recorrida": h_obt, "sha256_archivada": h_esp,
            "bytes_recorrida": len(obtenido.encode("utf-8")),
            "bytes_archivada": len(esperado.encode("utf-8")),
            "PARO": obtenido != esperado}


# ═══════════════════════════ P2 · corrida completa ═══════════════════════════
def _cuantil(v, q):
    v = sorted(v)
    return v[min(len(v) - 1, max(0, int(q * len(v))))]


def corre():
    ctrl_l6 = control_regresion_l6()
    if ctrl_l6.get("PARO"):
        return {"PARO": True, "control_regresion_l6": ctrl_l6,
                "nota": "§1.7.7: el estimador de L6 no reproduce su JSON. No se mide nada mas."}

    P, ctrl_lectura = panel_l3()
    D = tratamiento_l3()
    y, fuera = participacion(P)
    univ, perdidos = universo(P)
    T = transiciones(y, univ, D)

    res = {"control_regresion_l6": ctrl_l6,
           "n_transiciones_municipio": len(T),
           "n_municipios_por_entidad": {e: len(univ[e]) for e in univ},
           "municipios_perdidos": perdidos,
           "fuera_de_rango": fuera,
           "controles_lectura": ctrl_lectura}

    est = _ols(T)
    res["estimador"] = {
        "alpha_pp_por_anio": est["hueco"],
        "beta_pres_pp": est["ddp"],
        "beta_int_pp": est["ddi"],
        "wild_cluster_alpha": wild_cluster(T, "hueco"),
        "wild_cluster_beta_pres": wild_cluster(T, "ddp"),
        "wild_cluster_beta_int": wild_cluster(T, "ddi"),
        "wild_cluster_contraste_int_menos_pres": wild_cluster_contraste(T),
        "ic95_bootstrap_municipio": boot_municipio(T)}

    # variante SIN alpha (§0.3): se reporta pase lo que pase
    est_sa = _ols(T, ("ddp", "ddi"))
    res["variante_sin_alpha"] = {"beta_pres_pp": est_sa["ddp"], "beta_int_pp": est_sa["ddi"]}

    # transiciones por clase, para saber que identifica que
    clases = {}
    for t in T:
        k = identifica(t["ddp"], t["ddi"])
        clases.setdefault(k, {"n_municipio_transicion": 0, "transiciones": set()})
        clases[k]["n_municipio_transicion"] += 1
        clases[k]["transiciones"].add(f'{t["entidad"]} {t["de"]}->{t["a"]}')
    res["identificacion_del_panel"] = {
        k: {"n_municipio_transicion": v["n_municipio_transicion"],
            "transiciones": sorted(v["transiciones"])} for k, v in clases.items()}
    res["n_transiciones_STAY"] = len({f'{t["entidad"]} {t["de"]}->{t["a"]}'
                                      for t in T if t["clase"] == "STAY"})

    # resumen por transicion
    res["por_transicion"] = []
    for ent, anios in SERIES_L3.items():
        for a, b in zip(anios, anios[1:]):
            sub = [t for t in T if t["entidad"] == ent and t["de"] == a]
            if not sub:
                continue
            res["por_transicion"].append({
                "entidad": ent, "de": a, "a": b, "hueco": b - a,
                "tipo_de": etiqueta_pata(a, D[(ent, a)] != (0, 0)),
                "tipo_a": etiqueta_pata(b, D[(ent, b)] != (0, 0)),
                "ddp": sub[0]["ddp"], "ddi": sub[0]["ddi"], "clase": sub[0]["clase"],
                "n": len(sub),
                "y_de_media": sum(t["y0"] for t in sub) / len(sub),
                "y_a_media": sum(t["y1"] for t in sub) / len(sub),
                "dy_media": sum(t["dy"] for t in sub) / len(sub),
                "dy_mediana": _cuantil([t["dy"] for t in sub], 0.5)})

    # ATT por cohorte (§1.7.1): cada SWITCH contra la alpha estimada SOLO de STAY
    stay = [t for t in T if t["clase"] == "STAY"]
    a_stay = (sum(t["dy"] for t in stay) / sum(t["hueco"] for t in stay)) if stay else None
    res["att_por_transicion"] = {"alpha_solo_STAY_pp_por_anio": a_stay,
                                 "n_transiciones_STAY_municipio": len(stay)}
    if a_stay is not None:
        for t0 in sorted({(t["entidad"], t["de"], t["a"]) for t in T if t["clase"] == "SWITCH"}):
            ent, a, b = t0
            sub = [t for t in T if t["entidad"] == ent and t["de"] == a and t["clase"] == "SWITCH"]
            media = sum(t["dy"] for t in sub) / len(sub)
            res["att_por_transicion"][f"{ent}_{a}_{b}"] = {
                "delta_bruto_pp": media, "hueco": b - a, "ddp": sub[0]["ddp"], "ddi": sub[0]["ddi"],
                "att_pp": media - a_stay * (b - a), "n": len(sub)}

    # agregados estatales
    res["agregado_estatal"] = []
    for (ent, anio), d in sorted(P.items()):
        if ent not in SERIES_L3 or anio not in SERIES_L3[ent]:
            continue
        sub = {m: d[m] for m in univ[ent] if m in d}
        v = sum(x[0] for x in sub.values()); l = sum(x[1] for x in sub.values())
        res["agregado_estatal"].append({
            "entidad": ent, "anio": anio, "n": len(sub), "votos": v, "lista_nominal": l,
            "participacion_agregada_pp": 100.0 * v / l if l else None,
            "D_pres": D[(ent, anio)][0], "D_int": D[(ent, anio)][1]})

    # heterogeneidad por tamano (§1.7.4)
    tam = {}
    for ent, anios in SERIES_L3.items():
        base = P[(ent, anios[0])]
        for m in univ[ent]:
            if m in base:
                tam[(ent, m)] = base[m][1]
    vals = sorted(tam.values()); q1 = _cuantil(vals, 1/3); q2 = _cuantil(vals, 2/3)
    res["heterogeneidad_tamano"] = {"cortes_lista_nominal": [q1, q2]}
    for nom, cond in (("chico", lambda v: v <= q1), ("mediano", lambda v: q1 < v <= q2),
                      ("grande", lambda v: v > q2)):
        Ts = [t for t in T if cond(tam.get((t["entidad"], t["municipio"]), 0))]
        e = _ols(Ts) if len(Ts) > 3 else None
        res["heterogeneidad_tamano"][nom] = {"n": len(Ts)} | (
            {"alpha": e["hueco"], "beta_pres": e["ddp"], "beta_int": e["ddi"]} if e else {})

    # sensibilidades (§1.7.5)
    res["sensibilidad"] = {}
    for nom, filtro in (("sin_hueco_1", lambda t: t["hueco"] != 1),
                        ("sin_Coahuila", lambda t: t["entidad"] != "Coahuila"),
                        ("sin_Durango", lambda t: t["entidad"] != "Durango"),
                        ("solo_entidades_nuevas_de_L3",
                         lambda t: t["entidad"] in ("Baja California", "Chihuahua")),
                        ("solo_panel_de_L6",
                         lambda t: t["entidad"] in ("Coahuila", "Nayarit", "Zacatecas", "Durango"))):
        Ts = [t for t in T if filtro(t)]
        e = _ols(Ts) if len(Ts) > 3 else None
        res["sensibilidad"][nom] = {"n": len(Ts), "entidades": sorted({t["entidad"] for t in Ts})} | (
            {"alpha": e["hueco"], "beta_pres": e["ddp"], "beta_int": e["ddi"]} if e else
            {"nota": "no identificado con este subconjunto"})

    # descomposicion del Delta de MAESTRA34-L4 (§1.8)
    res["descomposicion_L4"] = {
        "delta_L4_pp": 10.4790,
        "componentes_para_2023_no_conc_-> 2024_presidencial_hueco_1": {
            "alpha_x_hueco": est["hueco"] * 1, "beta_pres": est["ddp"],
            "suma_explicada_pp": est["hueco"] * 1 + est["ddp"]}}
    return res

if __name__ == "__main__":
    main()
