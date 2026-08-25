#!/usr/bin/env python3
"""Ejecuta inspección estructural mínima cegada de activos locales T0."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import struct
import subprocess
import zipfile
from collections import Counter
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

import olefile
import openpyxl

try:
    from .barrido2_material import inspect_task, materialize_tasks
    from .pdf_extract import extract_pdf
except ImportError:
    from barrido2_material import inspect_task, materialize_tasks
    from pdf_extract import extract_pdf


FORBIDDEN_TASK_FIELDS = {
    "necesidad_id", "relacion_id", "adjudicacion_vigente", "constructo_esperado",
    "objeto_modelo_origen", "decision_humana_pendiente", "hipotesis",
    "interpretacion_deseada", "resultado_favorable", "signo_esperado",
    "clasificacion_supervisor",
}


def stable_id(prefix: str, *parts: str, size: int = 24) -> str:
    return prefix + hashlib.sha256("\x1f".join(parts).encode("utf-8")).hexdigest()[:size]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, fields: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", lineterminator="\n", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def decode_header(payload: bytes) -> tuple[str, list[str]]:
    candidates = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
    text = ""
    encoding = "NO_DETERMINADO"
    for candidate in candidates:
        try:
            text = payload.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    first = text.splitlines()[0] if text.splitlines() else ""
    delimiter = max((",", "\t", ";", "|"), key=first.count) if first else ","
    try:
        fields = next(csv.reader([first], delimiter=delimiter)) if first else []
    except csv.Error:
        fields = []
    return encoding, [value.strip() for value in fields]


def parse_dbf_fields(payload: bytes) -> tuple[int | None, list[str]]:
    if len(payload) < 32:
        return None, []
    records = struct.unpack("<I", payload[4:8])[0]
    header_size = struct.unpack("<H", payload[8:10])[0]
    fields: list[str] = []
    cursor = 32
    while cursor + 32 <= min(header_size, len(payload)) and payload[cursor] != 0x0D:
        name = payload[cursor:cursor + 11].split(b"\x00", 1)[0].decode("ascii", errors="replace")
        if name:
            fields.append(name)
        cursor += 32
    return records, fields


def inspect_zip(path: Path) -> tuple[dict[str, Any], list[dict[str, str]], str]:
    objects: list[dict[str, str]] = []
    with zipfile.ZipFile(path) as archive:
        members = archive.infolist()
        supported_compression = {
            zipfile.ZIP_STORED, zipfile.ZIP_DEFLATED,
            getattr(zipfile, "ZIP_BZIP2", 12), getattr(zipfile, "ZIP_LZMA", 14),
        }
        unsupported_methods = sorted({info.compress_type for info in members if info.compress_type not in supported_compression})
        bad = archive.testzip() if not unsupported_methods else None
        extensions = Counter(Path(info.filename).suffix.lower() or "(sin_extension)" for info in members if not info.is_dir())
        for info in members:
            if info.is_dir():
                continue
            variables: list[str] = []
            detail = ""
            suffix = Path(info.filename).suffix.lower()
            if info.compress_type not in supported_compression:
                detail = f"compresion_no_soportada_por_runtime={info.compress_type};contenido_no_abierto"
            elif suffix in {".csv", ".tsv", ".txt"}:
                with archive.open(info) as handle:
                    encoding, variables = decode_header(handle.read(65536))
                detail = f"encoding={encoding};campos={len(variables)}"
            elif suffix == ".dbf":
                with archive.open(info) as handle:
                    record_count, variables = parse_dbf_fields(handle.read(262144))
                detail = f"registros_declarados={record_count};campos={len(variables)}"
            objects.append({
                "objeto": info.filename,
                "tipo": suffix.lstrip(".").upper() or "SIN_EXTENSION",
                "tamano": str(info.file_size),
                "campos": json.dumps(variables, ensure_ascii=False, separators=(",", ":")),
                "detalle": detail,
            })
        listing = "\n".join(f"{info.filename}\t{info.file_size}\t{info.CRC}" for info in members)
        structure = {
            "tipo": "ZIP",
            "miembros_totales": len(members),
            "archivos": sum(not info.is_dir() for info in members),
            "directorios": sum(info.is_dir() for info in members),
            "extensiones": dict(sorted(extensions.items())),
            "sha256_listado_central": sha256_bytes(listing.encode("utf-8")),
            "primer_miembro_corrupto": bad or "NINGUNO",
            "integridad_contenido": "NO_DETERMINADO_COMPRESION_NO_SOPORTADA" if unsupported_methods else ("CORRUPTO" if bad else "SIN_ERROR_OBSERVADO"),
            "metodos_compresion_no_soportados": unsupported_methods,
            "miembros_con_campos_extraidos": sum(bool(json.loads(row["campos"])) for row in objects),
        }
        boundary = "Se abrió el contenedor completo y se enumeró su directorio central; solo se leyeron encabezados de miembros CSV/TSV/TXT/DBF con compresión soportada, no el contenido completo de los demás miembros."
        return structure, objects, boundary


def inspect_xlsx(path: Path) -> tuple[dict[str, Any], list[dict[str, str]], str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    objects: list[dict[str, str]] = []
    for sheet in workbook.worksheets:
        first_row: list[str] = []
        try:
            first_row = [str(value) if value is not None else "" for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        except StopIteration:
            pass
        objects.append({
            "objeto": sheet.title,
            "tipo": "HOJA_XLSX",
            "tamano": f"filas={sheet.max_row};columnas={sheet.max_column}",
            "campos": json.dumps(first_row, ensure_ascii=False, separators=(",", ":")),
            "detalle": "encabezado_primera_fila",
        })
    structure = {"tipo": "XLSX", "hojas": len(objects), "nombres_hojas": [row["objeto"] for row in objects]}
    workbook.close()
    return structure, objects, "Se enumeraron todas las hojas, dimensiones declaradas y primera fila; no se leyeron todas las celdas."


def biff_sheet_names(payload: bytes) -> list[str]:
    names: list[str] = []
    cursor = 0
    while cursor + 4 <= len(payload):
        record_type, length = struct.unpack("<HH", payload[cursor:cursor + 4])
        data = payload[cursor + 4:cursor + 4 + length]
        if record_type == 0x0085 and len(data) >= 8:
            count, flags = data[6], data[7]
            name_data = data[8:8 + count * (2 if flags & 0x01 else 1)]
            names.append(name_data.decode("utf-16le" if flags & 0x01 else "latin-1", errors="replace"))
        cursor += 4 + length
    return names


def inspect_xls(path: Path) -> tuple[dict[str, Any], list[dict[str, str]], str]:
    with olefile.OleFileIO(path) as document:
        streams = ["/".join(parts) for parts in document.listdir(streams=True, storages=False)]
        workbook_stream = "Workbook" if document.exists("Workbook") else "Book" if document.exists("Book") else ""
        names = biff_sheet_names(document.openstream(workbook_stream).read()) if workbook_stream else []
        metadata = document.get_metadata()
        objects = [{"objeto": name, "tipo": "HOJA_XLS", "tamano": "NO_DETERMINADO", "campos": "[]", "detalle": "nombre BIFF BOUNDSHEET"} for name in names]
        structure = {
            "tipo": "XLS_OLE_BIFF",
            "streams": streams,
            "hojas": len(names),
            "nombres_hojas": names,
            "titulo": str(metadata.title or ""),
        }
    return structure, objects, "Se abrió el contenedor OLE y se enumeraron streams/hojas BIFF; no se decodificaron todas las celdas."


class StructuralHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.title = ""
        self._in_title = False
        self.tables = 0
        self.headings: list[str] = []
        self._heading_tag = ""
        self._heading_text: list[str] = []
        self.links = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "title": self._in_title = True
        if tag == "table": self.tables += 1
        if tag == "a": self.links += 1
        if re.fullmatch(r"h[1-6]", tag): self._heading_tag, self._heading_text = tag, []

    def handle_endtag(self, tag: str) -> None:
        if tag == "title": self._in_title = False
        if tag == self._heading_tag:
            value = " ".join("".join(self._heading_text).split())
            if value: self.headings.append(value)
            self._heading_tag, self._heading_text = "", []

    def handle_data(self, data: str) -> None:
        if self._in_title: self.title += data
        if self._heading_tag: self._heading_text.append(data)


def inspect_html(path: Path) -> tuple[dict[str, Any], list[dict[str, str]], str]:
    parser = StructuralHTMLParser()
    parser.feed(path.read_text(encoding="utf-8", errors="replace"))
    objects = [{"objeto": value, "tipo": "ENCABEZADO_HTML", "tamano": "NO_APLICA", "campos": "[]", "detalle": ""} for value in parser.headings]
    return {"tipo": "HTML", "titulo": " ".join(parser.title.split()), "tablas": parser.tables, "enlaces": parser.links, "encabezados": len(parser.headings)}, objects, "Se parseó la estructura HTML completa; no se siguieron enlaces ni se ejecutó JavaScript."


def inspect_csv(path: Path) -> tuple[dict[str, Any], list[dict[str, str]], str]:
    with path.open("rb") as handle:
        encoding, fields = decode_header(handle.read(65536))
    objects = [{"objeto": path.name, "tipo": "TABLA_DELIMITADA", "tamano": str(path.stat().st_size), "campos": json.dumps(fields, ensure_ascii=False, separators=(",", ":")), "detalle": f"encoding={encoding}"}]
    return {"tipo": "TABLA_DELIMITADA", "campos": len(fields), "encoding": encoding}, objects, "Se leyó el encabezado; no se recorrieron todas las filas."


def inspect_pdf(path: Path, pdf_mode: str = "union") -> tuple[dict[str, Any], list[dict[str, str]], str]:
    metadata: dict[str, str] = {}
    metadata_warning = "NINGUNA"
    try:
        info = subprocess.run(["pdfinfo", str(path)], capture_output=True, text=True, timeout=30)
        if info.returncode != 0:
            metadata_warning = f"PDFINFO_FALLO:rc={info.returncode}"
        else:
            for line in info.stdout.splitlines():
                if ":" in line:
                    key, value = line.split(":", 1)
                    metadata[key.strip()] = value.strip()
    except (OSError, subprocess.SubprocessError) as exc:
        metadata_warning = f"PDFINFO_NO_DISPONIBLE:{type(exc).__name__}"
    declared_pages = int(metadata.get("Pages", "0") or "0")
    selected = range(1, min(declared_pages, 5) + 1) if declared_pages else None
    extraction = extract_pdf(path, mode=pdf_mode, pages=selected)
    extracted_pages = len(extraction.pages)
    if len(extraction.pages) > 5:
        extraction = type(extraction)(
            extraction.pages[:5], extraction.page_numbers[:5],
            extraction.extractors, extraction.warnings,
        )
    pages = declared_pages or extracted_pages
    extraction_pages = len(extraction.pages)
    text = extraction.text
    lines = [" ".join(line.split()) for line in text.splitlines() if line.strip()]
    headings = [line for line in lines if len(line) <= 140 and (line.isupper() or re.match(r"^(\d+\.?|[IVX]+\.)\s", line))][:100]
    objects = [{"objeto": value, "tipo": "ENCABEZADO_PDF_OBSERVADO", "tamano": "NO_APLICA", "campos": "[]", "detalle": f"paginas_1_a_{extraction_pages}"} for value in headings]
    visible_warnings = list(extraction.warnings)
    if metadata_warning != "NINGUNA":
        visible_warnings.append(metadata_warning)
    structure = {"tipo": "PDF", "paginas": pages, "cifrado": metadata.get("Encrypted", "NO_DETERMINADO"), "titulo": metadata.get("Title", ""), "paginas_texto_inspeccionadas": extraction_pages, "encabezados_observados": len(headings), "extractores_texto": list(extraction.extractors), "advertencias_extraccion": visible_warnings}
    warning_text = "|".join(visible_warnings) if visible_warnings else "NINGUNA"
    boundary = f"Se abrió el PDF, se leyó metadato de {pages} páginas y texto de páginas 1-{extraction_pages}; páginas posteriores y elementos no textuales no se inspeccionaron semánticamente; advertencias={warning_text}."
    return structure, objects, boundary


def inspect_json(path: Path) -> tuple[dict[str, Any], list[dict[str, str]], str]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    keys = sorted(payload) if isinstance(payload, dict) else []
    length = len(payload) if hasattr(payload, "__len__") else 1
    objects = [{"objeto": str(key), "tipo": "CLAVE_JSON", "tamano": "NO_APLICA", "campos": "[]", "detalle": "nivel_raiz"} for key in keys]
    return {"tipo": "JSON", "tipo_raiz": type(payload).__name__, "elementos_raiz": length, "claves_raiz": keys}, objects, "Se parseó el JSON completo y se enumeró el nivel raíz; no se caracterizaron recursivamente todos los valores."


def inspect_xml(path: Path) -> tuple[dict[str, Any], list[dict[str, str]], str]:
    import xml.etree.ElementTree as ET
    root = ET.parse(path).getroot()
    children = list(root)
    objects = [{"objeto": child.tag, "tipo": "ELEMENTO_XML", "tamano": str(len(list(child))), "campos": json.dumps(sorted(child.attrib), ensure_ascii=False), "detalle": "hijo_raiz"} for child in children]
    return {"tipo": "XML", "raiz": root.tag, "hijos_raiz": len(children), "atributos_raiz": sorted(root.attrib)}, objects, "Se parseó el XML completo y se enumeró la raíz; no se interpretó semánticamente cada nodo descendiente."


def inspect_one(path: Path) -> tuple[dict[str, Any], list[dict[str, str]], str]:
    suffix = path.suffix.lower()
    if suffix == ".zip": return inspect_zip(path)
    if suffix == ".pdf": return inspect_pdf(path)
    if suffix == ".xlsx": return inspect_xlsx(path)
    if suffix == ".xls": return inspect_xls(path)
    if suffix == ".html": return inspect_html(path)
    if suffix in {".csv", ".tsv", ".txt"}: return inspect_csv(path)
    if suffix == ".json": return inspect_json(path)
    if suffix == ".xml": return inspect_xml(path)
    raise NotImplementedError(f"FORMATO_NO_SOPORTADO:{suffix or '(sin_extension)'}")


REPORT_FIELDS = [
    "reporte_id", "tarea_observacion_id", "activo_id", "objeto_logico_id", "afirmacion_tipo",
    "objeto_inspeccionado", "universo_inspeccionado", "metodo", "valor_o_descripcion",
    "evidencia_ref", "localizador", "limitacion", "bloqueo", "siguiente_objeto_no_inspeccionado",
]


def reusable_inspections(output_dir: Path, output_root: Path) -> dict[Path, dict[str, Any]]:
    """Carga solo inspecciones previas cuyo expediente y hashes siguen íntegros."""
    reusable: dict[Path, dict[str, Any]] = {}
    for state in read_tsv(output_dir / "estado-activos.tsv") if (output_dir / "estado-activos.tsv").is_file() else []:
        task_id = state.get("tarea_observacion_id", "")
        evidence_path = Path(state.get("evidencia", ""))
        if state.get("estado_descriptivo") != "INSPECCIONADO_ESTRUCTURALMENTE" or not evidence_path.is_file():
            continue
        matches = sorted(output_root.glob(f"RUN-*/inspector/{task_id}"))
        if len(matches) != 1:
            continue
        directory = matches[0]
        hashes_path = directory / "hashes.json"
        if not hashes_path.is_file():
            continue
        recorded = json.loads(hashes_path.read_text(encoding="utf-8"))
        if not isinstance(recorded, dict) or any(
            not (directory / name).is_file() or sha256_file(directory / name) != expected
            for name, expected in recorded.items()
        ):
            continue
        old_input = json.loads((directory / "input.json").read_text(encoding="utf-8"))
        if FORBIDDEN_TASK_FIELDS.intersection(old_input):
            continue
        reports = read_tsv(directory / "reporte-inspeccion.tsv")
        facts = [row for row in reports if row.get("afirmacion_tipo") == "HECHO_OBSERVADO"]
        if len(facts) != 1 or facts[0].get("evidencia_ref") != f"sha256:{sha256_file(evidence_path)}":
            continue
        try:
            structure = json.loads(facts[0]["valor_o_descripcion"])
        except json.JSONDecodeError:
            continue
        reusable[evidence_path.resolve()] = {
            "task_id": task_id,
            "structure": structure,
            "objects": read_tsv(directory / "objetos-observados.tsv"),
            "boundary": facts[0]["limitacion"],
            "worker_dir": directory.as_posix(),
        }
    return reusable


def execute(universe_path: Path, snapshot_path: Path, corpus_root: Path, output_dir: Path, output_root: Path) -> dict[str, int]:
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    snapshot_hash = snapshot["snapshot_t0_sha256"]
    run_id = "RUN-" + snapshot_hash[:16]
    universe = read_tsv(universe_path)
    plan: list[dict[str, str]] = []
    reports: list[dict[str, str]] = []
    states: list[dict[str, str]] = []
    exceptions: list[dict[str, str]] = []
    reusable = reusable_inspections(output_dir, output_root)
    reuse_rows: list[dict[str, str]] = []

    for asset in universe:
        acquired = asset["estado_adquisicion"] == "ADQUIRIDO"
        if not acquired:
            states.append({
                "activo_id": asset["activo_id"], "objeto_logico_id": asset["objeto_logico_id"],
                "estado_descriptivo": "DECLARADO_NO_ADQUIRIDO", "adquirido": "NO",
                "inspeccionable": "NO_DETERMINADO", "tarea_observacion_id": "NO_APLICA",
                "reporte_inspeccion_ref": "NO_APLICA", "excepcion_inspeccion_ref": "NO_APLICA",
                "evidencia": asset["url_localizador_principal"], "reserva": "No se descargó durante T0",
            })
            continue
        task_id = stable_id("TOBS-", snapshot_hash, asset["activo_id"])
        report_id = stable_id("RINS-", task_id)
        relative = asset["ruta_local"]
        path = corpus_root / relative
        task = {
            "tarea_observacion_id": task_id,
            "run_id": run_id,
            "snapshot_t0_sha256": snapshot_hash,
            "activo_id": asset["activo_id"],
            "objeto_logico_id": asset["objeto_logico_id"],
            "familia_logica_id": asset["familia_logica_id"],
            "rutas_localizadores": [str(path)],
            "objetos_a_abrir": [relative],
            "grado_inspeccion": "MINIMA_ESTRUCTURAL_NEUTRAL",
            "campos_descriptivos": ["formato", "estructura", "objetos", "campos_observables", "frontera"],
            "metodo": "INSPECTOR_FORMATO_DETERMINISTA",
            "criterio_parada": "estructura enumerada y frontera declarada, o excepción material documentada",
        }
        forbidden = FORBIDDEN_TASK_FIELDS.intersection(task)
        if forbidden:
            raise AssertionError(f"task no cegada: {sorted(forbidden)}")
        worker_dir = output_root / run_id / "inspector" / task_id
        worker_dir.mkdir(parents=True, exist_ok=True)
        (worker_dir / "input.json").write_text(json.dumps(task, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        plan_row = {
            "tarea_observacion_id": task_id, "run_id": run_id, "snapshot_t0_sha256": snapshot_hash,
            "activo_id": asset["activo_id"], "objeto_logico_id": asset["objeto_logico_id"],
            "familia_logica_id": asset["familia_logica_id"], "ruta_localizador": str(path),
            "grado_inspeccion": "MINIMA_ESTRUCTURAL_NEUTRAL", "metodo": "INSPECTOR_FORMATO_DETERMINISTA",
            "criterio_parada": task["criterio_parada"], "estado_plan": "PENDIENTE",
        }
        worker_reports: list[dict[str, str]] = []
        objects: list[dict[str, str]] = []
        status = ""
        exception_id = "NO_APLICA"
        try:
            reused = reusable.get(path.resolve())
            if reused and asset["hash_local"] == sha256_file(path):
                structure = reused["structure"]
                objects = reused["objects"]
                boundary = reused["boundary"]
                reuse_rows.append({
                    "tarea_observacion_id": task_id,
                    "tarea_origen_id": reused["task_id"],
                    "activo_id": asset["activo_id"],
                    "ruta_local": relative,
                    "hash_local": asset["hash_local"],
                    "join_reutilizacion": "RUTA_LOCAL+SHA256+EXPEDIENTE_HASH_VALIDO",
                    "expediente_origen": reused["worker_dir"],
                })
            else:
                structure, objects, boundary = inspect_one(path)
            evidence = f"sha256:{sha256_file(path)}"
            worker_reports.extend([
                {
                    "reporte_id": report_id, "tarea_observacion_id": task_id, "activo_id": asset["activo_id"], "objeto_logico_id": asset["objeto_logico_id"],
                    "afirmacion_tipo": "HECHO_OBSERVADO", "objeto_inspeccionado": relative,
                    "universo_inspeccionado": "representación local completa al nivel estructural especificado",
                    "metodo": "apertura con parser determinista según formato", "valor_o_descripcion": json.dumps(structure, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
                    "evidencia_ref": evidence, "localizador": str(path), "limitacion": boundary,
                    "bloqueo": "NINGUNO", "siguiente_objeto_no_inspeccionado": "contenido más allá de la frontera declarada",
                },
                {
                    "reporte_id": report_id, "tarea_observacion_id": task_id, "activo_id": asset["activo_id"], "objeto_logico_id": asset["objeto_logico_id"],
                    "afirmacion_tipo": "NO_DETERMINADO", "objeto_inspeccionado": relative,
                    "universo_inspeccionado": "estructura y metadatos observables de la representación local",
                    "metodo": "inspección estructural sin adjudicación", "valor_o_descripcion": "unidad de observación y significado semántico no determinados por inspección mínima",
                    "evidencia_ref": evidence, "localizador": str(path), "limitacion": "No se infiere unidad desde nombres de archivo o campo",
                    "bloqueo": "NINGUNO", "siguiente_objeto_no_inspeccionado": "documentación semántica o contenido completo, si una orden superior lo solicita",
                },
                {
                    "reporte_id": report_id, "tarea_observacion_id": task_id, "activo_id": asset["activo_id"], "objeto_logico_id": asset["objeto_logico_id"],
                    "afirmacion_tipo": "NO_INSPECCIONADO", "objeto_inspeccionado": relative,
                    "universo_inspeccionado": "frontera exacta descrita en limitación", "metodo": "registro explícito de frontera",
                    "valor_o_descripcion": boundary, "evidencia_ref": evidence, "localizador": str(path),
                    "limitacion": boundary, "bloqueo": "NINGUNO", "siguiente_objeto_no_inspeccionado": "contenido fuera de la frontera",
                },
            ])
            status = "INSPECCIONADO_ESTRUCTURALMENTE"
            plan_row["estado_plan"] = "COMPLETADO"
        except (zipfile.BadZipFile, ValueError, OSError, EOFError) as exc:
            cause = "CORRUPCION" if isinstance(exc, (zipfile.BadZipFile, EOFError, ValueError)) else "FORMATO_NO_SOPORTADO"
            exception_id = stable_id("EXIN-", task_id, cause)
            exceptions.append({
                "excepcion_inspeccion_id": exception_id, "objeto_logico_id": asset["objeto_logico_id"],
                "causa": cause, "evidencia": f"{type(exc).__name__}:{str(exc)[:500]}",
                "intentos": "1 parser de formato + validación estructural", "autoridad": "SUPERVISOR_REGLA_MATERIAL",
                "estado": "ABIERTA", "accion_futura": "obtener representación íntegra o parser compatible",
            })
            status = "CORRUPTO" if cause == "CORRUPCION" else "FORMATO_NO_SOPORTADO"
            plan_row["estado_plan"] = "EXCEPCION_MATERIAL"
            worker_reports.append({
                "reporte_id": report_id, "tarea_observacion_id": task_id, "activo_id": asset["activo_id"], "objeto_logico_id": asset["objeto_logico_id"],
                "afirmacion_tipo": "NO_ACCESIBLE", "objeto_inspeccionado": relative,
                "universo_inspeccionado": "intento de apertura estructural de la representación local", "metodo": "parser de formato",
                "valor_o_descripcion": f"{type(exc).__name__}:{str(exc)[:500]}", "evidencia_ref": f"sha256:{sha256_file(path)}",
                "localizador": str(path), "limitacion": "No fue posible enumerar estructura completa", "bloqueo": cause,
                "siguiente_objeto_no_inspeccionado": relative,
            })
        except NotImplementedError as exc:
            exception_id = stable_id("EXIN-", task_id, "FORMATO_NO_SOPORTADO")
            exceptions.append({
                "excepcion_inspeccion_id": exception_id, "objeto_logico_id": asset["objeto_logico_id"],
                "causa": "FORMATO_NO_SOPORTADO", "evidencia": str(exc), "intentos": "1 detección por extensión",
                "autoridad": "SUPERVISOR_REGLA_MATERIAL", "estado": "ABIERTA", "accion_futura": "incorporar parser compatible",
            })
            status, plan_row["estado_plan"] = "FORMATO_NO_SOPORTADO", "EXCEPCION_MATERIAL"
            worker_reports.append({
                "reporte_id": report_id, "tarea_observacion_id": task_id, "activo_id": asset["activo_id"], "objeto_logico_id": asset["objeto_logico_id"],
                "afirmacion_tipo": "NO_ACCESIBLE", "objeto_inspeccionado": relative, "universo_inspeccionado": "detección de formato",
                "metodo": "extensión y dispatch de parser", "valor_o_descripcion": str(exc), "evidencia_ref": f"sha256:{sha256_file(path)}",
                "localizador": str(path), "limitacion": "estructura no abierta", "bloqueo": "FORMATO_NO_SOPORTADO",
                "siguiente_objeto_no_inspeccionado": relative,
            })

        write_tsv(worker_dir / "reporte-inspeccion.tsv", REPORT_FIELDS, worker_reports)
        write_tsv(worker_dir / "objetos-observados.tsv", ["objeto", "tipo", "tamano", "campos", "detalle"], objects)
        summary = {"tarea_observacion_id": task_id, "estado": status, "afirmaciones": len(worker_reports), "objetos_observados": len(objects), "cegado": True}
        (worker_dir / "resumen.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        hashes = {name: sha256_file(worker_dir / name) for name in ("input.json", "reporte-inspeccion.tsv", "objetos-observados.tsv", "resumen.json")}
        (worker_dir / "hashes.json").write_text(json.dumps(hashes, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        plan.append(plan_row)
        reports.extend(worker_reports)
        states.append({
            "activo_id": asset["activo_id"], "objeto_logico_id": asset["objeto_logico_id"],
            "estado_descriptivo": status, "adquirido": "SI", "inspeccionable": "SI" if status == "INSPECCIONADO_ESTRUCTURALMENTE" else "EXCLUIDO_POR_EXCEPCION_MATERIAL",
            "tarea_observacion_id": task_id, "reporte_inspeccion_ref": report_id,
            "excepcion_inspeccion_ref": exception_id, "evidencia": str(path),
            "reserva": "Inspección mínima estructural; no adjudica contenido" if status == "INSPECCIONADO_ESTRUCTURALMENTE" else "Excepción material documentada",
        })

    write_tsv(output_dir / "estado-activos.tsv", ["activo_id", "objeto_logico_id", "estado_descriptivo", "adquirido", "inspeccionable", "tarea_observacion_id", "reporte_inspeccion_ref", "excepcion_inspeccion_ref", "evidencia", "reserva"], sorted(states, key=lambda row: row["activo_id"]))
    write_tsv(output_dir / "plan-inspeccion.tsv", ["tarea_observacion_id", "run_id", "snapshot_t0_sha256", "activo_id", "objeto_logico_id", "familia_logica_id", "ruta_localizador", "grado_inspeccion", "metodo", "criterio_parada", "estado_plan"], sorted(plan, key=lambda row: row["tarea_observacion_id"]))
    write_tsv(output_dir / "excepciones-inspeccion.tsv", ["excepcion_inspeccion_id", "objeto_logico_id", "causa", "evidencia", "intentos", "autoridad", "estado", "accion_futura"], sorted(exceptions, key=lambda row: row["excepcion_inspeccion_id"]))
    write_tsv(output_dir / "reportes-inspeccion.tsv", REPORT_FIELDS, sorted(reports, key=lambda row: (row["reporte_id"], row["afirmacion_tipo"])))
    write_tsv(output_dir / "objetos-observados-no-representados.tsv", ["objeto_observado_id", "tarea_observacion_id", "activo_id", "objeto_logico_id", "objeto_observado", "reporte_inspeccion_ref", "localizador", "descripcion_literal", "posible_necesidad", "razon_inferencia"], [])
    write_tsv(output_dir / "reutilizacion-inspecciones.tsv", [
        "tarea_observacion_id", "tarea_origen_id", "activo_id", "ruta_local",
        "hash_local", "join_reutilizacion", "expediente_origen",
    ], sorted(reuse_rows, key=lambda row: row["tarea_observacion_id"]))
    return {
        "activos_t0": len(universe), "activos_adquiridos": len(plan),
        "inspecciones_completadas": sum(row["estado_plan"] == "COMPLETADO" for row in plan),
        "inspecciones_reutilizadas": len(reuse_rows),
        "inspecciones_ejecutadas_nuevas": len(plan) - len(reuse_rows),
        "excepciones_materiales": len(exceptions), "afirmaciones": len(reports),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--barrido2-materialize", action="store_true")
    mode.add_argument("--barrido2-inspect", action="store_true")
    parser.add_argument("--contract", type=Path)
    parser.add_argument("--task-root", type=Path)
    parser.add_argument("--ledger", type=Path)
    parser.add_argument("--task", type=Path)
    parser.add_argument("--roots-config", type=Path)
    parser.add_argument("--staging-dir", type=Path)
    parser.add_argument("--staging-root", type=Path)
    parser.add_argument("--reuse-source-dir", type=Path)
    parser.add_argument("--universe", type=Path)
    parser.add_argument("--snapshot", type=Path)
    parser.add_argument("--corpus-root", type=Path)
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument("--output-root", type=Path)
    args = parser.parse_args()
    if args.barrido2_materialize:
        if not all((args.snapshot, args.contract, args.task_root, args.ledger)):
            parser.error("--barrido2-materialize requiere --snapshot, --contract, --task-root y --ledger")
        result = materialize_tasks(
            args.snapshot.resolve(), args.contract.resolve(), args.task_root.resolve(), args.ledger.resolve(),
            args.staging_root.resolve() if args.staging_root else None,
        )
        print(json.dumps({"ok": True, **result}, ensure_ascii=False, indent=2, sort_keys=True))
        return 0
    if args.barrido2_inspect:
        if not all((args.task, args.roots_config, args.contract, args.staging_dir)):
            parser.error("--barrido2-inspect requiere --task, --roots-config, --contract y --staging-dir")
        result = inspect_task(
            args.task.resolve(), args.roots_config.resolve(), args.contract.resolve(), args.staging_dir.resolve(),
            reuse_source_dir=args.reuse_source_dir.resolve() if args.reuse_source_dir else None,
        )
        print(json.dumps({"ok": True, **result}, ensure_ascii=False, indent=2, sort_keys=True))
        return 0
    if not all((args.universe, args.snapshot, args.corpus_root, args.output_dir, args.output_root)):
        parser.error("modo T0 histórico requiere --universe, --snapshot, --corpus-root, --output-dir y --output-root")
    result = execute(args.universe.resolve(), args.snapshot.resolve(), args.corpus_root.resolve(), args.output_dir.resolve(), args.output_root.resolve())
    print(json.dumps({"ok": True, **result}, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
