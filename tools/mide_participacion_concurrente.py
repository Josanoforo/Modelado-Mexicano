#!/usr/bin/env python3
"""ACTO MAESTRA34-L4 · P3 · COMMIT 2 — mide el Δ de participación municipal entre
una elección local NO concurrente (2023) y una CONCURRENTE con la federal (2024),
en Coahuila y el Estado de México.

Ejecuta, sin desviarse, la spec congelada en
`forense/notas/2026-09-02-MAESTRA34-L4-P3-spec.md` (commit 1), firmada por mesa
como `F232-b` sobre `FP-232`. Regla del modelo bajo prueba: `R7.1`; necesidad del
registro del curador: `N25`.

Determinista por construcción: `seed = 42` fijo, sin red, sin fecha, sin
`random` sin semilla. Lee sólo de `data/raw/electoral_local_2023_2024/`, de
`data/raw/ine_marco_geografico_electoral/` (lectura secundaria 1) y de
`data/raw/20240603_2005_PREP.zip` (misma lectura secundaria).

    python3 tools/mide_participacion_concurrente.py [--json <ruta>]
"""

import argparse
import io
import json
import os
import re
import sys
import unicodedata
import zipfile

import numpy as np
import openpyxl

RAIZ = "data/raw"
BASE = os.path.join(RAIZ, "electoral_local_2023_2024")
MGE = os.path.join(RAIZ, "ine_marco_geografico_electoral")
PREP_ZIP = os.path.join(RAIZ, "20240603_2005_PREP.zip")

SEED = 42
B_BOOT = 10_000

# ────────────────────────────── utilidades ──────────────────────────────


def desmojibake(s):
    """Los XLSX por casilla del IEC traen UTF-8 releído como mac-roman
    ('ACU√ëA' por 'ACUÑA'). Se deshace cuando es posible; si no, se deja igual."""
    if "√" not in s:
        return s
    try:
        return s.encode("mac_roman").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return s


def norm(s):
    """Nombre de municipio comparable: sin acentos, mayúsculas, sin puntuación,
    espacios colapsados. NO se le quitan dígitos: en Edomex el cruce es por
    ID_MUNICIPIO justamente porque en 2024 algunos nombres traen sufijo."""
    s = desmojibake(str(s)).strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def num(v):
    """Entero de una celda; None si la celda no es un conteo."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    t = str(v).strip().replace(",", "").replace(" ", "")
    if t in ("", "-", "N/A", "NA"):
        return None
    try:
        return int(float(t))
    except ValueError:
        return None


def cabecera(ws, buscada, limite=14):
    """Fila (1-indexada) donde vive la cabecera, localizada por una columna que
    debe estar en ella. Se busca en vez de fijarse porque no es la misma fila en
    los dos años del Edomex (6 en 2023, 8 en 2024) — medido en P1."""
    objetivo = norm(buscada)
    for r in range(1, limite + 1):
        for c in ws[r]:
            if c.value is not None and norm(c.value) == objetivo:
                return r
    raise SystemExit("PARO: no se halló la cabecera con la columna %r" % buscada)


def indice(ws, fila, nombres):
    cab = {norm(c.value): c.column for c in ws[fila] if c.value is not None}
    fuera = {}
    for k, alias in nombres.items():
        for a in alias:
            if norm(a) in cab:
                fuera[k] = cab[norm(a)]
                break
        else:
            raise SystemExit("PARO: falta la columna %s (%r) en la cabecera" % (k, alias))
    return fuera


# ─────────────────────── lectura de las tablas municipales ───────────────────────


def lee_coahuila(path, etiqueta):
    """Tablas del IEC: cabecera en la fila 6, columna `No.` = clave 1..38.
    Se conserva SÓLO lo que tiene `No.` entero en 1..38: eso excluye la fila
    agregada (`No.`=0, VMRE_VA_VPPP) y las 3 filas de nota al pie (`No.` es texto)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    f = cabecera(ws, "nom_mun")
    ix = indice(ws, f, {"no": ["No."], "mun": ["nom_mun"],
                        "votos": ["TOTAL VOTOS", "TOTAL"],
                        "ln": ["Lista Nominal"], "part": ["% PART"]})
    filas, excl = {}, []
    for r in range(f + 1, ws.max_row + 1):
        crudo_no = ws.cell(r, ix["no"]).value
        crudo_mun = ws.cell(r, ix["mun"]).value
        if crudo_no is None and crudo_mun is None:
            continue
        n = num(crudo_no)
        nombre = norm(crudo_mun) if crudo_mun is not None else ""
        if n is None or not (1 <= n <= 38):
            excl.append({"fila": r, "no": str(crudo_no), "nombre": str(crudo_mun),
                         "motivo": "clave `No.` fuera de 1..38 (agregado o nota al pie)"})
            continue
        filas[n] = {"clave": n, "nombre": nombre,
                    "votos": num(ws.cell(r, ix["votos"]).value),
                    "ln": num(ws.cell(r, ix["ln"]).value),
                    "part_publicada": ws.cell(r, ix["part"]).value}
    wb.close()
    return {"etiqueta": etiqueta, "hoja": ws.title, "fila_cabecera": f,
            "filas": filas, "excluidas": excl, "max_row": ws.max_row}


def lee_edomex(path, etiqueta):
    """Tablas del IEEM: cabecera localizada por ID_MUNICIPIO (fila 6 en 2023, 8
    en 2024). Se conserva SÓLO ID_MUNICIPIO entero en 1..125: eso excluye las 3
    filas que no son municipios (voto anticipado / prisión preventiva /
    extranjero), que en 2023 no traen clave numérica en ese rango."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    f = cabecera(ws, "ID_MUNICIPIO")
    ix = indice(ws, f, {"id": ["ID_MUNICIPIO"], "mun": ["MUNICIPIO"],
                        "votos": ["TOTAL_VOTOS"], "ln": ["LISTA_NOMINAL"]})
    filas, excl = {}, []
    for r in range(f + 1, ws.max_row + 1):
        crudo_id = ws.cell(r, ix["id"]).value
        crudo_mun = ws.cell(r, ix["mun"]).value
        if crudo_id is None and crudo_mun is None:
            continue
        n = num(crudo_id)
        if n is None or not (1 <= n <= 125):
            excl.append({"fila": r, "id": str(crudo_id), "nombre": str(crudo_mun),
                         "motivo": "ID_MUNICIPIO fuera de 1..125 (fila no municipal)"})
            continue
        filas[n] = {"clave": n, "nombre": norm(crudo_mun),
                    "votos": num(ws.cell(r, ix["votos"]).value),
                    "ln": num(ws.cell(r, ix["ln"]).value),
                    "part_publicada": None}
    wb.close()
    return {"etiqueta": etiqueta, "hoja": ws.title, "fila_cabecera": f,
            "filas": filas, "excluidas": excl, "max_row": ws.max_row}


# ───────────────────────────── estimación ─────────────────────────────


def participacion(d):
    v, ln = d["votos"], d["ln"]
    if v is None or ln is None or ln <= 0:
        return None
    return 100.0 * v / ln


def cruza(t23, t24, estado, llave):
    """Universo = municipios presentes en LOS DOS años con lista nominal > 0.
    Todo lo que no empata se devuelve nombrado, nunca descartado en silencio."""
    filas, faltan = [], {"solo_2023": [], "solo_2024": [], "sin_denominador": []}
    for k in sorted(set(t23["filas"]) | set(t24["filas"])):
        a, b = t23["filas"].get(k), t24["filas"].get(k)
        if a is None:
            faltan["solo_2024"].append({"clave": k, "nombre": b["nombre"]})
            continue
        if b is None:
            faltan["solo_2023"].append({"clave": k, "nombre": a["nombre"]})
            continue
        pa, pb = participacion(a), participacion(b)
        if pa is None or pb is None:
            faltan["sin_denominador"].append({"clave": k, "nombre": a["nombre"]})
            continue
        filas.append({"estado": estado, "clave": k, "nombre_2023": a["nombre"],
                      "nombre_2024": b["nombre"], "llave": llave,
                      "votos_2023": a["votos"], "ln_2023": a["ln"], "part_2023": pa,
                      "votos_2024": b["votos"], "ln_2024": b["ln"], "part_2024": pb,
                      "delta": pb - pa})
    return filas, faltan


def boot_media(x, seed=SEED, b=B_BOOT):
    """IC95 por percentiles, remuestreando MUNICIPIOS con reemplazo."""
    x = np.asarray(x, dtype=float)
    if x.size == 0:
        return None
    rng = np.random.default_rng(seed)
    idx = rng.integers(0, x.size, size=(b, x.size))
    medias = x[idx].mean(axis=1)
    return {"media": float(x.mean()), "mediana": float(np.median(x)),
            "de": float(x.std(ddof=1)) if x.size > 1 else None,
            "ic95": [float(np.percentile(medias, 2.5)),
                     float(np.percentile(medias, 97.5))],
            "n": int(x.size), "B": b, "seed": seed}


def boot_diferencia(xa, xb, seed=SEED, b=B_BOOT):
    """IC95 de (media de xa − media de xb), remuestreando cada grupo por separado."""
    xa, xb = np.asarray(xa, float), np.asarray(xb, float)
    rng = np.random.default_rng(seed)
    ia = rng.integers(0, xa.size, size=(b, xa.size))
    ib = rng.integers(0, xb.size, size=(b, xb.size))
    d = xa[ia].mean(axis=1) - xb[ib].mean(axis=1)
    return {"diferencia": float(xa.mean() - xb.mean()),
            "ic95": [float(np.percentile(d, 2.5)), float(np.percentile(d, 97.5))],
            "n_a": int(xa.size), "n_b": int(xb.size), "B": b, "seed": seed}


# ───────────────────────────── controles ─────────────────────────────


def reagrega_casilla(path, col_votos, col_ln, col_mun, filtro_dtto_loc=None):
    """Control 1: reconstruye el agregado municipal desde las actas de casilla."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    f = cabecera(ws, col_mun)
    ix = indice(ws, f, {"mun": [col_mun], "votos": [col_votos], "ln": [col_ln]})
    acc, actas = {}, 0
    for fila in ws.iter_rows(min_row=f + 1, values_only=True):
        nombre = fila[ix["mun"] - 1]
        if nombre is None:
            continue
        nombre = norm(nombre)
        if nombre in ("", "N A", "NA"):
            continue
        v, ln = num(fila[ix["votos"] - 1]), num(fila[ix["ln"] - 1])
        if v is None or ln is None:
            continue
        a = acc.setdefault(nombre, {"votos": 0, "ln": 0, "actas": 0})
        a["votos"] += v
        a["ln"] += ln
        a["actas"] += 1
        actas += 1
    wb.close()
    return acc, actas


def lee_crosswalk_edomex():
    wb = openpyxl.load_workbook(os.path.join(MGE, "CRFE_22032016_1aSO_P07_1_3.xlsx"),
                                read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    f = cabecera(ws, "SECCION", limite=12)
    ix = indice(ws, f, {"sec": ["SECCION"], "mun": ["MUNICIPIO"],
                        "nom": ["NOMBRE_MUNICIPIO"]})
    m = {}
    for fila in ws.iter_rows(min_row=f + 1, values_only=True):
        s, mu = num(fila[ix["sec"] - 1]), num(fila[ix["mun"] - 1])
        if s is None or mu is None:
            continue
        m[s] = (mu, norm(fila[ix["nom"] - 1] or ""))
    wb.close()
    return m


def lee_crosswalk_coahuila():
    import tempfile

    import py7zr
    # py7zr 1.1.3 no expone `read()`: se extrae el miembro a un directorio
    # temporal (nunca al corpus ni al worktree) y se abre desde ahi.
    with tempfile.TemporaryDirectory() as tmp:
        with py7zr.SevenZipFile(os.path.join(MGE, "CRFE_17052016_5aSE_P07_2.zip")) as a:
            objetivo = [n for n in a.getnames() if n.endswith("Anexo1_COAH.xlsx")][0]
            a.extract(path=tmp, targets=[objetivo])
        with open(os.path.join(tmp, objetivo), "rb") as fh:
            datos = fh.read()
    wb = openpyxl.load_workbook(io.BytesIO(datos), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    f = cabecera(ws, "SECCION", limite=12)
    ix = indice(ws, f, {"sec": ["SECCION"], "mun": ["MUNICIPIO"],
                        "nom": ["NOMBRE_MUNICIPIO"]})
    m = {}
    for fila in ws.iter_rows(min_row=f + 1, values_only=True):
        s, mu = num(fila[ix["sec"] - 1]), num(fila[ix["mun"] - 1])
        if s is None or mu is None:
            continue
        m[s] = (mu, norm(fila[ix["nom"] - 1] or ""))
    wb.close()
    return m


def prep_federal_por_seccion(entidades=(5, 15)):
    """Lectura secundaria 1: PREP presidencial 2024 por sección, para las dos
    entidades. Devuelve {entidad: {seccion: (votos, lista_nominal)}} sumando
    actas, y cuenta las que no aportan cifra."""
    z = zipfile.ZipFile(PREP_ZIP)
    z2 = zipfile.ZipFile(io.BytesIO(z.read("20240603_2005_PREP_PRES.zip")))
    with z2.open("PRES_2024.csv") as fh:
        crudo = fh.read().decode("utf-8-sig", errors="replace")
    lineas = re.split(r"\r\n|\n|\r", crudo)
    icab = next(i for i, l in enumerate(lineas) if l.startswith("CLAVE_CASILLA,"))
    cab = lineas[icab].split(",")
    c = {n: cab.index(n) for n in ("ID_ENTIDAD", "SECCION", "TOTAL_VOTOS_CALCULADO",
                                   "LISTA_NOMINAL")}
    fuera = {e: {} for e in entidades}
    leidas = aportan = 0
    for l in lineas[icab + 1:]:
        if not l:
            continue
        p = l.split(",")
        if len(p) <= max(c.values()):
            continue
        leidas += 1
        ent = num(p[c["ID_ENTIDAD"]])
        if ent not in fuera:
            continue
        sec = num(p[c["SECCION"]])
        v = num(p[c["TOTAL_VOTOS_CALCULADO"]])
        ln = num(p[c["LISTA_NOMINAL"]])
        if sec is None or v is None or ln is None:
            continue
        a = fuera[ent].setdefault(sec, [0, 0, 0])
        a[0] += v
        a[1] += ln
        a[2] += 1
        aportan += 1
    return fuera, leidas, aportan


# ────────────────────────────────── main ──────────────────────────────────


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", default=None)
    args = ap.parse_args()
    R = {"acto": "MAESTRA34-L4", "pieza": "P3", "seed": SEED, "B": B_BOOT}
    P = print

    P("=" * 92)
    P("ACTO MAESTRA34-L4 · P3 · COMMIT 2 — participación municipal, local no")
    P("concurrente (2023) vs local concurrente (2024). Spec:")
    P("forense/notas/2026-09-02-MAESTRA34-L4-P3-spec.md · firma de mesa F232-b")
    P("=" * 92)

    c23 = lee_coahuila(f"{BASE}/iec_coahuila_2023/Gubernatura2023_X_Municipio.xlsx",
                       "Coahuila 2023 gubernatura")
    c24 = lee_coahuila(f"{BASE}/iec_coahuila_2024/Ayuntamientos2024_X_Municipio.xlsx",
                       "Coahuila 2024 ayuntamientos")
    e23 = lee_edomex(f"{BASE}/ieem_edomex_2023/"
                     "RESULTADOS_DEFINITIVOS_GUBERNATURA_2023_POR_MUNICIPIO.xlsx",
                     "Edomex 2023 gubernatura")
    e24 = lee_edomex(f"{BASE}/ieem_edomex_2024/Resultados_definitivos_ayu_municipio.xlsx",
                     "Edomex 2024 ayuntamientos")

    P("\n§2.1 · LECTURA Y EXCLUSIONES (A.13 — se declara qué se examinó)")
    R["lectura"] = {}
    for t in (c23, c24, e23, e24):
        P("  %-30s hoja=%-24s cabecera=fila %d  max_row=%-5d  municipios=%-4d excluidas=%d"
          % (t["etiqueta"], t["hoja"], t["fila_cabecera"], t["max_row"],
             len(t["filas"]), len(t["excluidas"])))
        for x in t["excluidas"]:
            P("        excluida fila %-4s %-46s  %s"
              % (x["fila"], (x.get("nombre") or "")[:46], x["motivo"]))
        R["lectura"][t["etiqueta"]] = {"hoja": t["hoja"], "fila_cabecera": t["fila_cabecera"],
                                       "max_row": t["max_row"], "municipios": len(t["filas"]),
                                       "excluidas": t["excluidas"]}

    P("\n§2.2 · UNIVERSO (municipios en LOS DOS años, lista nominal > 0)")
    fc, ffc = cruza(c23, c24, "Coahuila", "nombre normalizado")
    fe, ffe = cruza(e23, e24, "Edomex", "ID_MUNICIPIO")
    # Coahuila se cruza por NOMBRE, según la spec: se comprueba que la clave `No.`
    # y el nombre den el MISMO emparejamiento, y si no, manda el nombre.
    desajuste = [x for x in fc if x["nombre_2023"] != x["nombre_2024"]]
    todo = fc + fe
    for est, filas, faltan, nominal in (("Coahuila", fc, ffc, 38), ("Edomex", fe, ffe, 125)):
        P("  %-9s universo efectivo %3d de %3d nominal" % (est, len(filas), nominal))
        for k, v in faltan.items():
            if v:
                P("      %s: %s" % (k, ", ".join("%s(%s)" % (i["nombre"], i["clave"]) for i in v)))
    P("  Coahuila, control de la llave: municipios cuyo nombre 2023 != nombre 2024 = %d %s"
      % (len(desajuste), [d["nombre_2023"] + " / " + d["nombre_2024"] for d in desajuste] or ""))
    R["universo"] = {"coahuila": {"efectivo": len(fc), "nominal": 38, "faltantes": ffc},
                     "edomex": {"efectivo": len(fe), "nominal": 125, "faltantes": ffe},
                     "coahuila_desajuste_nombre": len(desajuste),
                     "total": len(todo)}

    P("\n§2.3 · CONTROL DE RANGO (ninguna participación fuera de (0,100] pp)")
    malas = [x for x in todo if not (0 < x["part_2023"] <= 100 and 0 < x["part_2024"] <= 100)]
    P("  fuera de rango: %d" % len(malas))
    for x in malas[:10]:
        P("      %s %s: 2023=%.4f 2024=%.4f" % (x["estado"], x["nombre_2023"],
                                                x["part_2023"], x["part_2024"]))
    R["fuera_de_rango"] = [{"estado": x["estado"], "nombre": x["nombre_2023"],
                            "part_2023": x["part_2023"], "part_2024": x["part_2024"]}
                           for x in malas]
    if malas:
        P("  PARO por la §1.8.4 de la spec: hay participaciones fuera de (0,100].")

    P("\n§2.4 · CONTRASTE PRINCIPAL — Δ = participación 2024 − 2023, en pp")
    d_todo = [x["delta"] for x in todo]
    principal = boot_media(d_todo)
    P("  n municipios         : %d" % principal["n"])
    P("  media Δ              : %+.4f pp" % principal["media"])
    P("  mediana Δ            : %+.4f pp" % principal["mediana"])
    P("  desviación estándar  : %.4f pp" % principal["de"])
    P("  IC95 bootstrap       : [%+.4f, %+.4f] pp   (B=%d, seed=%d, percentiles)"
      % (principal["ic95"][0], principal["ic95"][1], principal["B"], principal["seed"]))
    cruza_cero = principal["ic95"][0] <= 0 <= principal["ic95"][1]
    P("  ¿el IC95 cruza cero? : %s" % ("SÍ" if cruza_cero else "NO"))
    R["principal"] = dict(principal, cruza_cero=bool(cruza_cero))

    P("\n§2.5 · DIAGNÓSTICO por estado")
    dc = [x["delta"] for x in fc]
    de_ = [x["delta"] for x in fe]
    bc, be = boot_media(dc), boot_media(de_)
    for nom, b in (("Coahuila", bc), ("Edomex", be)):
        P("  %-9s n=%-4d media Δ=%+.4f pp  mediana=%+.4f  IC95=[%+.4f, %+.4f]"
          % (nom, b["n"], b["media"], b["mediana"], b["ic95"][0], b["ic95"][1]))
    dif = boot_diferencia(de_, dc)
    P("  Edomex − Coahuila: %+.4f pp  IC95=[%+.4f, %+.4f]"
      % (dif["diferencia"], dif["ic95"][0], dif["ic95"][1]))
    R["diagnostico"] = {"coahuila": bc, "edomex": be, "edomex_menos_coahuila": dif}

    P("\n§2.6 · CONTROL 1 — reconstrucción del agregado municipal desde las actas")
    ctrl = {}
    for etq, path, cv, cl, tabla in (
        ("Coahuila 2023 gubernatura", f"{BASE}/iec_coahuila_2023/Gubernatura2023_X_Casilla.xlsx",
         "total_votos", "lista", c23),
        ("Coahuila 2024 ayuntamientos", f"{BASE}/iec_coahuila_2024/AYUNTAMIENTOS2024_X_CASILLA.xlsx",
         "TOTAL", "lista", c24),
    ):
        acc, actas = reagrega_casilla(path, cv, cl, "nom_mun")
        dv = dl = 0
        comparados = 0
        peor = None
        for m in tabla["filas"].values():
            a = acc.get(m["nombre"])
            if a is None:
                continue
            comparados += 1
            ev, el = a["votos"] - (m["votos"] or 0), a["ln"] - (m["ln"] or 0)
            dv += abs(ev)
            dl += abs(el)
            if peor is None or abs(ev) > abs(peor[1]):
                peor = (m["nombre"], ev, el)
        P("  %-28s actas=%-5d municipios comparados=%-4d  |Δvotos|=%-7d |Δlista|=%-7d  peor=%s"
          % (etq, actas, comparados, dv, dl, peor))
        ctrl[etq] = {"actas": actas, "municipios_comparados": comparados,
                     "abs_delta_votos": dv, "abs_delta_lista": dl,
                     "peor": None if peor is None else {"municipio": peor[0],
                                                        "delta_votos": peor[1],
                                                        "delta_lista": peor[2]}}
    R["control_reconstruccion"] = ctrl

    P("\n§2.7 · CONTROL 2 — %PART publicada por el IEC vs recalculada")
    ctrl2 = {}
    for etq, tabla in (("Coahuila 2023", c23), ("Coahuila 2024", c24)):
        difs = []
        for m in tabla["filas"].values():
            pub = m["part_publicada"]
            p = participacion(m)
            if pub is None or p is None:
                continue
            try:
                pubf = float(pub) * 100.0
            except (TypeError, ValueError):
                continue
            difs.append(abs(pubf - p))
        mx = max(difs) if difs else None
        P("  %-14s municipios con %%PART publicada=%-4d  diferencia máxima=%s pp"
          % (etq, len(difs), "%.6f" % mx if mx is not None else "n/a"))
        ctrl2[etq] = {"comparados": len(difs), "diferencia_maxima_pp": mx}
    R["control_part_publicada"] = ctrl2

    P("\n§2.8 · LECTURA SECUNDARIA 1 — participación FEDERAL 2024 llevada a municipio")
    P("  (el contraste principal ya NO la necesita: P1 trajo la mitad local de 2024)")
    try:
        cw = {5: lee_crosswalk_coahuila(), 15: lee_crosswalk_edomex()}
        prep, leidas, aportan = prep_federal_por_seccion()
        sec2 = {}
        for ent, nombre in ((5, "Coahuila"), (15, "Edomex")):
            secs = prep[ent]
            m = cw[ent]
            agg, sin_match, sec_sin = {}, 0, []
            for s, (v, ln, _n) in secs.items():
                if s not in m:
                    sin_match += 1
                    sec_sin.append(s)
                    continue
                k = m[s][0]
                x = agg.setdefault(k, [0, 0])
                x[0] += v
                x[1] += ln
            base = fc if ent == 5 else fe
            porclave = {}
            if ent == 15:
                porclave = {x["clave"]: x for x in base}
            else:
                nom2clave = {m[s][1]: m[s][0] for s in m}
                porclave = {}
                for x in base:
                    k = nom2clave.get(x["nombre_2023"])
                    if k is not None:
                        porclave[k] = x
            pares = []
            for k, (v, ln) in agg.items():
                x = porclave.get(k)
                if x is None or ln <= 0:
                    continue
                pares.append({"clave": k, "part_federal_2024": 100.0 * v / ln,
                              "part_local_2024": x["part_2024"],
                              "part_local_2023": x["part_2023"]})
            df_l23 = [p["part_federal_2024"] - p["part_local_2023"] for p in pares]
            df_l24 = [p["part_federal_2024"] - p["part_local_2024"] for p in pares]
            b1 = boot_media(df_l23) if df_l23 else None
            b2 = boot_media(df_l24) if df_l24 else None
            P("  %-9s secciones PREP=%-6d sin correspondencia en el crosswalk 2016=%-5d (%.2f%%)"
              % (nombre, len(secs), sin_match, 100.0 * sin_match / max(1, len(secs))))
            P("            municipios emparejados=%d" % len(pares))
            if b1:
                P("            Δ federal2024 − local2023 : %+.4f pp  IC95=[%+.4f, %+.4f]"
                  % (b1["media"], b1["ic95"][0], b1["ic95"][1]))
            if b2:
                P("            Δ federal2024 − local2024 : %+.4f pp  IC95=[%+.4f, %+.4f]"
                  % (b2["media"], b2["ic95"][0], b2["ic95"][1]))
            sec2[nombre] = {"secciones_prep": len(secs), "sin_match": sin_match,
                            "pct_sin_match": 100.0 * sin_match / max(1, len(secs)),
                            "municipios_emparejados": len(pares),
                            "delta_federal_vs_local2023": b1,
                            "delta_federal_vs_local2024": b2}
        sec2["_actas_prep_leidas"] = leidas
        sec2["_actas_prep_con_cifra"] = aportan
        R["secundaria_prep_federal"] = sec2
    except Exception as exc:                                   # noqa: BLE001
        P("  NO CORRIÓ: %r — se declara y se sigue (§1.7 de la spec)" % (exc,))
        R["secundaria_prep_federal"] = {"error": repr(exc)}

    P("\n§2.9 · LECTURA SECUNDARIA 2 — Coahuila, diputaciones locales 2023")
    try:
        acc, actas = reagrega_casilla(
            f"{BASE}/iec_coahuila_2023/Diputaciones2023_X_Casilla.xlsx",
            "total_votos", "lista", "nom_mun")
        pares = []
        for x in fc:
            a = acc.get(x["nombre_2023"])
            if a is None or a["ln"] <= 0:
                continue
            pares.append(x["part_2024"] - 100.0 * a["votos"] / a["ln"])
        b = boot_media(pares) if pares else None
        P("  actas=%d  municipios emparejados=%d" % (actas, len(pares)))
        if b:
            P("  Δ ayuntamiento2024 − diputaciones2023 : %+.4f pp  IC95=[%+.4f, %+.4f]"
              % (b["media"], b["ic95"][0], b["ic95"][1]))
        R["secundaria_diputaciones_coah"] = {"actas": actas, "emparejados": len(pares),
                                             "delta": b}
    except Exception as exc:                                   # noqa: BLE001
        P("  NO CORRIÓ: %r — se declara y se sigue" % (exc,))
        R["secundaria_diputaciones_coah"] = {"error": repr(exc)}

    P("\n§2.10 · FILAS MUNICIPALES (Δ ordenado)")
    for x in sorted(todo, key=lambda y: y["delta"]):
        P("  %-9s %-42s 2023=%7.3f  2024=%7.3f  Δ=%+8.3f"
          % (x["estado"], x["nombre_2023"][:42], x["part_2023"], x["part_2024"], x["delta"]))
    R["filas"] = todo

    if args.json:
        with open(args.json, "w", encoding="utf-8") as fh:
            json.dump(R, fh, ensure_ascii=False, indent=1, sort_keys=True)
        P("\nJSON: %s" % args.json)
    return 0


if __name__ == "__main__":
    sys.exit(main())
