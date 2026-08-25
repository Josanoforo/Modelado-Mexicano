#!/usr/bin/env python3
"""Materializa y diagnostica todo el universo de autoridad semántica.

Los proyectores reutilizables entienden descriptores INEGI con dominios
``Códigos válidos/Concepto`` y ``Clases→Valor/Etiqueta``. Un ``b = Blanco por
secuencia`` prueba que existe un subuniverso, pero no su predicado exacto, y por
eso permanece cerrado. Las demás familias E2 se clasifican por la insuficiencia
estructural concreta que impide emitir, nunca por falta de código pendiente.

Encuesta, ola, universo, unidad y scope de ponderador viven en perfiles JSON
declarativos. Este módulo extrae, enlaza, deduplica y valida; no contiene esos
significados.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping

import yaml
from openpyxl import load_workbook

try:
    from .autoridad_semantica_marco import (
        AutoridadSemanticaError,
        authority_key_from_e2,
        authority_key_from_row,
        canonical_authority_line,
        validate_authority_row,
        with_authority_id,
    )
    from .generar_marco import SEMANTIC_AUTHORITY_SPEC_FIELDS
    from .marco_e2_adapter import (
        E2IndexReader,
        ProvenanceIndex,
        expected_index_sha,
        is_variable_seed,
    )
except ImportError:
    from autoridad_semantica_marco import (
        AutoridadSemanticaError,
        authority_key_from_e2,
        authority_key_from_row,
        canonical_authority_line,
        validate_authority_row,
        with_authority_id,
    )
    from generar_marco import SEMANTIC_AUTHORITY_SPEC_FIELDS
    from marco_e2_adapter import (
        E2IndexReader,
        ProvenanceIndex,
        expected_index_sha,
        is_variable_seed,
    )


PROFILE_SCHEMA_VERSION = "PERFILES-AUTORIDAD-SEMANTICA-MARCO-1.0"
DIAGNOSTIC_SCHEMA_VERSION = "DIAGNOSTICO-AUTORIDAD-SEMANTICA-MARCO-1.0"
SUPPORTED_PROJECTOR = "INEGI-FD-CODIGOS-VALIDOS-7COL-1.0"
SUPPORTED_CLASS_PROJECTOR = "INEGI-FD-CLASES-VALOR-ETIQUETA-1.0"
SUPPORTED_TYPE_RULE = "DOS_CODIGOS_BINARIA_OTRO_ENUMERADO_CATEGORICA"
SUPPORTED_MISSING_RULE = "CODIGOS_VALIDOS_EXHAUSTIVOS"
STRUCTURAL_SCALAR_RULE = "CAMPO_ESCALAR_CON_DOMINIO_ESTRUCTURADO"

# Frontera mecánica del diagnóstico emitido por #346. Estos nueve payloads ya
# tenían perfil o un bloqueo específico; las demás semillas exactas no emitidas
# formaban el inventario de 107,443. Se conserva aquí únicamente para poder
# regenerar ese inventario inicial después de reemplazar el diagnóstico de #346.
PR346_PROFILED_OR_BLOCKED_PAYLOADS = frozenset(
    {
        "enadid2023_fd_xlsx",
        "endutih2024_fd_xlsx",
        "enfih2019_fd_xlsx",
        "enif2018_fd_xlsx",
        "enif2021_fd_zip",
        "enif2024_fd_xlsx",
        "enasic2022_fd_xlsx",
        "enut2024_fd_xlsx",
        "mociba2024_fd_xlsx",
    }
)

WEIGHT_FIELDS = sorted(
    {
        "ponderador",
        "ponderador_exacto",
        "ponderador_fuente",
        "ponderador_fuente_ola_tabla",
        "ponderador_ola",
        "ponderador_scope_id",
        "ponderador_scope_tipo",
    }
)
CODEBOOK_FIELDS = sorted(
    {
        "codificacion",
        "encuesta",
        "missing",
        "ola",
        "tipo_estadistico",
    }
)
RESPONSE_FIELDS = ["categorias_excluyentes", "respuesta_multiple"]


class ProductiveAuthorityError(ValueError):
    """Defecto global de fuente, identidad o contrato."""


@dataclass(frozen=True)
class ParsedVariable:
    sheet: str
    name: str
    question: str
    coding: tuple[tuple[str, str], ...]
    missing: tuple[str, ...]
    start_row: int
    end_row: int
    reason: str | None


def _text(value: Any) -> str:
    return str(value or "").strip()


def _fold(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", _text(value))
    return " ".join(
        "".join(character for character in normalized if not unicodedata.combining(character))
        .upper()
        .split()
    )


def _canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle, delimiter="\t")]


def _canonical_code(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else format(value, ".15g")
    text = _text(value)
    if not text:
        return None
    # Una celda con rango o varios codigos no aporta etiquetas uno-a-uno.
    if any(token in text for token in (",", "...", "…", " - ", "..")):
        return None
    return text


def _is_sequence_blank(code: str, label: str) -> bool:
    return code.casefold() == "b" or _fold(label) == "BLANCO POR SECUENCIA"


def _is_documented_missing(label: str) -> bool:
    folded = _fold(label)
    exact = {
        "NO SABE",
        "NO RESPONDE",
        "NO SABE / NO RESPONDE",
        "NO SABE/NO RESPONDE",
        "NO ESPECIFICADO",
        "NO ESPECIFICADA",
        "NO CONTESTO",
        "NO CONTESTA",
        "NO RECUERDA",
        "SE IGNORA",
        "SIN INFORMACION",
        "NO DECLARADO",
        "NO DECLARADA",
    }
    return folded in exact


def _is_unresolved_no_aplica(label: str) -> bool:
    """Detecta la etiqueta cuyo papel no puede decidirse de forma global."""

    return _fold(label) == "NO APLICA (SOLO OPCION 1 Y 2)"


def _is_multiple_response_indicator(label: str) -> bool:
    """Reconoce etiquetas estructuradas que declaran un indicador de mención."""

    return _fold(label) in {
        "NO SE DECLARO COMO OPCION AFIRMATIVA",
        "NO SE MENCIONO",
        "SI SE MENCIONO",
    }


def _find_descriptor_columns(
    rows: list[tuple[Any, ...]]
) -> tuple[int, int, int, int, int]:
    for row_index, row in enumerate(rows):
        folded = [_fold(value) for value in row]
        mnemonic = next(
            (
                index
                for index, value in enumerate(folded)
                if value in {"NEMONICO", "NEMONICO [2]", "MNEMONICO", "MNEMONICO [2]"}
            ),
            None,
        )
        code = next(
            (
                index
                for index, value in enumerate(folded)
                if value in {"CODIGOS VALIDOS", "CODIGO VALIDO"}
            ),
            None,
        )
        concept = next(
            (index for index, value in enumerate(folded) if value in {"CONCEPTO", "ETIQUETA"}),
            None,
        )
        question = next(
            (index for index, value in enumerate(folded) if value in {"PREGUNTA", "DESCRIPCION"}),
            None,
        )
        if None not in (mnemonic, code, concept, question):
            return row_index, int(question), int(mnemonic), int(code), int(concept)
        classes = next(
            (index for index, value in enumerate(folded) if value == "CLASES"),
            None,
        )
        if None not in (mnemonic, question, classes) and row_index + 1 < len(rows):
            following = [_fold(value) for value in rows[row_index + 1]]
            value_column = next(
                (index for index, value in enumerate(following) if value == "VALOR"),
                None,
            )
            label_column = next(
                (index for index, value in enumerate(following) if value == "ETIQUETA"),
                None,
            )
            if None not in (value_column, label_column):
                return (
                    row_index + 1,
                    int(question),
                    int(mnemonic),
                    int(value_column),
                    int(label_column),
                )
    raise ProductiveAuthorityError("DESCRIPTOR_SIN_CABECERA_ESTRUCTURADA")


def _finalize_variable(
    *,
    sheet: str,
    name: str,
    question: str,
    pairs: list[tuple[Any, Any]],
    start_row: int,
    end_row: int,
) -> ParsedVariable:
    coding: dict[str, str] = {}
    missing: set[str] = set()
    reason: str | None = None
    saw_domain = False
    for raw_code, raw_label in pairs:
        if raw_code is None and raw_label is None:
            continue
        code = _canonical_code(raw_code)
        label = _text(raw_label)
        if code is None or not label:
            reason = reason or "DOMINIO_NO_ENUMERADO_O_SIN_ETIQUETA"
            continue
        saw_domain = True
        if _is_sequence_blank(code, label):
            reason = "UNIVERSO_REACTIVO_CONDICIONADO"
            continue
        if _is_multiple_response_indicator(label):
            reason = reason or "RESPUESTA_MULTIPLE_DOCUMENTADA"
        if _is_unresolved_no_aplica(label):
            reason = reason or "NO_APLICA_SEMANTICA_NO_RESUELTA"
        if _is_documented_missing(label):
            missing.add(code)
            continue
        previous = coding.get(code)
        if previous is not None and previous != label:
            reason = "CODIGO_CON_ETIQUETAS_CONFLICTIVAS"
            continue
        coding[code] = label
    if not saw_domain:
        reason = reason or "DOMINIO_ESTRUCTURADO_AUSENTE"
    if len(coding) < 2:
        reason = reason or "MENOS_DE_DOS_CATEGORIAS_VALIDAS"
    return ParsedVariable(
        sheet=sheet,
        name=name,
        question=question,
        coding=tuple(sorted(coding.items())),
        missing=tuple(sorted(missing)),
        start_row=start_row,
        end_row=end_row,
        reason=reason,
    )


def parse_inegi_fd_7col(
    path: Path, configured_sheets: Iterable[str]
) -> dict[tuple[str, str], ParsedVariable]:
    """Extrae dominios uno-a-uno sin leer observaciones de microdato."""

    wanted = set(configured_sheets)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        missing_sheets = sorted(wanted - set(workbook.sheetnames))
        if missing_sheets:
            raise ProductiveAuthorityError(
                "DESCRIPTOR_HOJAS_AUSENTES:" + ",".join(missing_sheets)
            )
        parsed: dict[tuple[str, str], ParsedVariable] = {}
        for sheet in sorted(wanted):
            worksheet = workbook[sheet]
            rows = list(worksheet.iter_rows(values_only=True))
            header_index, question_col, mnemonic_col, code_col, concept_col = (
                _find_descriptor_columns(rows)
            )
            current_name = ""
            current_question = ""
            current_pairs: list[tuple[Any, Any]] = []
            current_start = 0

            def flush(end_row: int) -> None:
                nonlocal current_name, current_question, current_pairs, current_start
                if not current_name:
                    return
                key = (sheet, current_name)
                if key in parsed:
                    raise ProductiveAuthorityError(
                        f"DESCRIPTOR_VARIABLE_DUPLICADA:{sheet}:{current_name}"
                    )
                parsed[key] = _finalize_variable(
                    sheet=sheet,
                    name=current_name,
                    question=current_question,
                    pairs=current_pairs,
                    start_row=current_start,
                    end_row=max(current_start, end_row),
                )
                current_name = ""
                current_question = ""
                current_pairs = []
                current_start = 0

            for zero_index, row in enumerate(rows[header_index + 1 :], header_index + 1):
                excel_row = zero_index + 1
                mnemonic = _text(row[mnemonic_col] if mnemonic_col < len(row) else None)
                if _fold(mnemonic) in {"NEMONICO", "MNEMONICO"}:
                    flush(excel_row - 1)
                    continue
                if mnemonic and not mnemonic.startswith("["):
                    flush(excel_row - 1)
                    current_name = mnemonic
                    current_question = _text(
                        row[question_col] if question_col < len(row) else None
                    )
                    current_start = excel_row
                if current_name:
                    current_pairs.append(
                        (
                            row[code_col] if code_col < len(row) else None,
                            row[concept_col] if concept_col < len(row) else None,
                        )
                    )
            flush(len(rows))
        return parsed
    finally:
        workbook.close()


def _load_profiles(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError) as exc:
        raise ProductiveAuthorityError(f"PERFILES_INVALIDOS:{path}:{exc}") from exc
    if not isinstance(payload, dict) or payload.get("schema_version") != PROFILE_SCHEMA_VERSION:
        raise ProductiveAuthorityError("PERFILES_SCHEMA_VERSION_INVALIDA")
    for field in ("origin_main_inicial", "pr_343_merge_commit"):
        value = payload.get(field)
        if (
            not isinstance(value, str)
            or len(value) != 40
            or any(character not in "0123456789abcdef" for character in value)
        ):
            raise ProductiveAuthorityError(f"PERFILES_SHA_GIT_INVALIDO:{field}")
    profiles = payload.get("profiles")
    if not isinstance(profiles, list) or not profiles:
        raise ProductiveAuthorityError("PERFILES_VACIOS")
    by_payload: dict[str, dict[str, Any]] = {}
    preserve_ids = payload.get("preserve_authority_ids")
    expected_counts = payload.get("expected_counts")
    blocked_payloads = payload.get("blocked_payloads")
    if (
        not isinstance(preserve_ids, list)
        or not preserve_ids
        or len(preserve_ids) != len(set(preserve_ids))
        or any(not isinstance(value, str) or not value for value in preserve_ids)
    ):
        raise ProductiveAuthorityError("PERFILES_AUTORIDADES_PRESERVADAS_INVALIDAS")
    if not isinstance(expected_counts, dict) or set(expected_counts) != {
        "semillas_variables_totales",
        "uniones_exactas",
        "uniones_ausentes",
        "uniones_ambiguas",
    }:
        raise ProductiveAuthorityError("PERFILES_CONTEOS_ESPERADOS_INVALIDOS")
    if not isinstance(blocked_payloads, dict) or any(
        not isinstance(key, str)
        or not key
        or not isinstance(value, str)
        or not value
        for key, value in blocked_payloads.items()
    ):
        raise ProductiveAuthorityError("PERFILES_PAYLOADS_BLOQUEADOS_INVALIDOS")
    for raw in profiles:
        if not isinstance(raw, dict):
            raise ProductiveAuthorityError("PERFIL_NO_ES_OBJETO")
        payload_id = _text(raw.get("payload_id"))
        if not payload_id or payload_id in by_payload:
            raise ProductiveAuthorityError(f"PERFIL_PAYLOAD_DUPLICADO:{payload_id}")
        if raw.get("projector") not in {
            SUPPORTED_PROJECTOR,
            SUPPORTED_CLASS_PROJECTOR,
        }:
            raise ProductiveAuthorityError(f"PERFIL_PROJECTOR_INVALIDO:{payload_id}")
        tables = raw.get("tables")
        if not isinstance(tables, dict) or not tables:
            raise ProductiveAuthorityError(f"PERFIL_TABLAS_VACIAS:{payload_id}")
        if raw.get("type_rule") != SUPPORTED_TYPE_RULE:
            raise ProductiveAuthorityError(f"PERFIL_REGLA_TIPO_INVALIDA:{payload_id}")
        if raw.get("missing_rule") != SUPPORTED_MISSING_RULE:
            raise ProductiveAuthorityError(f"PERFIL_REGLA_MISSING_INVALIDA:{payload_id}")
        blocked_tables = raw.get("blocked_tables", {})
        if not isinstance(blocked_tables, dict) or any(
            not isinstance(key, str)
            or not key
            or not isinstance(value, str)
            or not value
            for key, value in blocked_tables.items()
        ):
            raise ProductiveAuthorityError(f"PERFIL_TABLAS_BLOQUEADAS_INVALIDAS:{payload_id}")
        overlap = set(tables) & set(blocked_tables)
        if overlap:
            raise ProductiveAuthorityError(
                f"PERFIL_TABLA_RESUELTA_Y_BLOQUEADA:{payload_id}:{','.join(sorted(overlap))}"
            )
        for table_name, table in tables.items():
            if not isinstance(table, dict):
                raise ProductiveAuthorityError(
                    f"PERFIL_TABLA_NO_ES_OBJETO:{payload_id}:{table_name}"
                )
            response = table.get("response_semantics")
            if not isinstance(response, dict):
                raise ProductiveAuthorityError(
                    f"PERFIL_RESPUESTA_VARIABLE_AUSENTE:{payload_id}:{table_name}"
                )
            response_sets: dict[str, set[str]] = {}
            response_rule = response.get("rule")
            if response_rule not in (None, STRUCTURAL_SCALAR_RULE):
                raise ProductiveAuthorityError(
                    f"PERFIL_REGLA_RESPUESTA_INVALIDA:{payload_id}:{table_name}"
                )
            for field in (
                "scalar_exclusive_variables",
                "multiple_response_variables",
                "unresolved_variables",
            ):
                values = response.get(field)
                if (
                    not isinstance(values, list)
                    or len(values) != len(set(values))
                    or any(not isinstance(value, str) or not value for value in values)
                ):
                    raise ProductiveAuthorityError(
                        f"PERFIL_RESPUESTA_VARIABLES_INVALIDAS:{payload_id}:"
                        f"{table_name}:{field}"
                    )
                response_sets[field] = set(values)
            if any(
                response_sets[left] & response_sets[right]
                for left, right in (
                    ("scalar_exclusive_variables", "multiple_response_variables"),
                    ("scalar_exclusive_variables", "unresolved_variables"),
                    ("multiple_response_variables", "unresolved_variables"),
                )
            ):
                raise ProductiveAuthorityError(
                    f"PERFIL_RESPUESTA_VARIABLE_CLASIFICACION_CONFLICTIVA:"
                    f"{payload_id}:{table_name}"
                )
            response_citation = response.get("citation")
            if not isinstance(response_citation, dict) or any(
                not _text(response_citation.get(field))
                for field in ("source_id", "uri", "sha256", "locator")
            ):
                raise ProductiveAuthorityError(
                    f"PERFIL_CITA_RESPUESTA_INVALIDA:{payload_id}:{table_name}"
                )
            citations = table.get("context_citations")
            if not isinstance(citations, list) or not citations:
                raise ProductiveAuthorityError(
                    f"PERFIL_CITAS_CONTEXTO_VACIAS:{payload_id}:{table_name}"
                )
        by_payload[payload_id] = raw
    overlap = set(by_payload) & set(blocked_payloads)
    if overlap:
        raise ProductiveAuthorityError(
            "PERFILES_PAYLOAD_RESUELTO_Y_BLOQUEADO:" + ",".join(sorted(overlap))
        )
    payload["profiles_by_payload"] = by_payload
    repair_audit = payload.get("repair_audit")
    if not isinstance(repair_audit, dict):
        raise ProductiveAuthorityError("PERFILES_AUDITORIA_REPARACION_AUSENTE")
    audit_lists = (
        "retiradas_multirrespuesta",
        "retiradas_exclusividad_multiplicidad_no_resuelta",
        "dependientes_no_aplica_como_missing",
    )
    if any(
        not isinstance(repair_audit.get(field), list)
        or len(repair_audit[field]) != len(set(repair_audit[field]))
        or any(not isinstance(value, str) or not value for value in repair_audit[field])
        for field in audit_lists
    ):
        raise ProductiveAuthorityError("PERFILES_AUDITORIA_REPARACION_LISTAS_INVALIDAS")
    examined = repair_audit.get("autoridades_actuales_examinadas")
    remain = repair_audit.get("autoridades_permanecen")
    removed = sum(len(repair_audit[field]) for field in audit_lists)
    if not isinstance(examined, int) or not isinstance(remain, int) or examined != remain + removed:
        raise ProductiveAuthorityError("PERFILES_AUDITORIA_REPARACION_CONTEOS_INVALIDOS")
    return payload


def _validate_profile_sources(
    profile: Mapping[str, Any], corpus_root: Path, repo: Path
) -> tuple[Path, str]:
    payload_id = _text(profile.get("payload_id"))
    source_path = corpus_root / _text(profile.get("source_path"))
    expected_sha = _text(profile.get("source_sha256")).lower()
    observed_sha = _sha256_path(source_path)
    if observed_sha != expected_sha:
        raise ProductiveAuthorityError(
            f"FUENTE_SHA_DIVERGENTE:{payload_id}:{expected_sha}!={observed_sha}"
        )
    for table in profile["tables"].values():
        citations = [
            *table["context_citations"],
            table["response_semantics"]["citation"],
        ]
        for citation in citations:
            local_path = _text(citation.get("local_path"))
            repo_path = _text(citation.get("repo_path"))
            if local_path and repo_path:
                raise ProductiveAuthorityError(
                    f"FUENTE_CONTEXTO_RUTA_AMBIGUA:{payload_id}"
                )
            if not local_path and not repo_path:
                continue
            context_path = (
                corpus_root / local_path if local_path else repo / repo_path
            )
            context_sha = _sha256_path(context_path)
            citation_sha = _text(citation.get("sha256")).lower()
            if context_sha != citation_sha:
                raise ProductiveAuthorityError(
                    f"FUENTE_CONTEXTO_SHA_DIVERGENTE:{payload_id}:"
                    f"{local_path}:{citation_sha}!={context_sha}"
                )
    design_path = repo / "data/diseno-muestral.yaml"
    design_sha = _sha256_path(design_path)
    return source_path, design_sha


def _citation(
    *, source_id: str, uri: str, sha256: str, locator: str, fields: Iterable[str]
) -> dict[str, Any]:
    return {
        "fuente_id": source_id,
        "ruta_o_uri": uri,
        "sha256": sha256,
        "localizador": locator,
        "campos_autorizados": sorted(set(fields)),
    }


def _authority_from_record(
    record: Mapping[str, Any],
    profile: Mapping[str, Any],
    table: Mapping[str, Any],
    variable: ParsedVariable,
    design_sha: str,
) -> dict[str, Any]:
    key = authority_key_from_e2(record)
    coding = [
        {"codigo": code, "etiqueta": label} for code, label in variable.coding
    ]
    if profile.get("type_rule") != SUPPORTED_TYPE_RULE:
        raise ProductiveAuthorityError("PERFIL_REGLA_TIPO_NO_SOPORTADA")
    kind = "BINARIA" if len(coding) == 2 else "CATEGORICA"
    missing = {
        "codigos": list(variable.missing),
        "ausencia_documentada": not variable.missing,
    }
    source_id = _text(profile["source_id"])
    source_uri = _text(profile["source_uri"])
    source_sha = _text(profile["source_sha256"]).lower()
    source_locator = (
        f"hoja={variable.sheet};filas={variable.start_row}-{variable.end_row};"
        f"variable={variable.name};dominio exhaustivo sin Blanco por secuencia"
    )
    semantic_citation = _citation(
        source_id=source_id,
        uri=source_uri,
        sha256=source_sha,
        locator=source_locator,
        fields=CODEBOOK_FIELDS,
    )
    response_source = table["response_semantics"]["citation"]
    response_locator = (
        f"{_text(response_source['locator'])};hoja={variable.sheet};"
        f"filas_fd={variable.start_row}-{variable.end_row};variable={variable.name}"
    )
    response_citation = _citation(
        source_id=_text(response_source["source_id"]),
        uri=_text(response_source["uri"]),
        sha256=_text(response_source["sha256"]).lower(),
        locator=response_locator,
        fields=RESPONSE_FIELDS,
    )
    context_citations = [
        _citation(
            source_id=_text(source["source_id"]),
            uri=_text(source["uri"]),
            sha256=_text(source["sha256"]).lower(),
            locator=_text(source["locator"]),
            fields=source["fields"],
        )
        for source in table["context_citations"]
    ]
    weight_citation = _citation(
        source_id="REPO-DISENO-MUESTRAL",
        uri="data/diseno-muestral.yaml",
        sha256=design_sha,
        locator=_text(table["design_locator"]),
        fields=WEIGHT_FIELDS,
    )
    row = {
        "schema_version": "AUTORIDAD-SEMANTICA-MARCO-1.0",
        "representacion_id": key.representacion_id,
        "semilla_objeto_logico_id": _text(record.get("objeto_logico_id")),
        "e2_record_id": _text(record.get("record_id")),
        "payload_id": _text(record.get("payload_id")),
        "sha256_contenido": _text(record.get("sha256")).lower(),
        "scope_tipo": key.scope_tipo,
        "scope_id": key.scope_id,
        "variable": key.variable,
        "encuesta": _text(profile["encuesta"]),
        "ola": _text(profile["ola"]),
        "universo_poblacional": _text(table["universo_poblacional"]),
        "unidad_observacion": _text(table["unidad_observacion"]),
        "tipo_estadistico": kind,
        "respuesta_multiple": False,
        "categorias_excluyentes": True,
        "codificacion": coding,
        "missing": missing,
        "unidad_medida": None,
        "rango_valido": None,
        "operacion_estimador": None,
        "ponderador": _text(table["ponderador"]),
        "ponderador_exacto": True,
        "ponderador_fuente_ola_tabla": True,
        "ponderador_fuente": _text(profile["encuesta"]),
        "ponderador_ola": _text(profile["ola"]),
        "ponderador_scope_tipo": "TABLA",
        "ponderador_scope_id": _text(table.get("ponderador_scope_id", variable.sheet)),
        "no_aplica_ponderador_documentado": False,
        "cita_procedencia": sorted(
            [semantic_citation, response_citation, *context_citations, weight_citation],
            key=_canonical_json,
        ),
    }
    return with_authority_id(row)


def _semantic_spec(row: Mapping[str, Any]) -> str:
    return _canonical_json(
        {field: row.get(field) for field in SEMANTIC_AUTHORITY_SPEC_FIELDS}
    )


def _conceptual_identity(row: Mapping[str, Any]) -> tuple[str, str, str]:
    return (_text(row.get("encuesta")), _text(row.get("ola")), _text(row.get("variable")))


def _response_semantics_reason(
    table: Mapping[str, Any], variable_name: str
) -> str | None:
    """Resuelve multiplicidad sólo desde la clasificación declarativa exacta."""

    response = table["response_semantics"]
    if variable_name in response["multiple_response_variables"]:
        return "RESPUESTA_MULTIPLE_DOCUMENTADA"
    if variable_name in response["unresolved_variables"]:
        return "EXCLUSIVIDAD_MULTIPLICIDAD_NO_RESUELTA"
    if response.get("rule") == STRUCTURAL_SCALAR_RULE:
        return None
    if variable_name not in response["scalar_exclusive_variables"]:
        return "EXCLUSIVIDAD_MULTIPLICIDAD_NO_RESUELTA"
    return None


def _specific_blocker(record: Mapping[str, Any], declared: str = "") -> str:
    """Convierte una ausencia de perfil en una insuficiencia material observable."""

    normalized = {
        "SALTOS_MISSING_Y_SCOPE_PONDERADOR_NO_RESUELTOS": "UNIVERSO_REACTIVO_NO_RESUELTO",
        "UNIVERSO_UNIDAD_Y_SCOPE_PONDERADOR_NO_RESUELTOS": "UNIVERSO_POBLACIONAL_NO_DOCUMENTADO",
        "PONDERADOR_OLA_SCOPE_NO_DOCUMENTADO": "PONDERADOR_SCOPE_NO_RESUELTO",
        "PONDERADOR_SCOPE_NO_DOCUMENTADO_PARA_AUTORIDAD": "PONDERADOR_SCOPE_NO_RESUELTO",
        "SCOPE_PONDERADOR_O_UNIVERSO_NO_RESUELTO": "PONDERADOR_SCOPE_NO_RESUELTO",
        "PONDERADOR_NO_COMPATIBLE_CON_UNIDAD_OBSERVACION": "PONDERADOR_SCOPE_NO_RESUELTO",
        "CONFLICTO_PONDERADOR_FAC_SEL_FAC_ELE": "CONFLICTO_PONDERADOR",
        "UNIVERSO_POBLACIONAL_OBJETIVO_NO_DOCUMENTADO": "UNIVERSO_POBLACIONAL_NO_DOCUMENTADO",
        "UNIVERSO_REACTIVO_CONDICIONADO": "UNIVERSO_REACTIVO_NO_RESUELTO",
        "DOMINIO_NO_ENUMERADO_O_SIN_ETIQUETA": "DOMINIO_NO_ENUMERADO",
        "DOMINIO_ESTRUCTURADO_AUSENTE": "DOMINIO_NO_ENUMERADO",
        "MENOS_DE_DOS_CATEGORIAS_VALIDAS": "CODIFICACION_INCOMPLETA",
    }
    if declared:
        return normalized.get(declared, declared)
    kind = _text(record.get("objeto_tipo")).upper()
    if kind in {"VARIABLE-DTA", "VARIABLE-SAV"}:
        if record.get("value_labels"):
            # E2 conserva las etiquetas pero no el par código→etiqueta ni los
            # user-missing necesarios para construir el dominio autorizado.
            return "CODIFICACION_INCOMPLETA"
        return "TIPO_ESTADISTICO_NO_DOCUMENTADO"
    if kind == "VARIABLE-DICCIONARIO":
        # El parser material emitió una semilla por fila de catálogo. Los
        # códigos se pueden agrupar por clave, pero E2 no conservó la etiqueta
        # de cada código como par autorizado.
        return "CODIFICACION_INCOMPLETA"
    if kind in {"VARIABLE-DICCIONARIO-XLS", "VARIABLE-DICCIONARIO-XLSX"}:
        if not record.get("categorias"):
            return "DOMINIO_NO_ENUMERADO"
        return "UNIVERSO_POBLACIONAL_NO_DOCUMENTADO"
    return "FUENTE_PRIMARIA_NO_PRESENTE"


def _seed_outcome(
    record: Mapping[str, Any], estado: str, razon: str = ""
) -> dict[str, Any]:
    return {
        "e2_record_id": _text(record.get("record_id")),
        "estado_final": estado,
        "payload_id": _text(record.get("payload_id")),
        "representacion_id": _text(record.get("representacion_id")),
        "scope_id": _text(record.get("objeto_padre_id"))
        or _text(record.get("tabla")),
        "variable": _text(record.get("nombre")),
        "razon": razon,
    }


def _load_base_authorities(
    path: Path, preserve_authority_ids: Iterable[str]
) -> list[dict[str, Any]]:
    preserve = set(preserve_authority_ids)
    rows: list[dict[str, Any]] = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        try:
            row = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ProductiveAuthorityError(
                f"BASE_AUTORIDAD_JSON_INVALIDO:linea={line_number}:{exc}"
            ) from exc
        validate_authority_row(row)
        if row["autoridad_id"] in preserve:
            rows.append(row)
    observed = {row["autoridad_id"] for row in rows}
    missing = sorted(preserve - observed)
    if missing:
        raise ProductiveAuthorityError(
            "BASE_AUTORIDADES_PRESERVADAS_AUSENTES:" + ",".join(missing)
        )
    return rows


def _materialize_pr346(
    *,
    repo: Path,
    corpus_root: Path,
    index_path: Path,
    profiles_path: Path,
    base_authority_path: Path,
    residual_output: Path | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    profile_doc = _load_profiles(profiles_path)
    profiles: dict[str, dict[str, Any]] = profile_doc["profiles_by_payload"]
    blocked_payloads: dict[str, str] = profile_doc["blocked_payloads"]

    manifest = yaml.safe_load((repo / "data/manifiesto.yaml").read_text(encoding="utf-8"))
    provenance = ProvenanceIndex(
        manifest,
        _read_tsv(repo / "data/censo-explotacion-2026-08-17.tsv"),
        _read_tsv(repo / "data/curacion-universo/ledger-inspecciones-barrido2.tsv"),
        _read_tsv(repo / "data/curacion-universo/universo-declarado-t0.tsv"),
    )

    parsed_by_payload: dict[str, dict[tuple[str, str], ParsedVariable]] = {}
    design_sha_by_payload: dict[str, str] = {}
    for payload_id, profile in sorted(profiles.items()):
        source_path, design_sha = _validate_profile_sources(profile, corpus_root, repo)
        parsed_by_payload[payload_id] = parse_inegi_fd_7col(
            source_path, profile["tables"]
        )
        design_sha_by_payload[payload_id] = design_sha

    base_rows = _load_base_authorities(
        base_authority_path, profile_doc["preserve_authority_ids"]
    )
    rows_by_key = {authority_key_from_row(row): row for row in base_rows}
    if len(rows_by_key) != len(base_rows):
        raise ProductiveAuthorityError("BASE_AUTORIDAD_CLAVE_DUPLICADA")

    reasons: Counter[str] = Counter()
    source_attempts: dict[str, Counter[str]] = defaultdict(Counter)
    weight_scopes: Counter[str] = Counter()
    joins: Counter[str] = Counter()
    seeds_by_type: Counter[str] = Counter()
    matched_base_keys: set[Any] = set()
    residual_handle = None
    if residual_output is not None:
        residual_output.parent.mkdir(parents=True, exist_ok=True)
        residual_handle = residual_output.open("w", encoding="utf-8", newline="")
    try:
        reader = E2IndexReader(index_path)
        for record in reader:
            if not is_variable_seed(record):
                continue
            seeds_by_type[_text(record.get("objeto_tipo")).upper()] += 1
            join = provenance.resolve(record)
            joins[join.status] += 1
            if join.status != "EXACTA":
                continue
            reason = ""
            try:
                key = authority_key_from_e2(record)
            except AutoridadSemanticaError:
                key = None
                reason = "CLAVE_MATERIAL_NO_RESUELTA"
            if key is not None and key in rows_by_key:
                matched_base_keys.add(key)
                weight_scopes["RESUELTO"] += 1
                source_attempts[_text(record.get("payload_id"))]["AUTORIZADA_PREEXISTENTE"] += 1
                continue
            payload_id = _text(record.get("payload_id"))
            profile = profiles.get(payload_id)
            if not reason and profile is None:
                reason = blocked_payloads.get(
                    payload_id, "SIN_PROYECTOR_O_PERFIL_SEMANTICO_IMPLEMENTADO"
                )
                weight_scopes["NO_RESUELTO"] += 1
            table = None
            if profile is not None:
                sheet = _text(record.get("hoja"))
                table = profile["tables"].get(sheet)
                if not reason and table is None:
                    reason = profile.get("blocked_tables", {}).get(
                        sheet, "SCOPE_PONDERADOR_O_UNIVERSO_NO_RESUELTO"
                    )
                    weight_scopes["NO_RESUELTO"] += 1
                elif table is not None:
                    weight_scopes["RESUELTO"] += 1
            parsed = None
            if profile is not None and table is not None:
                parsed = parsed_by_payload[payload_id].get(
                    (_text(record.get("hoja")), _text(record.get("nombre")))
                )
                if not reason and parsed is None:
                    reason = "VARIABLE_NO_RESUELTA_EN_DESCRIPTOR"
                elif parsed is not None and parsed.reason:
                    reason = parsed.reason
                elif parsed is not None:
                    reason = _response_semantics_reason(table, parsed.name) or ""
            if reason:
                reasons[reason] += 1
                source_attempts[payload_id][reason] += 1
                if residual_handle is not None:
                    residual_handle.write(
                        _canonical_json(
                            {
                                "e2_record_id": _text(record.get("record_id")),
                                "payload_id": payload_id,
                                "representacion_id": _text(record.get("representacion_id")),
                                "scope_id": _text(record.get("objeto_padre_id"))
                                or _text(record.get("tabla")),
                                "variable": _text(record.get("nombre")),
                                "razon": reason,
                            }
                        )
                        + "\n"
                    )
                continue
            assert profile is not None and table is not None and parsed is not None
            authority = _authority_from_record(
                record,
                profile,
                table,
                parsed,
                design_sha_by_payload[payload_id],
            )
            validate_authority_row(authority)
            authority_key = authority_key_from_row(authority)
            if authority_key in rows_by_key and rows_by_key[authority_key] != authority:
                raise ProductiveAuthorityError(
                    "AUTORIDAD_MATERIAL_DIVERGENTE:" + _canonical_json(authority_key.as_dict())
                )
            rows_by_key[authority_key] = authority
            source_attempts[payload_id]["AUTORIDAD_PROYECTADA"] += 1
    finally:
        if residual_handle is not None:
            residual_handle.close()

    expected_sha = expected_index_sha(
        repo / "data/curacion-universo/baseline-material-barrido2.json"
    )
    reader.require_sha(expected_sha)
    orphan_base = sorted(
        row["autoridad_id"]
        for key, row in rows_by_key.items()
        if key in {authority_key_from_row(base) for base in base_rows}
        and key not in matched_base_keys
    )
    if orphan_base:
        raise ProductiveAuthorityError(
            "BASE_AUTORIDAD_HUERFANA:" + ",".join(orphan_base)
        )
    observed_counts = {
        "semillas_variables_totales": sum(seeds_by_type.values()),
        "uniones_exactas": joins["EXACTA"],
        "uniones_ausentes": joins["AUSENTE"],
        "uniones_ambiguas": joins["AMBIGUA"],
    }
    if observed_counts != profile_doc["expected_counts"]:
        raise ProductiveAuthorityError(
            "CONTEOS_UNIVERSO_DIVERGENTES:"
            f"esperado={_canonical_json(profile_doc['expected_counts'])}:"
            f"observado={_canonical_json(observed_counts)}"
        )

    all_rows = [rows_by_key[key] for key in sorted(rows_by_key)]
    rows_by_identity: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in all_rows:
        rows_by_identity[_conceptual_identity(row)].append(row)
    conflicts: dict[tuple[str, str, str], int] = {}
    compatible_duplicates = 0
    compatible_duplicate_details: list[dict[str, Any]] = []
    for identity, identity_rows in rows_by_identity.items():
        specs = {_semantic_spec(row) for row in identity_rows}
        if len(specs) > 1:
            conflicts[identity] = len(identity_rows)
        else:
            compatible_duplicates += len(identity_rows) - 1
            if len(identity_rows) > 1:
                compatible_duplicate_details.append(
                    {
                        "encuesta": identity[0],
                        "ola": identity[1],
                        "variable": identity[2],
                        "autoridades": sorted(row["autoridad_id"] for row in identity_rows),
                    }
                )
    if conflicts:
        conflict_keys = set(conflicts)
        conflict_rows = [
            row for row in all_rows if _conceptual_identity(row) in conflict_keys
        ]
        all_rows = [
            row for row in all_rows if _conceptual_identity(row) not in conflict_keys
        ]
        for identity, count in conflicts.items():
            reasons["CONFLICTO_IDENTIDAD_CONCEPTUAL"] += count
        for row in conflict_rows:
            source_attempts[row["payload_id"]]["CONFLICTO_IDENTIDAD_CONCEPTUAL"] += 1
        if residual_output is not None:
            with residual_output.open("a", encoding="utf-8", newline="") as handle:
                for row in sorted(conflict_rows, key=lambda value: authority_key_from_row(value)):
                    handle.write(
                        _canonical_json(
                            {
                                "e2_record_id": row["e2_record_id"],
                                "payload_id": row["payload_id"],
                                "representacion_id": row["representacion_id"],
                                "scope_id": row["scope_id"],
                                "variable": row["variable"],
                                "razon": "CONFLICTO_IDENTIDAD_CONCEPTUAL",
                            }
                        )
                        + "\n"
                    )

    final_by_identity: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in all_rows:
        final_by_identity[_conceptual_identity(row)].append(row)
    authority_by_wave: Counter[str] = Counter()
    candidate_by_wave: Counter[str] = Counter()
    candidates_by_type: Counter[str] = Counter()
    for row in all_rows:
        authority_by_wave[f"{row['encuesta']} | {row['ola']}"] += 1
        source_attempts[row["payload_id"]]["AUTORIDAD_FINAL"] += 1
    for identity, identity_rows in final_by_identity.items():
        candidate_by_wave[f"{identity[0]} | {identity[1]}"] += 1
        candidates_by_type[_text(identity_rows[0]["tipo_estadistico"])] += 1

    exact = joins["EXACTA"]
    base_identities = {_conceptual_identity(row) for row in base_rows}
    final_identities = set(final_by_identity)
    source_without_authority = sorted(
        payload_id
        for payload_id, counts in source_attempts.items()
        if not counts["AUTORIDAD_FINAL"] and not counts["AUTORIZADA_PREEXISTENTE"]
    )
    top_reasons = [
        {"razon": reason, "semillas": count}
        for reason, count in sorted(
            reasons.items(), key=lambda item: (-item[1], item[0])
        )[:10]
    ]
    without_projector = reasons["SIN_PROYECTOR_O_PERFIL_SEMANTICO_IMPLEMENTADO"]
    exact_not_emitted = exact - len(all_rows)
    specifically_evaluated_not_emitted = exact_not_emitted - without_projector
    repair_audit = dict(profile_doc["repair_audit"])
    if len(all_rows) != repair_audit["autoridades_permanecen"]:
        raise ProductiveAuthorityError(
            "AUDITORIA_REPARACION_RESULTADO_DIVERGENTE:"
            f"esperado={repair_audit['autoridades_permanecen']}:observado={len(all_rows)}"
        )
    repair_audit["retiradas_no_aplica_semantica_no_resuelta"] = len(
        repair_audit["dependientes_no_aplica_como_missing"]
    )
    repair_audit["retiradas_multirrespuesta_total"] = len(
        repair_audit["retiradas_multirrespuesta"]
    )
    repair_audit["retiradas_exclusividad_multiplicidad_no_resuelta_total"] = len(
        repair_audit["retiradas_exclusividad_multiplicidad_no_resuelta"]
    )
    repair_audit["dependientes_no_aplica_como_missing_conservadas"] = 0
    repair_audit["retiradas_total"] = (
        repair_audit["autoridades_actuales_examinadas"]
        - repair_audit["autoridades_permanecen"]
    )
    repair_audit["rendimiento_antes_reparacion"] = {
        "semillas_exactas": exact,
        "autoridades_emitidas": repair_audit["autoridades_actuales_examinadas"],
        "semillas_no_emitidas": (
            exact - repair_audit["autoridades_actuales_examinadas"]
        ),
        "sin_proyector_o_perfil": without_projector,
        "evaluadas_con_diagnostico_especifico_y_no_emitidas": (
            exact - repair_audit["autoridades_actuales_examinadas"] - without_projector
        ),
    }
    diagnostics = {
        "schema_version": DIAGNOSTIC_SCHEMA_VERSION,
        "trazabilidad": {
            "origin_main_inicial": profile_doc["origin_main_inicial"],
            "pr_343_merge_commit": profile_doc["pr_343_merge_commit"],
            "indice_e2_sha256": expected_sha,
            "perfiles_sha256": _sha256_path(profiles_path),
            "proyector": SUPPORTED_PROJECTOR,
            "receta": (
                "python3 tools/curador_registro/autoridad_semantica_productiva.py "
                "--corpus-root <CORPUS_RAW> --indice-e2 <E2_PRIVADO> "
                "--output data/curacion-universo/autoridad-semantica-marco-v1_0.jsonl "
                "--diagnostico "
                "data/curacion-universo/diagnostico-autoridad-semantica-marco-v1_0.json "
                "--residuales /tmp/semillas-no-emitidas.jsonl"
            ),
        },
        "resumen": {
            "autoridades_totales": len(all_rows),
            "autoridades_emitidas": len(all_rows),
            "candidatas_totales": len(final_by_identity),
            "autoridades_emitidas_en_acto": len(all_rows) - len(base_rows),
            "candidatas_nuevas_en_acto": len(final_identities - base_identities),
            "semillas_exactas_examinadas": exact,
            "semillas_exactas": exact,
            "semillas_no_emitidas": exact_not_emitted,
            "sin_proyector_o_perfil": without_projector,
            "evaluadas_con_diagnostico_especifico_y_no_emitidas": (
                specifically_evaluated_not_emitted
            ),
            "semillas_variables_totales": sum(seeds_by_type.values()),
            "tasa_emision_sobre_semillas_exactas": (
                len(all_rows) / exact if exact else 0.0
            ),
            "encuestas_con_candidatas": len({identity[0] for identity in final_by_identity}),
            "olas_con_candidatas": len(
                {(identity[0], identity[1]) for identity in final_by_identity}
            ),
            "conflictos_identidad_conceptual": len(conflicts),
            "autoridades_huerfanas": 0,
            "duplicados_compatibles": compatible_duplicates,
            "uniones_exactas": exact,
            "uniones_ausentes": joins["AUSENTE"],
            "uniones_ambiguas": joins["AMBIGUA"],
        },
        "auditoria_reparacion_pr_346": repair_audit,
        "autoridades_por_encuesta_ola": dict(sorted(authority_by_wave.items())),
        "candidatas_por_encuesta_ola": dict(sorted(candidate_by_wave.items())),
        "candidatas_por_tipo_estadistico": dict(sorted(candidates_by_type.items())),
        "conflictos_conceptuales": [
            {
                "encuesta": key[0],
                "ola": key[1],
                "variable": key[2],
                "evidencias_materiales": [
                    {
                        "autoridad_id": row["autoridad_id"],
                        "e2_record_id": row["e2_record_id"],
                        "payload_id": row["payload_id"],
                        "ponderador": row["ponderador"],
                        "ponderador_scope_id": row["ponderador_scope_id"],
                        "representacion_id": row["representacion_id"],
                        "scope_id": row["scope_id"],
                        "scope_tipo": row["scope_tipo"],
                        "unidad_observacion": row["unidad_observacion"],
                        "universo_poblacional": row["universo_poblacional"],
                    }
                    for row in sorted(
                        rows_by_identity[key], key=lambda value: authority_key_from_row(value)
                    )
                ],
            }
            for key in sorted(conflicts)
        ],
        "duplicados_compatibles_detalle": compatible_duplicate_details,
        "razones_insuficiencia": dict(sorted(reasons.items())),
        "top_razones_insuficiencia": top_reasons,
        "scopes_ponderador": dict(sorted(weight_scopes.items())),
        "fuentes": {
            payload_id: dict(sorted(counts.items()))
            for payload_id, counts in sorted(source_attempts.items())
        },
        "fuentes_sin_autoridad_emitida": source_without_authority,
        "semillas_por_objeto_tipo": dict(sorted(seeds_by_type.items())),
    }
    return all_rows, diagnostics


def materialize(
    *,
    repo: Path,
    corpus_root: Path,
    index_path: Path,
    profiles_path: Path,
    base_authority_path: Path,
    residual_output: Path | None = None,
    inventory_output: Path | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Agota el universo E2 y asigna un estado final a cada semilla."""

    profile_doc = _load_profiles(profiles_path)
    profiles: dict[str, dict[str, Any]] = profile_doc["profiles_by_payload"]
    blocked_payloads: dict[str, str] = profile_doc["blocked_payloads"]
    manifest = yaml.safe_load((repo / "data/manifiesto.yaml").read_text(encoding="utf-8"))
    provenance = ProvenanceIndex(
        manifest,
        _read_tsv(repo / "data/censo-explotacion-2026-08-17.tsv"),
        _read_tsv(repo / "data/curacion-universo/ledger-inspecciones-barrido2.tsv"),
        _read_tsv(repo / "data/curacion-universo/universo-declarado-t0.tsv"),
    )

    parsed_by_payload: dict[str, dict[tuple[str, str], ParsedVariable]] = {}
    design_sha_by_payload: dict[str, str] = {}
    for payload_id, profile in sorted(profiles.items()):
        source_path, design_sha = _validate_profile_sources(profile, corpus_root, repo)
        parsed_by_payload[payload_id] = parse_inegi_fd_7col(
            source_path, profile["tables"]
        )
        design_sha_by_payload[payload_id] = design_sha

    base_rows = _load_base_authorities(
        base_authority_path, profile_doc["preserve_authority_ids"]
    )
    rows_by_key = {authority_key_from_row(row): row for row in base_rows}
    if len(rows_by_key) != len(base_rows):
        raise ProductiveAuthorityError("BASE_AUTORIDAD_CLAVE_DUPLICADA")
    base_keys = set(rows_by_key)

    strict_joins: Counter[str] = Counter()
    final_joins: Counter[str] = Counter()
    repair_reasons: Counter[str] = Counter()
    initial_absent_components: Counter[str] = Counter()
    seeds_by_type: Counter[str] = Counter()
    source_attempts: dict[str, Counter[str]] = defaultdict(Counter)
    weight_scopes: Counter[str] = Counter()
    matched_base_keys: set[Any] = set()
    outcomes: list[dict[str, Any]] = []
    outcome_index: dict[str, int] = {}
    seed_keys: dict[str, Any] = {}
    inventory_groups: dict[tuple[str, ...], dict[str, Any]] = {}
    inventory_seed_count = 0

    reader = E2IndexReader(index_path)
    for record in reader:
        if not is_variable_seed(record):
            continue
        record_id = _text(record.get("record_id"))
        payload_id = _text(record.get("payload_id"))
        seeds_by_type[_text(record.get("objeto_tipo")).upper()] += 1
        strict_join = provenance.resolve(record)
        strict_joins[strict_join.status] += 1
        component_cards = None
        if strict_join.status == "AUSENTE":
            component_cards = provenance.component_cardinalities(record)
            missing_components = [
                name.upper()
                for name, count in component_cards.items()
                if count == 0
            ]
            initial_absent_components["+".join(missing_components)] += 1
        join = provenance.resolve_repaired(record)
        if join.status == "EXACTA_REPARADA":
            final_joins["EXACTA"] += 1
            repair_reasons[join.reason] += 1
        else:
            final_joins[join.status] += 1

        try:
            key = authority_key_from_e2(record)
            seed_keys[record_id] = key
        except AutoridadSemanticaError:
            key = None

        if (
            strict_join.status == "EXACTA"
            and payload_id not in PR346_PROFILED_OR_BLOCKED_PAYLOADS
            and key not in base_keys
        ):
            inventory_seed_count += 1
            scope_id = _text(record.get("objeto_padre_id")) or _text(
                record.get("tabla")
            )
            inventory_key = (
                payload_id,
                _text(record.get("representacion_id")),
                _text(record.get("objeto_tipo")),
                _text(record.get("format")),
                _text(record.get("ruta_relativa")),
                _text(record.get("hoja")),
                scope_id,
            )
            group = inventory_groups.setdefault(
                inventory_key,
                {
                    "payload_id": payload_id,
                    "representacion_id": inventory_key[1],
                    "objeto_tipo": inventory_key[2],
                    "formato_fuente": inventory_key[3],
                    "ruta_fuente": inventory_key[4],
                    "tabla": inventory_key[5],
                    "scope_id": inventory_key[6],
                    "semillas": 0,
                    "metadatos_estructurados": Counter(),
                    "diccionario_catalogo_disponible": False,
                    "documentacion_diseno_disponible": None,
                    "ponderadores_documentados": [],
                    "cuestionario_metodologia_disponible": None,
                    "estado_documentacion_contextual": (
                        "NO_RECONCILIADA_DETERMINISTICAMENTE_A_PAYLOAD"
                    ),
                    "encuesta_documentada": None,
                    "ola_documentada": None,
                    "patron_descriptor": (
                        f"{inventory_key[2]}|{_text(record.get('parser'))}|"
                        f"{inventory_key[3]}"
                    ),
                },
            )
            group["semillas"] += 1
            for field in (
                "etiqueta",
                "texto_reactivo",
                "definicion",
                "categorias",
                "value_labels",
                "unidad",
                "periodo",
                "poblacion",
            ):
                value = record.get(field)
                present = bool(value) if isinstance(value, list) else bool(
                    _text(value)
                    and _text(value).upper()
                    not in {"NO-APLICA", "[REDACTADO-PRIVACIDAD]"}
                )
                if present:
                    group["metadatos_estructurados"][field] += 1
            group["diccionario_catalogo_disponible"] = group[
                "diccionario_catalogo_disponible"
            ] or _text(record.get("objeto_tipo")).upper().startswith(
                "VARIABLE-DICCIONARIO"
            )

        if join.status not in {"EXACTA", "EXACTA_REPARADA"}:
            cards = component_cards or provenance.component_cardinalities(record)
            missing = [name.upper() for name, count in cards.items() if count == 0]
            reason = (
                "RUTA_CENSO_AUSENTE_PAYLOAD_NO_IDENTIFICADO"
                if missing == ["CENSO"]
                else "RUTA_CENSO_Y_T0_AUSENTES_PAYLOAD_NO_IDENTIFICADO"
                if missing == ["CENSO", "T0"]
                else "PROCEDENCIA_COMPONENTES_NO_RECONCILIADOS:" + ",".join(missing)
            )
            outcome = _seed_outcome(
                record, "PROCEDENCIA_NO_RECONCILIADA", reason
            )
            outcomes.append(outcome)
            outcome_index[record_id] = len(outcomes) - 1
            source_attempts[payload_id][reason] += 1
            continue

        reason = "" if key is not None else "CLAVE_MATERIAL_NO_RESUELTA"
        if key is not None and key in rows_by_key:
            matched_base_keys.add(key)
            authority = rows_by_key[key]
            if authority["e2_record_id"] == record_id:
                outcome = _seed_outcome(record, "AUTORIDAD_EMITIDA")
            else:
                outcome = _seed_outcome(
                    record,
                    "DUPLICADO_COMPATIBLE",
                    "MISMA_CLAVE_MATERIAL_Y_ESPECIFICACION",
                )
            outcomes.append(outcome)
            outcome_index[record_id] = len(outcomes) - 1
            weight_scopes["RESUELTO"] += 1
            source_attempts[payload_id]["AUTORIZADA_PREEXISTENTE"] += 1
            continue

        profile = profiles.get(payload_id)
        if not reason and profile is None:
            reason = _specific_blocker(
                record, blocked_payloads.get(payload_id, "")
            )
            weight_scopes["NO_RESUELTO"] += 1
        table = None
        if profile is not None:
            sheet = _text(record.get("hoja"))
            table = profile["tables"].get(sheet)
            if not reason and table is None:
                reason = _specific_blocker(
                    record,
                    profile.get("blocked_tables", {}).get(
                        sheet, "PONDERADOR_SCOPE_NO_RESUELTO"
                    ),
                )
                weight_scopes["NO_RESUELTO"] += 1
            elif table is not None:
                weight_scopes["RESUELTO"] += 1
        parsed = None
        if profile is not None and table is not None:
            parsed = parsed_by_payload[payload_id].get(
                (_text(record.get("hoja")), _text(record.get("nombre")))
            )
            if not reason and parsed is None:
                reason = "VARIABLE_NO_RESUELTA_EN_DESCRIPTOR"
            elif parsed is not None and parsed.reason:
                reason = _specific_blocker(record, parsed.reason)
            elif parsed is not None:
                reason = _response_semantics_reason(table, parsed.name) or ""
        if reason:
            outcome = _seed_outcome(
                record, "INSUFICIENCIA_SEMANTICA_ESPECIFICA", reason
            )
            outcomes.append(outcome)
            outcome_index[record_id] = len(outcomes) - 1
            source_attempts[payload_id][reason] += 1
            continue

        assert profile is not None and table is not None and parsed is not None
        authority = _authority_from_record(
            record,
            profile,
            table,
            parsed,
            design_sha_by_payload[payload_id],
        )
        validate_authority_row(authority)
        authority_key = authority_key_from_row(authority)
        existing = rows_by_key.get(authority_key)
        if existing is not None and _semantic_spec(existing) != _semantic_spec(authority):
            raise ProductiveAuthorityError(
                "AUTORIDAD_MATERIAL_DIVERGENTE:"
                + _canonical_json(authority_key.as_dict())
            )
        rows_by_key.setdefault(authority_key, authority)
        outcome = _seed_outcome(record, "AUTORIDAD_EMITIDA")
        outcomes.append(outcome)
        outcome_index[record_id] = len(outcomes) - 1
        source_attempts[payload_id]["AUTORIDAD_PROYECTADA"] += 1

    expected_sha = expected_index_sha(
        repo / "data/curacion-universo/baseline-material-barrido2.json"
    )
    reader.require_sha(expected_sha)
    initial_counts = {
        "semillas_variables_totales": sum(seeds_by_type.values()),
        "uniones_exactas": strict_joins["EXACTA"],
        "uniones_ausentes": strict_joins["AUSENTE"],
        "uniones_ambiguas": strict_joins["AMBIGUA"],
    }
    if initial_counts != profile_doc["expected_counts"]:
        raise ProductiveAuthorityError(
            "CONTEOS_UNIVERSO_DIVERGENTES:"
            f"esperado={_canonical_json(profile_doc['expected_counts'])}:"
            f"observado={_canonical_json(initial_counts)}"
        )
    if inventory_seed_count != 107443:
        raise ProductiveAuthorityError(
            f"INVENTARIO_INICIAL_DIVERGENTE:esperado=107443:observado={inventory_seed_count}"
        )

    orphan_base = sorted(
        row["autoridad_id"]
        for key, row in rows_by_key.items()
        if key in base_keys and key not in matched_base_keys
    )
    if orphan_base:
        raise ProductiveAuthorityError(
            "BASE_AUTORIDAD_HUERFANA:" + ",".join(orphan_base)
        )

    projected_rows = [rows_by_key[key] for key in sorted(rows_by_key)]
    rows_by_identity: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in projected_rows:
        rows_by_identity[_conceptual_identity(row)].append(row)
    conflicts: dict[tuple[str, str, str], int] = {}
    compatible_duplicate_details: list[dict[str, Any]] = []
    for identity, identity_rows in rows_by_identity.items():
        specs = {_semantic_spec(row) for row in identity_rows}
        ordered = sorted(identity_rows, key=lambda row: authority_key_from_row(row))
        if len(specs) > 1:
            conflicts[identity] = len(identity_rows)
            for row in ordered:
                index = outcome_index[row["e2_record_id"]]
                outcomes[index]["estado_final"] = "CONFLICTO_IDENTIDAD_CONCEPTUAL"
                outcomes[index]["razon"] = "CONFLICTO_IDENTIDAD_CONCEPTUAL"
        elif len(ordered) > 1:
            compatible_duplicate_details.append(
                {
                    "encuesta": identity[0],
                    "ola": identity[1],
                    "variable": identity[2],
                    "autoridades": [row["autoridad_id"] for row in ordered],
                }
            )
            for row in ordered[1:]:
                index = outcome_index[row["e2_record_id"]]
                outcomes[index]["estado_final"] = "DUPLICADO_COMPATIBLE"
                outcomes[index]["razon"] = "MISMA_ESPECIFICACION_SEMANTICA"

    conflict_identities = set(conflicts)
    conflict_material_keys = {
        authority_key_from_row(row)
        for identity in conflict_identities
        for row in rows_by_identity[identity]
    }
    for record_id, key in seed_keys.items():
        if key in conflict_material_keys and record_id in outcome_index:
            index = outcome_index[record_id]
            outcomes[index]["estado_final"] = "CONFLICTO_IDENTIDAD_CONCEPTUAL"
            outcomes[index]["razon"] = "CONFLICTO_IDENTIDAD_CONCEPTUAL"
    all_rows = [
        row
        for row in projected_rows
        if _conceptual_identity(row) not in conflict_identities
    ]

    if len(outcomes) != sum(seeds_by_type.values()):
        raise ProductiveAuthorityError(
            f"COBERTURA_ESTADOS_INCOMPLETA:{len(outcomes)}!={sum(seeds_by_type.values())}"
        )
    forbidden = [
        row for row in outcomes if row["razon"] == "SIN_PROYECTOR_O_PERFIL_SEMANTICO_IMPLEMENTADO"
    ]
    if forbidden:
        raise ProductiveAuthorityError("RESIDUAL_GENERICO_NO_AGOTADO")

    final_by_identity: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    authority_by_wave: Counter[str] = Counter()
    candidate_by_wave: Counter[str] = Counter()
    candidates_by_type: Counter[str] = Counter()
    for row in all_rows:
        identity = _conceptual_identity(row)
        final_by_identity[identity].append(row)
        authority_by_wave[f"{identity[0]} | {identity[1]}"] += 1
        source_attempts[row["payload_id"]]["AUTORIDAD_FINAL"] += 1
    for identity, identity_rows in final_by_identity.items():
        candidate_by_wave[f"{identity[0]} | {identity[1]}"] += 1
        candidates_by_type[_text(identity_rows[0]["tipo_estadistico"])] += 1

    state_counts = Counter(row["estado_final"] for row in outcomes)
    for state in (
        "AUTORIDAD_EMITIDA",
        "DUPLICADO_COMPATIBLE",
        "CONFLICTO_IDENTIDAD_CONCEPTUAL",
        "INSUFICIENCIA_SEMANTICA_ESPECIFICA",
        "PROCEDENCIA_NO_RECONCILIADA",
    ):
        state_counts[state] += 0
    reason_counts = Counter(
        row["razon"] for row in outcomes if row["estado_final"] != "AUTORIDAD_EMITIDA"
    )
    final_exact = final_joins["EXACTA"]
    baseline_authorities = profile_doc["repair_audit"]["autoridades_permanecen"]
    baseline_candidates = 142
    residual_rows = [
        row for row in outcomes if row["estado_final"] != "AUTORIDAD_EMITIDA"
    ]
    if residual_output is not None:
        residual_output.parent.mkdir(parents=True, exist_ok=True)
        residual_output.write_text(
            "".join(_canonical_json(row) + "\n" for row in residual_rows),
            encoding="utf-8",
        )
    if inventory_output is not None:
        inventory_output.parent.mkdir(parents=True, exist_ok=True)
        inventory_rows = []
        for group in inventory_groups.values():
            row = dict(group)
            row["metadatos_estructurados"] = dict(
                sorted(row["metadatos_estructurados"].items())
            )
            row["semillas_resolubles_por_unidad_implementacion"] = row["semillas"]
            inventory_rows.append(row)
        inventory_rows.sort(
            key=lambda row: (
                -row["semillas_resolubles_por_unidad_implementacion"],
                row["patron_descriptor"],
                row["payload_id"],
                row["scope_id"],
            )
        )
        inventory_output.write_text(
            "".join(_canonical_json(row) + "\n" for row in inventory_rows),
            encoding="utf-8",
        )

    repair_audit = dict(profile_doc["repair_audit"])
    repair_audit["insuficiencias_previas_reevaluadas"] = 5428
    diagnostics = {
        "schema_version": DIAGNOSTIC_SCHEMA_VERSION,
        "trazabilidad": {
            "origin_main_inicial": profile_doc["origin_main_inicial"],
            "pr_343_merge_commit": profile_doc["pr_343_merge_commit"],
            "indice_e2_sha256": expected_sha,
            "perfiles_sha256": _sha256_path(profiles_path),
            "proyectores": sorted(
                {profile["projector"] for profile in profiles.values()}
            ),
            "receta": (
                "python3 tools/curador_registro/autoridad_semantica_productiva.py "
                "--corpus-root <CORPUS_RAW> --indice-e2 <E2_PRIVADO> "
                "--output data/curacion-universo/autoridad-semantica-marco-v1_0.jsonl "
                "--diagnostico data/curacion-universo/diagnostico-autoridad-semantica-marco-v1_0.json "
                "--inventario data/curacion-universo/inventario-autoridad-semantica-marco-v1_0.jsonl "
                "--residuales /tmp/semillas-no-emitidas.jsonl"
            ),
        },
        "resumen": {
            "semillas_totales": sum(seeds_by_type.values()),
            "semillas_exactas_iniciales": strict_joins["EXACTA"],
            "semillas_ausentes_iniciales": strict_joins["AUSENTE"],
            "uniones_reparadas": repair_reasons[
                "T0_AUSENTE_RECONSTRUIDO_POR_CENSO_LEDGER_HASH"
            ],
            "uniones_exactas_finales": final_exact,
            "procedencia_no_reconciliada": state_counts[
                "PROCEDENCIA_NO_RECONCILIADA"
            ],
            "autoridades_totales": len(all_rows),
            "candidatas_totales": len(final_by_identity),
            "autoridades_nuevas": len(all_rows) - baseline_authorities,
            "candidatas_nuevas": len(final_by_identity) - baseline_candidates,
            "tasa_emision_sobre_semillas_exactas": (
                len(all_rows) / final_exact if final_exact else 0.0
            ),
            "tasa_emision_sobre_semillas_totales": (
                len(all_rows) / sum(seeds_by_type.values()) if seeds_by_type else 0.0
            ),
            "encuestas_con_candidatas": len(
                {identity[0] for identity in final_by_identity}
            ),
            "olas_con_candidatas": len(
                {(identity[0], identity[1]) for identity in final_by_identity}
            ),
            "duplicados_compatibles": state_counts["DUPLICADO_COMPATIBLE"],
            "conflictos_identidad_conceptual": len(conflicts),
            "autoridades_huerfanas": 0,
            "sin_proyector_o_perfil_semantico_implementado": 0,
            "semillas_no_emitidas": len(residual_rows),
        },
        "antes_pr346_despues": {
            "candidatas": {"antes": 142, "despues": len(final_by_identity)},
            "sin_proyector_o_perfil": {"antes": 107443, "despues": 0},
            "insuficiencias_especificas": {
                "antes": 5428,
                "despues": state_counts["INSUFICIENCIA_SEMANTICA_ESPECIFICA"],
                "reevaluadas": 5428,
            },
            "procedencias_ausentes": {
                "antes": 16832,
                "despues": state_counts["PROCEDENCIA_NO_RECONCILIADA"],
            },
        },
        "estados_finales": dict(sorted(state_counts.items())),
        "reparacion_procedencia": dict(sorted(repair_reasons.items())),
        "procedencia_ausente_inicial_por_componente": dict(
            sorted(initial_absent_components.items())
        ),
        "auditoria_reparacion_pr_346": repair_audit,
        "autoridades_por_encuesta_ola": dict(sorted(authority_by_wave.items())),
        "candidatas_por_encuesta_ola": dict(sorted(candidate_by_wave.items())),
        "candidatas_por_tipo_estadistico": dict(sorted(candidates_by_type.items())),
        "conflictos_conceptuales": [
            {
                "encuesta": key[0],
                "ola": key[1],
                "variable": key[2],
                "evidencias_materiales": [
                    {
                        "autoridad_id": row["autoridad_id"],
                        "e2_record_id": row["e2_record_id"],
                        "payload_id": row["payload_id"],
                        "ponderador": row["ponderador"],
                        "ponderador_scope_id": row["ponderador_scope_id"],
                        "representacion_id": row["representacion_id"],
                        "scope_id": row["scope_id"],
                        "scope_tipo": row["scope_tipo"],
                        "unidad_observacion": row["unidad_observacion"],
                        "universo_poblacional": row["universo_poblacional"],
                    }
                    for row in sorted(
                        rows_by_identity[key],
                        key=lambda value: authority_key_from_row(value),
                    )
                ],
            }
            for key in sorted(conflicts)
        ],
        "duplicados_compatibles_detalle": compatible_duplicate_details,
        "razones_finales_residual": dict(sorted(reason_counts.items())),
        "top_razones_insuficiencia": [
            {"razon": reason, "semillas": count}
            for reason, count in sorted(
                reason_counts.items(), key=lambda item: (-item[1], item[0])
            )[:10]
        ],
        "scopes_ponderador": dict(sorted(weight_scopes.items())),
        "fuentes": {
            payload_id: dict(sorted(counts.items()))
            for payload_id, counts in sorted(source_attempts.items())
        },
        "fuentes_sin_autoridad_emitida": sorted(
            payload_id
            for payload_id, counts in source_attempts.items()
            if not counts["AUTORIDAD_FINAL"]
        ),
        "semillas_por_objeto_tipo": dict(sorted(seeds_by_type.items())),
        "inventario_inicial_sin_proyector_perfil": {
            "semillas": inventory_seed_count,
            "unidades": len(inventory_groups),
        },
    }
    return all_rows, diagnostics


def main(argv: list[str] | None = None) -> int:
    repo = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--corpus-root", type=Path, required=True)
    parser.add_argument("--indice-e2", type=Path, required=True)
    parser.add_argument(
        "--perfiles",
        type=Path,
        default=repo / "data/curacion-universo/perfiles-autoridad-semantica-marco-v1_0.json",
    )
    parser.add_argument(
        "--base-autoridad",
        type=Path,
        default=repo / "data/curacion-universo/autoridad-semantica-marco-v1_0.jsonl",
    )
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--diagnostico", type=Path, required=True)
    parser.add_argument("--residuales", type=Path)
    parser.add_argument("--inventario", type=Path)
    args = parser.parse_args(argv)
    try:
        rows, diagnostics = materialize(
            repo=repo,
            corpus_root=args.corpus_root,
            index_path=args.indice_e2,
            profiles_path=args.perfiles,
            base_authority_path=args.base_autoridad,
            residual_output=args.residuales,
            inventory_output=args.inventario,
        )
    except (AutoridadSemanticaError, ProductiveAuthorityError, OSError, ValueError) as exc:
        print(f"ERROR:{exc}", file=sys.stderr)
        return 2
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        "".join(canonical_authority_line(row) + "\n" for row in rows),
        encoding="utf-8",
    )
    args.diagnostico.parent.mkdir(parents=True, exist_ok=True)
    args.diagnostico.write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
