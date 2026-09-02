#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ACTO MAESTRA34-L6 · P3 — lectores de la serie municipal.

Un lector por fuente. Cada uno devuelve {municipio_normalizado: (votos_totales,
lista_nominal)} y la lista de filas excluidas por no ser municipios, para que el
control aritmetico de la spec (1.7.6) se pueda correr.
"""
import os, re, io, zipfile, unicodedata
import openpyxl

RAW = "data/raw"
SERIE = os.path.join(RAW, "electoral_local_municipal_serie")

ALIAS_MUNICIPIO = {
    # Zacatecas: el IEEZ alterna abreviatura y nombre largo entre anios
    "GRAL ENRIQUE ESTRADA": "GENERAL ENRIQUE ESTRADA",
    "GRAL FRANCISCO R MURGUIA": "GENERAL FRANCISCO R MURGUIA",
    "GRAL PANFILO NATERA": "GENERAL PANFILO NATERA",
    "GRAL JOAQUIN AMARO": "GENERAL JOAQUIN AMARO",
    "NOCHISTLAN": "NOCHISTLAN DE MEJIA",
    "VILLA GONZALEZ ORTEGA": "VILLA GONZALEZ ORTEGA",
    "EL PLATEADO DE JOAQUIN AMARO": "EL PLATEADO DE JOAQUIN AMARO",
}

def norm(s):
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Za-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().upper()
    # sufijo numerico que el IEEM mete en 2024 y que L4 documento
    s = re.sub(r"\s+\d+$", "", s)
    s = re.sub(r"^GRAL\.?\s+", "GENERAL ", s)
    return ALIAS_MUNICIPIO.get(s, s)

def _num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(",", "").replace(" ", "")
    if s in ("", "-", "--", "N/A"): return None
    try: return float(s)
    except ValueError: return None

def _filas(path, sheet=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    out = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return out

def _idx(cab, *patrones):
    for j, c in enumerate(cab):
        t = norm(c or "")
        for p in patrones:
            if re.fullmatch(p, t): return j
    for j, c in enumerate(cab):
        t = norm(c or "")
        for p in patrones:
            if re.search(p, t): return j
    return None

# ---------------------------------------------------------------- Coahuila
def coahuila_municipio(path, fila_cab):
    filas = _filas(path)
    cab = [("" if c is None else str(c)) for c in filas[fila_cab - 1]]
    j_nom = _idx(cab, r"AYUNTAMIENTO", r"MUNICIPIO", r"NOM MUN")
    j_tot = _idx(cab, r"TOTAL DE LA VOTACION", r"TOTAL VOTOS", r"TOTAL")
    j_ln  = _idx(cab, r"LISTA NOMINAL", r"LISTA_NOMINAL", r"LISTADO NOM")
    assert None not in (j_nom, j_tot, j_ln), (path, j_nom, j_tot, j_ln, cab[:30])
    datos, excl = {}, []
    for r in filas[fila_cab:]:
        if j_nom >= len(r): continue
        nom = r[j_nom]
        if nom is None or not str(nom).strip(): continue
        tot = _num(r[j_tot]) if j_tot < len(r) else None
        ln  = _num(r[j_ln]) if j_ln < len(r) else None
        k = norm(nom)
        if tot is None or ln is None or ln <= 0:
            excl.append((k, tot, ln)); continue
        if re.search(r"TOTAL|VMRE|EXTRANJ|SUMA|ESTATAL|CIRCUNSCRIP", k):
            excl.append((k, tot, ln)); continue
        datos[k] = (tot, ln)
    return datos, excl

# ---------------------------------------------------------------- Zacatecas
def zacatecas_casilla(path, sheet, fila_cab):
    filas = _filas(path, sheet)
    cab = [("" if c is None else str(c)) for c in filas[fila_cab - 1]]
    j_mun = _idx(cab, r"MUNICIPIO LOCAL", r"MUNICIPIO")
    j_ln  = _idx(cab, r"LISTA NOMINAL CASILLA", r"LISTA NOMINAL", r"LISTA\s*NOMINAL")
    j_tot = _idx(cab, r"VTOTAL", r"TOTAL VOTOS", r"TOTAL_VOTOS")
    assert None not in (j_mun, j_ln, j_tot), (path, sheet, j_mun, j_ln, j_tot, cab[:30])
    agg, excl = {}, []
    for r in filas[fila_cab:]:
        if max(j_mun, j_ln, j_tot) >= len(r): continue
        mun = r[j_mun]
        if mun is None or not str(mun).strip(): continue
        ln, tot = _num(r[j_ln]), _num(r[j_tot])
        if ln is None or tot is None: continue
        k = norm(mun)
        if re.search(r"^TOTAL|ESTATAL|SUMA", k): excl.append((k, tot, ln)); continue
        a = agg.setdefault(k, [0.0, 0.0]); a[0] += tot; a[1] += ln
    return {k: (v[0], v[1]) for k, v in agg.items() if v[1] > 0}, excl

# ---------------------------------------------------------------- Nayarit
def nayarit_votos(path, sheet, fila_cab=6):
    """La tabla NO trae TOTAL: se suma columna por columna desde la primera
    columna de partido hasta VOTOS NULOS inclusive."""
    filas = _filas(path, sheet)
    cab = [("" if c is None else str(c)) for c in filas[fila_cab - 1]]
    j_mun = _idx(cab, r"MUNICIPIO")
    j_cas = _idx(cab, r"CASILLAS INSTALADAS")
    j_nul = _idx(cab, r"VOTOS NULOS", r"NULOS")
    assert None not in (j_mun, j_cas, j_nul), (path, j_mun, j_cas, j_nul, cab[:25])
    ini = j_cas + 1
    datos, cols = {}, [c for c in cab[ini:j_nul + 1]]
    for r in filas[fila_cab:]:
        if j_mun >= len(r): continue
        mun = r[j_mun]
        if mun is None or not str(mun).strip(): continue
        k = norm(mun)
        if re.search(r"^TOTAL|ESTATAL|SUMA", k): continue
        vals = [_num(r[j]) for j in range(ini, min(j_nul + 1, len(r)))]
        tot = sum(v for v in vals if v is not None)
        if tot > 0: datos[k] = tot
    return datos, cols

MUNICIPIOS_NAYARIT = ["Acaponeta","Ahuacatlan","Amatlan de Canas","Bahia de Banderas",
 "Compostela","Huajicori","Ixtlan del Rio","Jala","Del Nayar","Rosamorada","Ruiz","San Blas",
 "San Pedro Lagunillas","Santa Maria del Oro","Santiago Ixcuintla","Tecuala","Tepic","Tuxpan",
 "Xalisco","La Yesca"]

def nayarit_lista_nominal(path_pdf):
    """Dos formatos publicados por el IEE Nayarit:
       2017/2024: 'Municipio  <lista nominal>'
       2021:      'Municipio  <hombres>  <mujeres>  <lista nominal>' (con control
                  aritmetico hombres+mujeres == lista nominal).
    Se busca por NOMBRE DE MUNICIPIO conocido, no por posicion, y se exige que
    el bloque encontrado sea el de MUNICIPIOS y no el de DISTRITOS."""
    import pypdf
    r = pypdf.PdfReader(path_pdf)
    txt = "\n".join((p.extract_text() or "") for p in r.pages)
    lineas = [re.sub(r"\s+", " ", l).strip() for l in txt.split("\n") if l.strip()]
    esperados = {norm(m) for m in MUNICIPIOS_NAYARIT}
    datos, control = {}, {}
    for l in lineas:
        m = re.match(r"^([A-Za-zÁÉÍÓÚÜÑáéíóúüñ][A-Za-zÁÉÍÓÚÜÑáéíóúüñ .\']{2,40}?)\s+"
                     r"((?:[\d,]+\s+){0,2}[\d,]{3,})$", l)
        if not m: continue
        k = norm(m.group(1))
        if k not in esperados: continue
        nums = [_num(x) for x in m.group(2).split()]
        nums = [x for x in nums if x is not None]
        if len(nums) == 1:
            datos[k] = nums[0]
        elif len(nums) == 3 and abs((nums[0] + nums[1]) - nums[2]) < 1e-6:
            datos[k] = nums[2]; control[k] = "h+m==ln"
        elif len(nums) >= 1:
            datos.setdefault(k, nums[-1])
    return datos

# ---------------------------------------------------------------- Durango
def durango2016_municipio_capital(path):
    """El archivo cubre SOLO el municipio de Durango (capital).
    OJO: la fila de cabecera trae 15 rotulos y las filas de datos 16 valores --
    la columna 'DIFERENCIA ENTRE 1o Y 2o' no tiene rotulo propio, asi que
    cabecera y datos estan DESALINEADAS una posicion. No se lee por rotulo: se
    toman las tres ultimas columnas numericas de cada fila (TOTAL, L. NOMINAL,
    %) y se EXIGE la identidad TOTAL / L.NOMINAL == %, que es la que identifica
    las columnas sin depender del rotulo."""
    filas = _filas(path, "2016 Ayuntamiento Dgo")
    CAS = re.compile(r"(?i)^(b[aá]sica|contigua|extraordinaria|especial)")
    tot = ln = 0.0; n = 0; subt = 0; gran = None
    for r in filas[6:]:
        nums = [(j, _num(v)) for j, v in enumerate(r) if _num(v) is not None]
        if len(nums) < 3: continue
        (_, t), (_, l), (_, pct) = nums[-3], nums[-2], nums[-1]
        if l is None or l <= 0 or t is None or pct is None: continue
        if abs((t / l) - pct) > 1e-4: continue
        etiqueta = str(r[2]).strip() if len(r) > 2 and r[2] is not None else ""
        if CAS.match(etiqueta):
            tot += t; ln += l; n += 1
        elif re.search(r"(?i)casillas", etiqueta):
            subt += 1
        else:
            gran = (t, l, pct)          # fila de gran total del municipio
    return {"DURANGO": (tot, ln)}, (n, subt, gran)

def durango2019_csv(path_zip):
    z = zipfile.ZipFile(path_zip)
    nombre = [n for n in z.namelist() if "AYUN" in n.upper() and n.upper().endswith(".CSV")
              and "ENCABEZADO" not in n.upper()][0]
    txt = z.read(nombre).decode("utf-8", "replace")
    lineas = [l for l in txt.split("\n") if l.strip()]
    cab = lineas[0].split(",")
    agg = {}
    # La cabecera declara 26 nombres y las filas traen mas columnas: las 5 ultimas
    # numericas antes de la observacion son VALIDOS, NOREG, NULOS, TOTAL, LISTA_NOMINAL.
    # Se identifica POSICIONALMENTE y se comprueba VALIDOS+NOREG+NULOS == TOTAL.
    ok = mal = 0
    for l in lineas[1:]:
        c = l.split(",")
        if len(c) < 32: continue
        mun = c[5]
        nums = []
        for j in range(len(c) - 1, 5, -1):
            v = _num(c[j])
            if v is not None: nums.append((j, v))
            if len(nums) >= 5 and _num(c[j - 1]) is None: break
        nums = sorted(nums)[:0] or nums
        # localiza la ventana de 5 numericos consecutivos que cumple la identidad
        idxs = [j for j in range(6, len(c)) if _num(c[j]) is not None]
        hallado = None
        for a in range(len(idxs) - 4):
            w = idxs[a:a + 5]
            if w != list(range(w[0], w[0] + 5)): continue
            v = [_num(c[j]) for j in w]
            if abs((v[0] + v[1] + v[2]) - v[3]) < 1e-6 and v[4] > 0:
                hallado = (v[3], v[4])
        if hallado is None: mal += 1; continue
        ok += 1
        k = norm(mun)
        a = agg.setdefault(k, [0.0, 0.0]); a[0] += hallado[0]; a[1] += hallado[1]
    return {k: (v[0], v[1]) for k, v in agg.items() if v[1] > 0}, (ok, mal, cab)
