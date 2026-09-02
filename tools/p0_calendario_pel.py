#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ACTO MAESTRA34-L6 · P0 — tabla de tratamiento de la homologacion escalonada.

Deriva, para cada entidad federativa y cada jornada electoral local
2015-2024, si en esa jornada se eligieron AYUNTAMIENTOS y si la jornada
coincidio con la federal, a partir de los acuerdos del Consejo General del
INE que aprueban el Plan Integral y los Calendarios de Coordinacion de cada
Proceso Electoral Local (payloads en data/raw/electoral_calendario_pel_ine/).

No lee ningun resultado electoral: solo calendario y cargos. Salida:
data/p0-calendario-ayuntamientos-v1_0.tsv
"""
import os, re, sys, json, hashlib, unicodedata

RAW = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "raw",
                   "electoral_calendario_pel_ine")
SALIDA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data",
                      "p0-calendario-ayuntamientos-v1_0.tsv")
SALIDA_TRAT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data",
                           "p0-tratamiento-homologacion-v1_0.tsv")

# EXCEPCIONES DOCUMENTADAS a la regla "anio federal => jornada local concurrente".
# Cada una cita verbatim el acuerdo del CG del INE que la establece.
EXCEPCIONES_NO_CONCURRENTE = {
 ("Chiapas", 2015): (
   "el 19 de julio de 2015 se celebro la eleccion de diputados y ayuntamientos "
   "del estado de Chiapas",
   "CG extraordinaria 14/oct/2015 (handle 123456789/79710), antecedente II"),
}
# Anadidos documentados que la cadena de calendarios de coordinacion no cubre.
FILAS_DOCUMENTALES = [
 {"entidad": "Chiapas", "anio_jornada": 2015, "ayuntamientos": "SI",
  "n_actividades_municipales": 1,
  "cargos_declarados": "Diputaciones y Ayuntamientos (texto verbatim del acuerdo)",
  "ejemplo_actividad": "el 19 de julio de 2015 se celebro la eleccion de diputados y ayuntamientos del estado de Chiapas",
  "fuente_archivo": "pel2015b_CGex201510-14_ap4.pdf.txt", "fuente_hoja": "antecedente II",
  "fuente_acuerdo_ine": "CG extraordinaria 14/oct/2015", "handle_dspace": "123456789/79710"},
]

# Anios de jornada electoral FEDERAL (art. 25 LGIPE: primer domingo de junio).
ANIOS_FEDERALES = {2015, 2018, 2021, 2024}

# Cargo municipal: la actividad tiene que nombrar el cargo, no el organo del OPL.
# "Organos Municipales" (consejos municipales del OPL) queda FUERA a proposito:
# nombra la estructura, no la eleccion.
CARGO_MUNI = re.compile(
    r"(?i)(ayuntamient|alcald[ií]a|presidencias?\s+municipal|junta[s]?\s+municipal"
    r"|concejal|regidur|sindicatur|mun[ií]cipe)")
ACTO_ELECTORAL = re.compile(
    r"(?i)(campa[nñ]a|precampa[nñ]a|registro|candidatur|topes?\s+de\s+gastos"
    r"|coalici[oó]n|c[oó]mputo|boleta|apoyo\s+ciudadano|plataforma|paridad"
    r"|cargos?\s+a\s+elegir)")

ENTIDADES = [
 "Aguascalientes","Baja California","Baja California Sur","Campeche","Coahuila",
 "Colima","Chiapas","Chihuahua","Ciudad de México","Durango","Guanajuato","Guerrero",
 "Hidalgo","Jalisco","Estado de México","Michoacán","Morelos","Nayarit","Nuevo León",
 "Oaxaca","Puebla","Querétaro","Quintana Roo","San Luis Potosí","Sinaloa","Sonora",
 "Tabasco","Tamaulipas","Tlaxcala","Veracruz","Yucatán","Zacatecas"]

ALIAS = {
 "distrito federal":"Ciudad de México", "ciudad de mexico":"Ciudad de México",
 "cdmx":"Ciudad de México", "mexico":"Estado de México",
 "estado de mexico":"Estado de México", "edomex":"Estado de México",
 "coahuila de zaragoza":"Coahuila", "michoacan":"Michoacán",
 "michoacan de ocampo":"Michoacán", "nuevo leon":"Nuevo León",
 "queretaro":"Querétaro", "san luis potosi":"San Luis Potosí",
 "yucatan":"Yucatán", "veracruz de ignacio de la llave":"Veracruz",
 "sonora":"Sonora", "veracruz":"Veracruz",
}

def norm(s):
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().lower()

_NORM_ENT = {norm(e): e for e in ENTIDADES}

def canon_entidad(s):
    n = norm(s)
    if n in _NORM_ENT: return _NORM_ENT[n]
    if n in ALIAS: return ALIAS[n]
    return None

def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""): h.update(chunk)
    return h.hexdigest()

# ---------------------------------------------------------------- extractores

def _celdas_texto(ws, limite=None):
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if limite and i >= limite: break
        out.append([("" if c is None else str(c).strip()) for c in row])
    return out

def evidencia_municipal(textos):
    """Devuelve (n_fuertes, ejemplos) sobre una lista de cadenas."""
    fuertes = []
    for t in textos:
        if not t or len(t) > 300: continue
        if CARGO_MUNI.search(t) and ACTO_ELECTORAL.search(t):
            fuertes.append(re.sub(r"\s+", " ", t))
    vistos, ej = set(), []
    for f in fuertes:
        k = f[:60]
        if k in vistos: continue
        vistos.add(k); ej.append(f)
    return len(fuertes), ej[:3]

def de_hojas_por_entidad(path):
    """Hojas nombradas por entidad (2017-2018, 2020-2021-a1, 2023-2024)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    res = {}
    for sn in wb.sheetnames:
        ent = canon_entidad(sn)
        if not ent: continue
        ws = wb[sn]
        filas = _celdas_texto(ws)
        planas = [c for fila in filas for c in fila if c]
        cargos_decl = ""
        for i, fila in enumerate(filas[:4]):
            if any(re.search(r"(?i)cargos?\s+a\s+elegir", c) for c in fila):
                for sig in filas[i:i + 3]:
                    cand = " ".join(c for c in sig if c and not re.search(r"(?i)cargos?\s+a\s+elegir", c))
                    if cand.strip():
                        cargos_decl = re.sub(r"\s+", " ", cand.strip()); break
                break
        n, ej = evidencia_municipal(planas)
        res[ent] = {"cargos_declarados": cargos_decl, "n_actividades_muni": n,
                    "ejemplos": ej, "hoja": sn}
    wb.close()
    return res

def de_concentrado(path, hoja=None, col_ent="Entidad"):
    """Hoja unica con columna Entidad (2021-2022, 2022-2023, 2024-2025)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    hojas = [hoja] if hoja else [h for h in wb.sheetnames
                                 if norm(h) in ("concentrado", "calendario")]
    res = {}
    for sn in hojas:
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        filas = _celdas_texto(ws)
        if not filas: continue
        cab = None
        for i, fila in enumerate(filas[:6]):
            if any(norm(c) == norm(col_ent) for c in fila): cab = i; break
        if cab is None: continue
        idx_ent = [j for j, c in enumerate(filas[cab]) if norm(c) == norm(col_ent)][0]
        for fila in filas[cab + 1:]:
            if idx_ent >= len(fila): continue
            ent = canon_entidad(fila[idx_ent])
            if not ent: continue
            d = res.setdefault(ent, {"cargos_declarados": "", "n_actividades_muni": 0,
                                     "ejemplos": [], "hoja": sn})
            n, ej = evidencia_municipal([c for c in fila if c])
            d["n_actividades_muni"] += n
            for e in ej:
                if len(d["ejemplos"]) < 3 and e not in d["ejemplos"]: d["ejemplos"].append(e)
    wb.close()
    return res

def de_texto_por_bloques(path):
    """Anexo PDF->txt con bloques 'Calendario de Actividades ...' por entidad
    (PEL 2016-2017)."""
    t = open(path, encoding="utf-8", errors="replace").read()
    lineas = [re.sub(r"\s+", " ", l).strip() for l in t.split("\n")]
    actual, res = None, {}
    for l in lineas:
        for e in ENTIDADES:
            if re.search(r"(?<![A-Za-zÁÉÍÓÚáéíóúÑñ])" + re.escape(e) +
                         r"(?![A-Za-zÁÉÍÓÚáéíóúÑñ])", l):
                actual = e; break
        if actual is None: continue
        d = res.setdefault(actual, {"cargos_declarados": "", "n_actividades_muni": 0,
                                    "ejemplos": [], "hoja": os.path.basename(path)})
        n, ej = evidencia_municipal([l])
        d["n_actividades_muni"] += n
        for e2 in ej:
            if len(d["ejemplos"]) < 3 and e2 not in d["ejemplos"]: d["ejemplos"].append(e2)
    return res

def de_tabla_cargos_2016(path):
    """Cuadro verbatim del acuerdo CGex1 16/dic/2015 (considerando 20):
    'Entidad Ayuntamientos Diputaciones MR Diputaciones RP Gubernatura'."""
    t = re.sub(r"\s+", " ", open(path, encoding="utf-8", errors="replace").read())
    m = re.search(r"Entidad\s+Ayuntamientos\s+Diputaciones\s+MR\s+Diputaciones\s+RP\s+Gubernatura(.{0,1200})", t)
    if not m: return {}
    cuerpo = m.group(1)
    res = {}
    patron = re.compile(r"\*?(" + "|".join(re.escape(e) for e in ENTIDADES) +
                        r")\s+(--|-|\d+)\s+(--|-|\d+)\s+(--|-|\d+)\s+(--|-|\d+)")
    for mm in patron.finditer(cuerpo):
        ent = canon_entidad(mm.group(1))
        if not ent: continue
        ay = mm.group(2)
        n_ay = 0 if ay in ("--", "-") else int(ay)
        res[ent] = {"cargos_declarados": (f"Ayuntamientos={ay} DipMR={mm.group(3)} "
                                          f"DipRP={mm.group(4)} Gub={mm.group(5)} "
                                          f"(cuadro verbatim del considerando 20)"),
                    "n_actividades_muni": 1 if n_ay > 0 else 0,
                    "ejemplos": [f"cuadro del acuerdo: {ent} Ayuntamientos={ay}"],
                    "hoja": "considerando 20", "n_ayuntamientos": n_ay}
    return res

def de_lista_entidades_2015(path):
    """Punto de Acuerdo del CGor 18/dic/2014 + CGex 14/oct/2015: entidades con
    jornada local coincidente con la federal del 7/jun/2015. NO trae cargo."""
    t = re.sub(r"\s+", " ", open(path, encoding="utf-8", errors="replace").read())
    m = re.search(r"(?i)(?:estados|entidades) de ((?:[A-ZÁÉÍÓÚÑ][^,.;]{2,30}, )+[^,.;]{2,30}(?:\s+y\s+[^,.;]{2,30})?)\.", t)
    if not m: return {}
    crudo = m.group(1)
    res = {}
    for pedazo in re.split(r",\s*|\s+y\s+", crudo):
        ent = canon_entidad(pedazo.strip())
        if ent:
            res[ent] = {"cargos_declarados": "NO-DETERMINADO-POR-ESTA-FUENTE "
                                             "(el acuerdo enumera entidades, no cargos)",
                        "n_actividades_muni": -1, "ejemplos": [], "hoja": "punto de acuerdo"}
    return res

# ------------------------------------------------------------------- catalogo
# (anio de jornada, archivo, extractor, handle DSpace, acuerdo)
FUENTES = [
 (2015, "pel2014-2015_CGor201412-18_ap13.pdf.txt", de_lista_entidades_2015,
  "123456789/87134", "CG ordinaria 18/dic/2014"),
 (2015, "pel2015b_CGex201510-14_ap4.pdf.txt", de_lista_entidades_2015,
  "123456789/79710", "CG extraordinaria 14/oct/2015"),
 (2016, "pel2015-2016_CGex1_201512-16_ap25.pdf.txt", de_tabla_cargos_2016,
  "123456789/87457", "CG extraordinaria 16/dic/2015, considerando 20"),
 (2017, "pel2016-2017_CGex201609-7-ap-9-a1.pdf.txt", de_texto_por_bloques,
  "123456789/85998", "CG extraordinaria 7/sep/2016, anexo 1"),
 (2018, "pel2017-2018.xlsx", de_hojas_por_entidad,
  "123456789/93570", "CG extraordinaria 8/sep/2017, anexo 2"),
 (2019, "pel2018-2019ord_CGex201808-6-ap-6-a1.xlsx", de_hojas_por_entidad,
  "123456789/97991", "CG extraordinaria 6/ago/2018, anexo 1 (Aguascalientes)"),
 (2019, "pel2018-2019ord_CGex201808-6-ap-6-a2.xlsx", de_hojas_por_entidad,
  "123456789/97991", "CG extraordinaria 6/ago/2018, anexo 2 (Baja California)"),
 (2019, "pel2018-2019ord_CGex201808-6-ap-6-a3.xlsx", de_hojas_por_entidad,
  "123456789/97991", "CG extraordinaria 6/ago/2018, anexo 3 (Durango)"),
 (2019, "pel2018-2019ord_CGex201808-6-ap-6-a4.xlsx", de_hojas_por_entidad,
  "123456789/97991", "CG extraordinaria 6/ago/2018, anexo 4 (Quintana Roo)"),
 (2019, "pel2018-2019ord_CGex201808-6-ap-6-a5.xlsx", de_hojas_por_entidad,
  "123456789/97991", "CG extraordinaria 6/ago/2018, anexo 5 (Tamaulipas)"),
 (2025, "pel2024-2025_CGor202409-26-ap-03-a2.xlsx", de_concentrado,
  "123456789/176887", "CG ordinaria 26/sep/2024, anexo 2"),
 (2020, "pel2019-2020_CGex201909-30-ap-1-a1.xlsx", de_hojas_por_entidad,
  "123456789/112696", "CG extraordinaria 30/sep/2019, anexo 1"),
 (2021, "pel2020-2021_CGex202008-07-ap-2-a1.xlsm", de_hojas_por_entidad,
  "123456789/114312", "CG extraordinaria 7/ago/2020, anexo 1"),
 (2022, "pel2021-2022.xlsx", de_concentrado,
  "123456789/122210", "CG ordinaria 28/jul/2021, anexo 1"),
 (2023, "pel2022-2023.xlsx", de_concentrado,
  "123456789/143140", "CG extraordinaria 26/sep/2022, anexo C"),
 (2024, "pel2023-2024.xlsx", de_hojas_por_entidad,
  "123456789/152565", "CG extraordinaria 20/jul/2023, anexo 2"),
]

def fin_campana_por_ciclo(path):
    """Verificacion A.13 de que NINGUNA entidad del ciclo se sale de la fecha
    comun de jornada: fecha de fin de la ultima campana, por entidad."""
    import datetime, collections
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    CAMP = re.compile(r"(?i)^campa[nñ]a")
    res = collections.defaultdict(list)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(values_only=True):
            txts = [("" if c is None else str(c).strip()) for c in row]
            fechas = [c for c in row if isinstance(c, datetime.datetime)]
            if fechas and any(CAMP.search(t) for t in txts):
                clave = canon_entidad(sn) or canon_entidad(txts[0] if txts else "") or sn
                res[clave].append(max(fechas).date())
    wb.close()
    return {k: max(v) for k, v in res.items() if v}

def main():
    filas, meta = [], []
    for anio, arch, fn, handle, acuerdo in FUENTES:
        path = os.path.join(RAW, arch)
        if not os.path.exists(path):
            print(f"AUSENTE: {path}", file=sys.stderr); continue
        try:
            res = fn(path)
        except Exception as e:
            print(f"ERROR {arch}: {type(e).__name__}: {e}", file=sys.stderr); continue
        meta.append({"anio": anio, "archivo": arch, "sha256": sha256(path),
                     "bytes": os.path.getsize(path), "handle": handle,
                     "acuerdo": acuerdo, "entidades": len(res)})
        fincamp = fin_campana_por_ciclo(path) if arch.lower().endswith((".xlsx", ".xlsm")) else {}
        fechas_distintas = sorted({str(v) for k, v in fincamp.items() if canon_entidad(str(k))})
        meta[-1]["fin_campana_distintos"] = fechas_distintas
        meta[-1]["fin_campana_n_entidades"] = len([k for k in fincamp if canon_entidad(str(k))])
        for ent, d in sorted(res.items()):
            exc = EXCEPCIONES_NO_CONCURRENTE.get((ent, anio))
            conc = "NO" if (anio not in ANIOS_FEDERALES or exc) else "SI"
            filas.append({
                "entidad": ent, "anio_jornada": anio,
                "concurrente_con_federal": conc,
                "ayuntamientos": ("SI" if d["n_actividades_muni"] > 0 else
                                  ("INDETERMINADO" if d["n_actividades_muni"] < 0 else "NO")),
                "n_actividades_municipales": d["n_actividades_muni"],
                "cargos_declarados": d["cargos_declarados"],
                "ejemplo_actividad": (d["ejemplos"][0] if d["ejemplos"] else ""),
                "fuente_archivo": arch, "fuente_hoja": d["hoja"],
                "fuente_acuerdo_ine": acuerdo, "handle_dspace": handle,
                "nota_concurrencia": (f"EXCEPCION documentada: {exc[0]} [{exc[1]}]" if exc
                                      else ("anio no federal" if anio not in ANIOS_FEDERALES
                                            else "anio federal, sin excepcion documentada"))})
    for fd in FILAS_DOCUMENTALES:
        if any(f["entidad"] == fd["entidad"] and f["anio_jornada"] == fd["anio_jornada"]
               and f["ayuntamientos"] == "SI" for f in filas):
            continue
        filas = [f for f in filas if not (f["entidad"] == fd["entidad"]
                                          and f["anio_jornada"] == fd["anio_jornada"])]
        exc = EXCEPCIONES_NO_CONCURRENTE.get((fd["entidad"], fd["anio_jornada"]))
        r = dict(fd)
        r["concurrente_con_federal"] = ("NO" if (fd["anio_jornada"] not in ANIOS_FEDERALES or exc)
                                        else "SI")
        r["nota_concurrencia"] = (f"EXCEPCION documentada: {exc[0]} [{exc[1]}]" if exc
                                  else "anio no federal")
        filas.append(r)
    cols = ["entidad","anio_jornada","concurrente_con_federal","ayuntamientos",
            "n_actividades_municipales","cargos_declarados","ejemplo_actividad",
            "fuente_archivo","fuente_hoja","fuente_acuerdo_ine","handle_dspace",
            "nota_concurrencia"]
    with open(SALIDA, "w", encoding="utf-8") as f:
        f.write("# GENERADO por tools/p0_calendario_pel.py (ACTO MAESTRA34-L6, P0). "
                "Fuente: acuerdos del Consejo General del INE, repositoriodocumental.ine.mx. "
                "No editar a mano.\n")
        f.write("\t".join(cols) + "\n")
        for r in sorted(filas, key=lambda r: (r["entidad"], r["anio_jornada"])):
            f.write("\t".join(str(r[c]).replace("\t", " ").replace("\n", " ") for c in cols) + "\n")
    # ---- tabla de tratamiento por entidad ----
    porent = {}
    for r in filas:
        if r["ayuntamientos"] != "SI": continue
        porent.setdefault(r["entidad"], []).append((int(r["anio_jornada"]),
                                                    r["concurrente_con_federal"]))
    trat_cols = ["entidad","elecciones_ayuntamiento_en_ventana","anios_no_concurrente",
                 "anios_concurrente","anio_tratamiento","cohorte","estatus",
                 "tiene_antes_y_despues","n_elecciones"]
    tfilas = []
    for ent in ENTIDADES:
        secu = sorted(porent.get(ent, []))
        nc = [a for a, c in secu if c == "NO"]
        cc = [a for a, c in secu if c == "SI"]
        indet = any(r["entidad"] == ent and r["ayuntamientos"] == "INDETERMINADO"
                    for r in filas)
        if not secu:
            estatus, g = "SIN-DATO", ""
        elif not cc:
            estatus, g = "NUNCA-TRATADO", ""
        elif not nc:
            estatus = "SIEMPRE-CONCURRENTE-EN-VENTANA" + ("-CON-2015-INDETERMINADO" if indet else "")
            g = ""
        else:
            g = min(a for a in cc if a > min(nc)) if any(a > min(nc) for a in cc) else ""
            estatus = "TRATADO" if g else "MIXTO"
        antes_y_despues = "SI" if (g and any(a < g for a in nc) and g in cc) else "NO"
        tfilas.append({"entidad": ent,
                       "elecciones_ayuntamiento_en_ventana": ",".join(str(a) for a, _ in secu),
                       "anios_no_concurrente": ",".join(map(str, nc)),
                       "anios_concurrente": ",".join(map(str, cc)),
                       "anio_tratamiento": g, "cohorte": (f"g{g}" if g else ""),
                       "estatus": estatus, "tiene_antes_y_despues": antes_y_despues,
                       "n_elecciones": len(secu)})
    with open(SALIDA_TRAT, "w", encoding="utf-8") as f:
        f.write("# GENERADO por tools/p0_calendario_pel.py (ACTO MAESTRA34-L6, P0). "
                "Derivada de p0-calendario-ayuntamientos-v1_0.tsv. No editar a mano.\n")
        f.write("\t".join(trat_cols) + "\n")
        for r in tfilas:
            f.write("\t".join(str(r[c]) for c in trat_cols) + "\n")
    resumen = {}
    for r in tfilas: resumen[r["estatus"]] = resumen.get(r["estatus"], 0) + 1
    print(json.dumps({"filas_calendario": len(filas), "entidades_tratamiento": len(tfilas),
                      "estatus": resumen,
                      "tratados_con_antes_y_despues": sum(1 for r in tfilas
                                                          if r["tiene_antes_y_despues"] == "SI"),
                      "fuentes": meta}, ensure_ascii=False, indent=1))

if __name__ == "__main__":
    main()
