"""Proyectores de metadatos primarios para las familias estructuradas de E2.

El modulo no decide encuesta, ola, universo, exclusividad ni ponderador. Solo
lee el artefacto registrado en ``data/manifiesto.yaml`` y prueba que metadatos
como etiqueta, tipo de almacenamiento, dominio, pares codigo-etiqueta y
user-missing estan (o no estan) en la fuente primaria. Esa separacion evita
convertir una limitacion del resumen E2 en una carencia epistemologica de la
fuente.
"""

from __future__ import annotations

import csv
import hashlib
import io
import math
import re
import struct
import tempfile
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Mapping
from zipfile import BadZipFile, ZipFile

import olefile
import yaml
from openpyxl import load_workbook
from pandas.io.stata import StataMissingValue, StataReader


DTA_PROJECTOR = "DTA-VALUE-LABELS-PANDAS-1.0"
SAV_PROJECTOR = "SAV-DICCIONARIO-CLASICO-1.0"
CSV_CATALOG_PROJECTOR = "CSV-DICCIONARIO-CATALOGO-1.0"
XLSX_PROJECTOR = "XLSX-DICCIONARIO-OPENPYXL-1.0"
XLS_PROJECTOR = "XLS-DICCIONARIO-BIFF8-1.0"

PROJECTED = "METADATA_PRIMARIA_PROYECTADA"
NOT_PROJECTABLE = "METADATA_PRIMARIA_NO_PROYECTABLE"
SOURCE_NOT_PRESENT = "FUENTE_PRIMARIA_NO_PRESENTE"

_ZIP_MEMBER_RE = re.compile(r"(?:^|!/)miembro=(\d+):([^!]+)")
_ROW_RE = re.compile(r"#diccionario-fila=(\d+):variable=")
_XLS_SHEET_RE = re.compile(r"/hoja=(\d+):nombre_sha256=")

_MISSING_LABELS = {
    "NO CONTESTA",
    "NO CONTESTO",
    "NO DECLARADA",
    "NO DECLARADO",
    "NO ESPECIFICADA",
    "NO ESPECIFICADO",
    "NO RECUERDA",
    "NO RESPONDE",
    "NO SABE",
    "NO SABE / NO RESPONDE",
    "NO SABE/NO RESPONDE",
    "SE IGNORA",
    "SIN INFORMACION",
}


def _text(value: Any) -> str:
    return str(value or "").strip()


def _fold(value: Any) -> str:
    import unicodedata

    normalized = unicodedata.normalize("NFKD", _text(value))
    return " ".join(
        "".join(
            character
            for character in normalized
            if not unicodedata.combining(character)
        )
        .upper()
        .split()
    )


def _is_missing_label(value: Any) -> bool:
    return _fold(value) in _MISSING_LABELS


def _sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _decode(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="replace")


def _number(value: Any) -> str:
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value):
            return "."
        return str(int(value)) if value.is_integer() else format(value, ".15g")
    return _text(value)


@dataclass(frozen=True)
class MetadataProjection:
    projector: str
    status: str
    recovered_fields: tuple[str, ...] = ()
    code_label_pairs: int = 0
    user_missing: int = 0
    survey_title_documented: bool = False
    source_locator: str = ""
    technical_reason: str = ""

    def as_summary(self) -> dict[str, Any]:
        summary = {
            "proyector": self.projector,
            "estado": self.status,
            "campos_recuperados": list(self.recovered_fields),
            "pares_codigo_etiqueta": self.code_label_pairs,
            "user_missing": self.user_missing,
            "encuesta_titulo_estructurado": self.survey_title_documented,
            "localizador_fuente": self.source_locator,
        }
        if self.status == NOT_PROJECTABLE:
            summary["razon_tecnica"] = self.technical_reason
        elif self.status == SOURCE_NOT_PRESENT:
            summary["razon_fuente_no_presente"] = self.technical_reason
        return summary


def _mapping_strings(value: Any) -> list[str]:
    if isinstance(value, Mapping):
        return [
            item
            for nested in value.values()
            for item in _mapping_strings(nested)
        ]
    if isinstance(value, list):
        return [item for nested in value for item in _mapping_strings(nested)]
    return [value] if isinstance(value, str) else []


def _design_contexts_by_payload(
    manifest: Mapping[str, Mapping[str, Any]], design_text: str
) -> dict[str, dict[str, Any]]:
    document = yaml.safe_load(design_text)
    entries = document if isinstance(document, list) else []
    contexts: dict[str, dict[str, Any]] = {}
    for payload_id in manifest:
        token = re.compile(
            rf"(?<![A-Za-z0-9_]){re.escape(payload_id)}(?![A-Za-z0-9_])"
        )
        matching = [
            entry
            for entry in entries
            if isinstance(entry, Mapping)
            and any(token.search(value) for value in _mapping_strings(entry))
        ]
        if not matching:
            continue
        contexts[payload_id] = {
            "fuentes": sorted(
                {_text(entry.get("fuente")) for entry in matching if entry.get("fuente")}
            ),
            "estados": sorted(
                {_text(entry.get("estado")) for entry in matching if entry.get("estado")}
            ),
            "ponderadores_documentados": sorted(
                {
                    _text(entry.get("ponderador"))
                    for entry in matching
                    if entry.get("ponderador")
                }
            ),
        }
    return contexts


class PrimaryMetadataRegistry:
    """Cachea una lectura primaria por miembro/hoja, nunca por encuesta."""

    def __init__(
        self,
        *,
        corpus_root: Path,
        manifest: list[Mapping[str, Any]],
        design_text: str,
        structural_locators: Mapping[str, str] | None = None,
        source_roots: Mapping[str, Path] | None = None,
    ) -> None:
        self.corpus_root = corpus_root
        self.source_roots = {
            "": corpus_root,
            "data_raw": corpus_root,
            **{
                _text(name): Path(path)
                for name, path in (source_roots or {}).items()
                if _text(name)
            },
        }
        self.manifest = {
            _text(row.get("id")): dict(row)
            for row in manifest
            if isinstance(row, Mapping) and _text(row.get("id"))
        }
        self._design_contexts = _design_contexts_by_payload(
            self.manifest, design_text
        )
        self.design_linked_payloads = frozenset(self._design_contexts)
        self.structural_locators = dict(structural_locators or {})
        self._source_checks: dict[str, tuple[Path | None, str]] = {}
        self._member_metadata: dict[tuple[str, str, str], Any] = {}
        self._zip_names: dict[str, tuple[str, ...]] = {}
        self._structured_titles: dict[tuple[str, str], bool] = {}

    def has_design_id_link(self, payload_id: str) -> bool:
        return payload_id in self.design_linked_payloads

    def design_context(self, payload_id: str) -> dict[str, Any] | None:
        context = self._design_contexts.get(payload_id)
        return dict(context) if context is not None else None

    def _source(self, payload_id: str) -> tuple[Path | None, str]:
        cached = self._source_checks.get(payload_id)
        if cached is not None:
            return cached
        entry = self.manifest.get(payload_id)
        if entry is None:
            result = (None, "MANIFIESTO_ID_AUSENTE")
        else:
            root_name = _text(entry.get("raiz"))
            root = self.source_roots.get(root_name)
            path = root / _text(entry.get("archivo")) if root is not None else None
            if path is None or not path.is_file():
                root_label = root_name or "data_raw"
                result = (
                    None,
                    f"ARTEFACTO_PRIMARIO_AUSENTE:raiz={root_label}",
                )
            else:
                expected = _text(entry.get("sha256")).lower()
                observed = _sha256_path(path)
                if not expected or observed != expected:
                    result = (None, "ARTEFACTO_PRIMARIO_SHA_DIVERGENTE")
                else:
                    result = (path, "")
        self._source_checks[payload_id] = result
        return result

    def _effective_locator(self, record: Mapping[str, Any]) -> str:
        locator = _text(record.get("localizador"))
        if locator and locator != "[REDACTADO-PRIVACIDAD]":
            return locator
        return self.structural_locators.get(_text(record.get("objeto_padre_id")), "")

    def _member_locators(
        self, record: Mapping[str, Any]
    ) -> tuple[tuple[int, str], ...]:
        matches = list(_ZIP_MEMBER_RE.finditer(self._effective_locator(record)))
        return tuple((int(match.group(1)), match.group(2)) for match in matches)

    def _member_locator(self, record: Mapping[str, Any]) -> tuple[int, str] | None:
        located = self._member_locators(record)
        return located[-1] if located else None

    def _member_key(self, record: Mapping[str, Any], fallback: str) -> str:
        located = self._member_locators(record)
        if not located:
            scope = _text(record.get("objeto_padre_id"))
            return f"{fallback}#scope={scope}" if scope else fallback
        return "!/".join(f"{ordinal}:{name}" for ordinal, name in located)

    def _zip_member(self, source: Path, record: Mapping[str, Any]) -> tuple[str, bytes]:
        located = self._member_locators(record)
        if not located:
            raise ValueError("LOCALIZADOR_MIEMBRO_AUSENTE")
        payload: bytes | None = None
        actual_name = ""
        for depth, (ordinal, _registered_name) in enumerate(located):
            archive_input: Path | io.BytesIO
            archive_input = source if depth == 0 else io.BytesIO(payload or b"")
            with ZipFile(archive_input) as archive:
                infos = archive.infolist()
                if ordinal < 1 or ordinal > len(infos):
                    raise ValueError("LOCALIZADOR_MIEMBRO_ORDINAL_FUERA_RANGO")
                info = infos[ordinal - 1]
                actual_name = info.filename
                payload = archive.read(info)
        assert payload is not None
        return actual_name, payload

    def _material_path(
        self, source: Path, record: Mapping[str, Any], expected_suffix: str
    ) -> tuple[Path, str, bool]:
        if source.suffix.casefold() == expected_suffix:
            return source, source.name, False
        try:
            name, payload = self._zip_member(source, record)
        except ValueError as exc:
            if str(exc) != "LOCALIZADOR_MIEMBRO_AUSENTE":
                raise
            candidates = _archive_suffix_candidates(source, expected_suffix)
            if len(candidates) != 1 and expected_suffix in {".xls", ".xlsx"}:
                candidates = [
                    candidate
                    for candidate in candidates
                    if _spreadsheet_candidate_matches(
                        candidate[1], expected_suffix, record
                    )
                ]
            if len(candidates) != 1:
                raise ValueError(
                    "LOCALIZADOR_MIEMBRO_NO_RESUELTO_POR_ESTRUCTURA:"
                    f"coincidencias={len(candidates)}"
                )
            name, payload = candidates[0]
        suffix = Path(name).suffix.casefold()
        if suffix != expected_suffix:
            raise ValueError(f"FORMATO_MIEMBRO_INESPERADO:{suffix}")
        temporary = tempfile.NamedTemporaryFile(suffix=expected_suffix, delete=False)
        try:
            temporary.write(payload)
            temporary.close()
        except Exception:
            temporary.close()
            Path(temporary.name).unlink(missing_ok=True)
            raise
        return Path(temporary.name), name, True

    def project(self, record: Mapping[str, Any]) -> MetadataProjection:
        payload_id = _text(record.get("payload_id"))
        kind = _text(record.get("objeto_tipo")).upper()
        source, failure = self._source(payload_id)
        if source is None:
            return MetadataProjection(
                projector=self._projector_for(kind),
                status=(
                    SOURCE_NOT_PRESENT
                    if failure.startswith(("ARTEFACTO_PRIMARIO_AUSENTE", "MANIFIESTO_ID_AUSENTE"))
                    else NOT_PROJECTABLE
                ),
                source_locator=_text(record.get("localizador")),
                technical_reason=failure,
            )
        try:
            if kind == "VARIABLE-DTA":
                return self._project_dta(payload_id, source, record)
            if kind == "VARIABLE-SAV":
                return self._project_sav(payload_id, source, record)
            if kind == "VARIABLE-DICCIONARIO":
                return self._project_csv_catalog(payload_id, source, record)
            if kind == "VARIABLE-DICCIONARIO-XLSX":
                return self._project_xlsx(payload_id, source, record)
            if kind == "VARIABLE-DICCIONARIO-XLS":
                return self._project_xls(payload_id, source, record)
            return MetadataProjection(
                projector="NO-APLICA",
                status=NOT_PROJECTABLE,
                source_locator=_text(record.get("localizador")),
                technical_reason=f"OBJETO_TIPO_NO_SOPORTADO:{kind}",
            )
        except (
            BadZipFile,
            EOFError,
            KeyError,
            OSError,
            RuntimeError,
            TypeError,
            ValueError,
            struct.error,
            csv.Error,
        ) as exc:
            return MetadataProjection(
                projector=self._projector_for(kind),
                status=NOT_PROJECTABLE,
                source_locator=_text(record.get("localizador")),
                technical_reason=f"{type(exc).__name__}:{str(exc)[:160]}",
            )

    @staticmethod
    def _projector_for(kind: str) -> str:
        return {
            "VARIABLE-DTA": DTA_PROJECTOR,
            "VARIABLE-SAV": SAV_PROJECTOR,
            "VARIABLE-DICCIONARIO": CSV_CATALOG_PROJECTOR,
            "VARIABLE-DICCIONARIO-XLSX": XLSX_PROJECTOR,
            "VARIABLE-DICCIONARIO-XLS": XLS_PROJECTOR,
        }.get(kind, "NO-APLICA")

    def _project_dta(
        self, payload_id: str, source: Path, record: Mapping[str, Any]
    ) -> MetadataProjection:
        member_key = self._member_key(record, source.name)
        key = (payload_id, member_key, DTA_PROJECTOR)
        variables = self._member_metadata.get(key)
        if variables is None:
            path, _name, temporary = self._material_path(source, record, ".dta")
            try:
                variables = _read_dta_dictionary(path)
            finally:
                if temporary:
                    path.unlink(missing_ok=True)
            self._member_metadata[key] = variables
        name = _text(record.get("nombre"))
        metadata = variables.get(name)
        if metadata is None:
            raise ValueError("VARIABLE_NO_RESUELTA_EN_DTA")
        fields = ["tipo_almacenamiento"]
        if metadata["label"]:
            fields.append("etiqueta_variable")
        if metadata["pairs"]:
            fields.append("codigo_etiqueta")
        if metadata["missing"]:
            fields.append("user_missing")
        return MetadataProjection(
            projector=DTA_PROJECTOR,
            status=PROJECTED,
            recovered_fields=tuple(fields),
            code_label_pairs=len(metadata["pairs"]),
            user_missing=len(metadata["missing"]),
            source_locator=member_key,
        )

    def _project_sav(
        self, payload_id: str, source: Path, record: Mapping[str, Any]
    ) -> MetadataProjection:
        member_key = self._member_key(record, source.name)
        key = (payload_id, member_key, SAV_PROJECTOR)
        variables = self._member_metadata.get(key)
        if variables is None:
            path, _name, temporary = self._material_path(source, record, ".sav")
            try:
                variables = _read_sav_dictionary(path)
            finally:
                if temporary:
                    path.unlink(missing_ok=True)
            self._member_metadata[key] = variables
        name = _text(record.get("nombre"))
        metadata = variables.get(name)
        if metadata is None:
            raise ValueError("VARIABLE_NO_RESUELTA_EN_SAV")
        fields = ["tipo_almacenamiento"]
        if metadata["label"]:
            fields.append("etiqueta_variable")
        if metadata["pairs"]:
            fields.append("codigo_etiqueta")
        if metadata["missing"]:
            fields.append("user_missing")
        return MetadataProjection(
            projector=SAV_PROJECTOR,
            status=PROJECTED,
            recovered_fields=tuple(fields),
            code_label_pairs=len(metadata["pairs"]),
            user_missing=len(metadata["missing"]),
            source_locator=member_key,
        )

    def _zip_member_names(self, payload_id: str, source: Path) -> tuple[str, ...]:
        names = self._zip_names.get(payload_id)
        if names is None:
            with ZipFile(source) as archive:
                names = tuple(info.filename for info in archive.infolist())
            self._zip_names[payload_id] = names
        return names

    def _structured_title(
        self, payload_id: str, source: Path, prefix: str
    ) -> bool:
        key = (payload_id, prefix)
        cached = self._structured_titles.get(key)
        if cached is not None:
            return cached
        candidates = [
            name
            for name in self._zip_member_names(payload_id, source)
            if PurePosixPath(name).parent.as_posix().casefold()
            == f"{prefix}/metadatos".casefold()
            and Path(name).suffix.casefold() == ".txt"
        ]
        title = False
        if len(candidates) == 1:
            with ZipFile(source) as archive:
                for line in _decode(archive.read(candidates[0])).splitlines():
                    if line.startswith("Title:") and line.partition(":")[2].strip():
                        title = True
                        break
        self._structured_titles[key] = title
        return title

    def _project_csv_catalog(
        self, payload_id: str, source: Path, record: Mapping[str, Any]
    ) -> MetadataProjection:
        located = self._member_locator(record)
        if located is None:
            raise ValueError("DICCIONARIO_CSV_NO_ESTA_EN_ZIP")
        _ordinal, _registered_name = located
        dictionary_name, dictionary_payload = self._zip_member(source, record)
        dictionary_key = (payload_id, dictionary_name, CSV_CATALOG_PROJECTOR)
        dictionary_rows = self._member_metadata.get(dictionary_key)
        if dictionary_rows is None:
            dictionary_rows = list(
                csv.reader(io.StringIO(_decode(dictionary_payload), newline=""))
            )
            self._member_metadata[dictionary_key] = dictionary_rows
        row_match = _ROW_RE.search(self._effective_locator(record))
        variable = _text(record.get("nombre"))
        if row_match is not None:
            # El extractor CSV registra el índice cero-basado de la fila de
            # datos en el contrato INEGI común; otros extractores registran
            # la fila física. Ambos se validan contra el mnemónico exacto.
            row_numbers = [int(row_match.group(1)) + 1]
        else:
            row_numbers = [
                number
                for number, row in enumerate(dictionary_rows, 1)
                if variable in {_text(value) for value in row if _text(value)}
            ]
        if not row_numbers or any(
            number < 1 or number > len(dictionary_rows) for number in row_numbers
        ):
            row_numbers = []
        if row_numbers and any(
            variable
            not in {
                _text(value)
                for value in dictionary_rows[number - 1]
                if _text(value)
            }
            for number in row_numbers
        ):
            row_numbers = [
                number
                for number, row in enumerate(dictionary_rows, 1)
                if variable in {_text(value) for value in row if _text(value)}
            ]
            if not row_numbers:
                raise ValueError("VARIABLE_NO_RESUELTA_EN_DICCIONARIO_CSV")
        if not row_numbers:
            row_numbers = [
                number
                for number, row in enumerate(dictionary_rows, 1)
                if variable in {_text(value) for value in row if _text(value)}
            ]
        if not row_numbers:
            raise ValueError("VARIABLE_NO_RESUELTA_EN_DICCIONARIO_CSV")
        marker = "/diccionario_de_datos/"
        folded = dictionary_name.casefold()
        if marker not in folded:
            raise ValueError("CONTRATO_DICCIONARIO_CSV_NO_RECONOCIDO")
        marker_index = folded.index(marker)
        prefix = dictionary_name[:marker_index]
        wanted = f"{prefix}/catalogos/{variable}.csv".casefold()
        matching = [
            name
            for name in self._zip_member_names(payload_id, source)
            if name.casefold() == wanted
        ]
        pairs: list[tuple[str, str]] = []
        missing: list[str] = []
        fields = ["etiqueta_variable", "dominio_declarado"]
        locator = (
            dictionary_name
            + "#filas="
            + ",".join(str(number) for number in row_numbers)
        )
        if len(matching) == 1:
            with ZipFile(source) as archive:
                rows = list(
                    csv.reader(
                        io.StringIO(_decode(archive.read(matching[0])), newline="")
                    )
                )
            for row in rows[1:]:
                if len(row) < 2 or not _text(row[0]) or not _text(row[1]):
                    continue
                code, label = _text(row[0]), _text(row[1])
                if _is_missing_label(label):
                    missing.append(code)
                else:
                    pairs.append((code, label))
            if pairs:
                fields.append("codigo_etiqueta")
            if missing:
                fields.append("user_missing")
            locator = matching[0]
        title = self._structured_title(payload_id, source, prefix)
        if title:
            fields.append("titulo_encuesta_estructurado")
        return MetadataProjection(
            projector=CSV_CATALOG_PROJECTOR,
            status=PROJECTED,
            recovered_fields=tuple(fields),
            code_label_pairs=len(pairs),
            user_missing=len(missing),
            survey_title_documented=title,
            source_locator=locator,
        )

    def _project_xlsx(
        self, payload_id: str, source: Path, record: Mapping[str, Any]
    ) -> MetadataProjection:
        member_key = self._member_key(record, source.name)
        key = (payload_id, member_key, XLSX_PROJECTOR)
        workbook_data = self._member_metadata.get(key)
        if workbook_data is None:
            path, _name, temporary = self._material_path(source, record, ".xlsx")
            try:
                workbook = load_workbook(path, read_only=True, data_only=True)
                workbook_data = {
                    sheet.title: [tuple(row) for row in sheet.iter_rows(values_only=True)]
                    for sheet in workbook.worksheets
                }
                workbook.close()
            finally:
                if temporary:
                    path.unlink(missing_ok=True)
            self._member_metadata[key] = workbook_data
        sheet_name = _text(record.get("hoja"))
        rows = workbook_data.get(sheet_name)
        if rows is None:
            raise ValueError("HOJA_NO_RESUELTA_EN_XLSX")
        match = _ROW_RE.search(self._effective_locator(record))
        if match is None:
            row_numbers = _matching_dictionary_rows(rows, record)
        else:
            row_numbers = [int(match.group(1))]
        if any(row_number < 1 or row_number > len(rows) for row_number in row_numbers):
            raise ValueError("FILA_DICCIONARIO_FUERA_RANGO")
        values = set().union(
            *(
                _dictionary_block_values(rows, row_number, record)
                for row_number in row_numbers
            )
        )
        categories = _documented_categories(record)
        fields = ["etiqueta_variable"]
        pairs = 0
        if record.get("categorias") and values:
            fields.append("dominio_declarado")
        if len(categories) == 2:
            fields.append("codigo_etiqueta")
            pairs = 1
        return MetadataProjection(
            projector=XLSX_PROJECTOR,
            status=PROJECTED,
            recovered_fields=tuple(fields),
            code_label_pairs=pairs,
            source_locator=(
                f"{member_key}#hoja={sheet_name};filas="
                + ",".join(str(row_number) for row_number in row_numbers)
            ),
        )

    def _project_xls(
        self, payload_id: str, source: Path, record: Mapping[str, Any]
    ) -> MetadataProjection:
        member_key = self._member_key(record, source.name)
        key = (payload_id, member_key, XLS_PROJECTOR)
        sheets = self._member_metadata.get(key)
        if sheets is None:
            path, _name, temporary = self._material_path(source, record, ".xls")
            try:
                sheet_rows, sheet_names = _read_biff_dictionary_rows(path)
                sheets = {"rows": sheet_rows, "names": sheet_names}
            finally:
                if temporary:
                    path.unlink(missing_ok=True)
            self._member_metadata[key] = sheets
        effective_locator = self._effective_locator(record)
        sheet_match = _XLS_SHEET_RE.search(effective_locator)
        row_match = _ROW_RE.search(effective_locator)
        sheet_ordinal = int(sheet_match.group(1)) if sheet_match else 0
        sheet_rows = sheets["rows"]
        if not sheet_ordinal:
            wanted_sheet = _text(record.get("hoja"))
            matching_sheets = [
                ordinal
                for ordinal, name in sheets["names"].items()
                if name == wanted_sheet
            ]
            if len(matching_sheets) == 1:
                sheet_ordinal = matching_sheets[0]
            else:
                matching_rows = [
                    (ordinal, _unique_dictionary_row_map(rows, record, allow_none=True))
                    for ordinal, rows in sheet_rows.items()
                ]
                matching_rows = [item for item in matching_rows if item[1] is not None]
                if len(matching_rows) != 1:
                    raise ValueError(
                        "LOCALIZADOR_HOJA_XLS_NO_RESUELTO:"
                        f"coincidencias={len(matching_rows)}"
                    )
                sheet_ordinal = matching_rows[0][0]
        rows = sheet_rows.get(sheet_ordinal)
        if rows is None:
            raise ValueError("HOJA_DICCIONARIO_NO_DECODIFICADA_EN_XLS")
        if row_match is None:
            row_numbers = _matching_dictionary_rows_map(rows, record)
        else:
            row_numbers = [int(row_match.group(1))]
        if any(row_number not in rows for row_number in row_numbers):
            raise ValueError("FILA_DICCIONARIO_NO_DECODIFICADA_EN_XLS")
        values = set().union(
            *(
                _dictionary_block_values_map(rows, row_number, record)
                for row_number in row_numbers
            )
        )
        categories = _documented_categories(record)
        fields = ["etiqueta_variable"]
        pairs = 0
        if record.get("categorias") and values:
            fields.append("dominio_declarado")
        if len(categories) == 2:
            fields.append("codigo_etiqueta")
            pairs = 1
        return MetadataProjection(
            projector=XLS_PROJECTOR,
            status=PROJECTED,
            recovered_fields=tuple(fields),
            code_label_pairs=pairs,
            source_locator=(
                f"{member_key}#hoja={sheet_ordinal};filas="
                + ",".join(str(row_number) for row_number in row_numbers)
            ),
        )


def _archive_suffix_candidates(
    source: Path, expected_suffix: str
) -> list[tuple[str, bytes]]:
    candidates: list[tuple[str, bytes]] = []

    def visit(archive_input: Path | io.BytesIO, prefix: str, depth: int) -> None:
        with ZipFile(archive_input) as archive:
            for ordinal, info in enumerate(archive.infolist(), 1):
                if info.is_dir():
                    continue
                payload = archive.read(info)
                locator = f"{prefix}!/{ordinal}:{info.filename}"
                suffix = Path(info.filename).suffix.casefold()
                if suffix == expected_suffix:
                    candidates.append((locator, payload))
                elif suffix == ".zip" and depth < 3:
                    try:
                        visit(io.BytesIO(payload), locator, depth + 1)
                    except BadZipFile:
                        continue

    visit(source, source.name, 0)
    return candidates


def _spreadsheet_candidate_matches(
    payload: bytes, suffix: str, record: Mapping[str, Any]
) -> bool:
    temporary = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    try:
        temporary.write(payload)
        temporary.close()
        path = Path(temporary.name)
        wanted_sheet = _text(record.get("hoja"))
        if suffix == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                if wanted_sheet not in workbook.sheetnames:
                    return False
                return any(
                    _row_matches_record(tuple(row), record)
                    for row in workbook[wanted_sheet].iter_rows(values_only=True)
                )
            finally:
                workbook.close()
        rows, names = _read_biff_dictionary_rows(path)
        return any(
            name == wanted_sheet
            and any(_row_matches_record(values, record) for values in rows[ordinal].values())
            for ordinal, name in names.items()
            if ordinal in rows
        )
    except (BadZipFile, KeyError, OSError, ValueError, struct.error):
        return False
    finally:
        temporary.close()
        Path(temporary.name).unlink(missing_ok=True)


def _read_dta_dictionary(path: Path) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    with StataReader(path, convert_categoricals=False) as reader:
        variable_labels = reader.variable_labels()
        value_tables = reader.value_labels()
        variables = list(reader._varlist)
        table_names = dict(zip(variables, reader._lbllist))
        raw_types = list(
            getattr(reader, "_dtyplist", None)
            or getattr(reader, "_typlist", None)
            or []
        )
        for index, name in enumerate(variables):
            mapping = value_tables.get(table_names.get(name, ""), {})
            pairs: list[tuple[str, str]] = []
            missing: list[str] = []
            for raw_code, raw_label in sorted(
                mapping.items(), key=lambda item: (_number(item[0]), _text(item[1]))
            ):
                code, label = _number(raw_code), _text(raw_label)
                if isinstance(raw_code, StataMissingValue) or _is_missing_label(label):
                    missing.append(code)
                else:
                    pairs.append((code, label))
            result[name] = {
                "label": _text(variable_labels.get(name, "")),
                "type": _text(raw_types[index]) if index < len(raw_types) else "",
                "pairs": pairs,
                "missing": missing,
            }
    return result


def _read_sav_dictionary(path: Path) -> dict[str, dict[str, Any]]:
    file_size = path.stat().st_size
    with path.open("rb") as stream:
        header = stream.read(176)
        if len(header) < 176 or header[:4] != b"$FL2":
            raise ValueError("SAV_CABECERA_INVALIDA")
        layout_raw = header[64:68]
        if int.from_bytes(layout_raw, "little", signed=True) in {2, 3}:
            endian = "<"
        elif int.from_bytes(layout_raw, "big", signed=True) in {2, 3}:
            endian = ">"
        else:
            raise ValueError("SAV_LAYOUT_NO_SOPORTADO")

        def read_exact(size: int) -> bytes:
            value = stream.read(size)
            if len(value) != size:
                raise ValueError("SAV_DICCIONARIO_TRUNCADO")
            return value

        def read_int() -> int:
            return struct.unpack(endian + "i", read_exact(4))[0]

        variables: dict[str, dict[str, Any]] = {}
        positions: dict[int, str] = {}
        variable_records = 0
        while stream.tell() + 4 <= file_size:
            record_type = read_int()
            if record_type == 2:
                variable_records += 1
                variable_type, has_label, missing_count, _print, _write = struct.unpack(
                    endian + "iiiii", read_exact(20)
                )
                name = read_exact(8).rstrip(b" \x00").decode("cp1252", errors="replace")
                label = ""
                if has_label:
                    label_size = read_int()
                    if label_size < 0 or label_size > 1_048_576:
                        raise ValueError("SAV_ETIQUETA_TAMANO_INVALIDO")
                    label = read_exact(label_size).decode("cp1252", errors="replace")
                    read_exact((-label_size) % 4)
                slots = abs(missing_count)
                if slots > 3:
                    raise ValueError("SAV_MISSING_VALUES_INVALIDOS")
                raw_missing = [read_exact(8) for _ in range(slots)]
                if variable_type == -1:
                    continue
                positions[variable_records] = name
                decoded_missing = [
                    _decode_sav_code(value, variable_type, endian)
                    for value in raw_missing
                ]
                if missing_count < 0 and len(decoded_missing) >= 2:
                    decoded_missing[0] = f"{decoded_missing[0]}..{decoded_missing[1]}"
                    del decoded_missing[1]
                variables[name] = {
                    "label": label.strip(),
                    "type": "NUMERICO" if variable_type == 0 else f"STRING-{variable_type}",
                    "pairs": [],
                    "missing": decoded_missing,
                    "variable_type": variable_type,
                }
            elif record_type == 3:
                count = read_int()
                if count < 0 or count > 1_000_000:
                    raise ValueError("SAV_VALUE_LABEL_COUNT_INVALIDO")
                raw_labels: list[tuple[bytes, str]] = []
                for _ in range(count):
                    raw_code = read_exact(8)
                    size = read_exact(1)[0]
                    label = read_exact(size).decode("cp1252", errors="replace")
                    read_exact((-(size + 1)) % 8)
                    raw_labels.append((raw_code, label.strip()))
                if read_int() != 4:
                    raise ValueError("SAV_VALUE_LABEL_SIN_VARIABLES")
                variable_count = read_int()
                indexes = [read_int() for _ in range(variable_count)]
                for index in indexes:
                    name = positions.get(index)
                    if name is None:
                        continue
                    metadata = variables[name]
                    pairs = []
                    for raw_code, label in raw_labels:
                        code = _decode_sav_code(
                            raw_code, metadata["variable_type"], endian
                        )
                        if _is_missing_label(label):
                            if code not in metadata["missing"]:
                                metadata["missing"].append(code)
                        else:
                            pairs.append((code, label))
                    metadata["pairs"] = pairs
            elif record_type == 6:
                count = read_int()
                read_exact(count * 80)
            elif record_type == 7:
                _subtype, element_size, element_count = struct.unpack(
                    endian + "iii", read_exact(12)
                )
                total = element_size * element_count
                if element_size < 0 or element_count < 0 or total > file_size - stream.tell():
                    raise ValueError("SAV_EXTENSION_TAMANO_INVALIDO")
                read_exact(total)
            elif record_type == 999:
                read_exact(4)
                break
            else:
                raise ValueError(f"SAV_RECORD_TYPE_NO_SOPORTADO:{record_type}")
    for metadata in variables.values():
        metadata.pop("variable_type", None)
    if not variables:
        raise ValueError("SAV_SIN_DICCIONARIO_VARIABLES")
    return variables


def _decode_sav_code(payload: bytes, variable_type: int, endian: str) -> str:
    if variable_type == 0:
        return _number(struct.unpack(endian + "d", payload)[0])
    return payload[: min(variable_type, 8)].decode("cp1252", errors="replace").rstrip()


def _cell_text(value: Any) -> str:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return _number(value)
    return _text(value)


def _documented_categories(record: Mapping[str, Any]) -> list[str]:
    return [
        _cell_text(value)
        for value in record.get("categorias", [])
        if _cell_text(value)
        and _cell_text(value).upper() != "[REDACTADO-PRIVACIDAD]"
    ]


def _row_matches_record(values: tuple[Any, ...], record: Mapping[str, Any]) -> bool:
    cells = {_cell_text(value) for value in values if _cell_text(value)}
    variable = _text(record.get("nombre"))
    return variable in cells


def _variable_column(values: tuple[Any, ...], record: Mapping[str, Any]) -> int:
    variable = _text(record.get("nombre"))
    matches = [
        index for index, value in enumerate(values) if _cell_text(value) == variable
    ]
    if len(matches) != 1:
        return -1
    return matches[0]


def _dictionary_block_values(
    rows: list[tuple[Any, ...]], row_number: int, record: Mapping[str, Any]
) -> set[str]:
    variable_column = _variable_column(rows[row_number - 1], record)
    values: set[str] = set()
    if variable_column < 0:
        return {
            _cell_text(value)
            for value in rows[row_number - 1]
            if _cell_text(value)
        }
    for current_number in range(row_number, len(rows) + 1):
        row = rows[current_number - 1]
        if current_number > row_number:
            marker = row[variable_column] if variable_column < len(row) else None
            if _cell_text(marker):
                break
        values.update(_cell_text(value) for value in row if _cell_text(value))
    return values


def _dictionary_block_values_map(
    rows: Mapping[int, tuple[Any, ...]],
    row_number: int,
    record: Mapping[str, Any],
) -> set[str]:
    variable_column = _variable_column(rows[row_number], record)
    values: set[str] = set()
    if variable_column < 0:
        return {
            _cell_text(value) for value in rows[row_number] if _cell_text(value)
        }
    for current_number in sorted(number for number in rows if number >= row_number):
        if current_number > row_number + 256:
            break
        row = rows[current_number]
        if current_number > row_number:
            marker = row[variable_column] if variable_column < len(row) else None
            if _cell_text(marker):
                break
        values.update(_cell_text(value) for value in row if _cell_text(value))
    return values


def _matching_dictionary_rows(
    rows: list[tuple[Any, ...]], record: Mapping[str, Any]
) -> list[int]:
    matches = [
        index
        for index, values in enumerate(rows, 1)
        if _row_matches_record(values, record)
    ]
    if not matches:
        raise ValueError(
            "VARIABLE_NO_RESUELTA_EN_DICCIONARIO_XLSX"
        )
    return matches


def _matching_dictionary_rows_map(
    rows: Mapping[int, tuple[Any, ...]], record: Mapping[str, Any]
) -> list[int]:
    matches = [
        row_number
        for row_number, values in rows.items()
        if _row_matches_record(values, record)
    ]
    if not matches:
        raise ValueError("VARIABLE_NO_RESUELTA_EN_DICCIONARIO_XLS")
    return matches


def _unique_dictionary_row_map(
    rows: Mapping[int, tuple[Any, ...]],
    record: Mapping[str, Any],
    *,
    allow_none: bool = False,
) -> int | None:
    try:
        matches = _matching_dictionary_rows_map(rows, record)
    except ValueError:
        if allow_none:
            return None
        raise
    if len(matches) != 1:
        if allow_none:
            return None
        raise ValueError(
            f"FILA_DICCIONARIO_NO_UNICA_EN_XLS:coincidencias={len(matches)}"
        )
    return matches[0]


def _biff_sst(payload: bytes) -> list[str]:
    chunks: list[bytes] = []
    cuts: list[int] = []
    total = 0
    inside = False
    cursor = 0
    while cursor + 4 <= len(payload):
        record_type, length = struct.unpack("<HH", payload[cursor : cursor + 4])
        data = payload[cursor + 4 : cursor + 4 + length]
        if record_type == 0x00FC:
            inside = True
            chunks.append(data)
            total += len(data)
        elif inside and record_type == 0x003C:
            cuts.append(total)
            chunks.append(data)
            total += len(data)
        elif inside:
            break
        cursor += 4 + length
    if not chunks:
        return []
    blob = b"".join(chunks)
    boundaries = set(cuts)
    if len(blob) < 8:
        return []
    unique = struct.unpack("<I", blob[4:8])[0]
    position = 8
    strings: list[str] = []
    for _ in range(min(unique, 200_000)):
        if position + 3 > len(blob):
            break
        characters = struct.unpack("<H", blob[position : position + 2])[0]
        position += 2
        flags = blob[position]
        position += 1
        high = flags & 1
        rich = 0
        extended = 0
        if flags & 8:
            rich = struct.unpack("<H", blob[position : position + 2])[0]
            position += 2
        if flags & 4:
            extended = struct.unpack("<I", blob[position : position + 4])[0]
            position += 4
        letters: list[str] = []
        for _ in range(characters):
            if position in boundaries and position < len(blob):
                high = blob[position] & 1
                position += 1
            width = 2 if high else 1
            raw = blob[position : position + width]
            position += width
            letters.append(
                raw.decode("utf-16le" if width == 2 else "latin-1", errors="replace")
            )
        position += rich * 4 + extended
        strings.append("".join(letters))
    return strings


def _biff_rows(
    payload: bytes, offset: int, strings: list[str], maximum_row: int = 256
) -> dict[int, tuple[str, ...]]:
    rows: dict[int, dict[int, str]] = {}
    cursor = offset
    while cursor + 4 <= len(payload):
        record_type, length = struct.unpack("<HH", payload[cursor : cursor + 4])
        data = payload[cursor + 4 : cursor + 4 + length]
        if record_type == 0x000A:
            break
        if record_type == 0x00FD and length >= 10:
            row, column, _xf, string_index = struct.unpack("<HHHI", data[:10])
            if row < maximum_row and 0 <= string_index < len(strings):
                rows.setdefault(row + 1, {})[column] = strings[string_index]
        elif record_type == 0x0203 and length >= 14:
            row, column, _xf = struct.unpack("<HHH", data[:6])
            if row < maximum_row:
                rows.setdefault(row + 1, {})[column] = _number(
                    struct.unpack("<d", data[6:14])[0]
                )
        cursor += 4 + length
    result: dict[int, tuple[str, ...]] = {}
    for row_number, columns in rows.items():
        width = max(columns) + 1
        result[row_number] = tuple(columns.get(column, "") for column in range(width))
    return result


def _read_biff_dictionary_rows(
    path: Path,
) -> tuple[dict[int, dict[int, tuple[str, ...]]], dict[int, str]]:
    with olefile.OleFileIO(path) as document:
        stream_name = (
            "Workbook"
            if document.exists("Workbook")
            else "Book"
            if document.exists("Book")
            else ""
        )
        if not stream_name:
            raise ValueError("XLS_SIN_STREAM_WORKBOOK")
        payload = document.openstream(stream_name).read()
    sheets: list[tuple[int, str]] = []
    cursor = 0
    while cursor + 4 <= len(payload):
        record_type, length = struct.unpack("<HH", payload[cursor : cursor + 4])
        data = payload[cursor + 4 : cursor + 4 + length]
        if record_type == 0x0085 and len(data) >= 8:
            offset = struct.unpack("<I", data[0:4])[0]
            name_size = data[6]
            flags = data[7]
            width = 2 if flags & 1 else 1
            raw_name = data[8 : 8 + (name_size * width)]
            name = raw_name.decode(
                "utf-16le" if width == 2 else "latin-1", errors="replace"
            )
            sheets.append((offset, name))
        cursor += 4 + length
    strings = _biff_sst(payload)
    rows = {
        ordinal: _biff_rows(payload, offset, strings)
        for ordinal, (offset, _name) in enumerate(sheets, 1)
        if 0 < offset < len(payload)
    }
    names = {ordinal: name for ordinal, (_offset, name) in enumerate(sheets, 1)}
    return rows, names
