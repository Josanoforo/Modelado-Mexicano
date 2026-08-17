from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import shutil
import struct
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest import mock

import openpyxl
import yaml

from tools.curador_registro.barrido2_material import (
    AUTHORIZED_ROOTS,
    MATERIAL_BUILD_VERSION,
    MaterialDriftError,
    NetworkIsolationError,
    assert_network_disabled,
    assign_wave,
    build_material_snapshot,
    inspect_e2,
    inspect_task,
    logical_object_id,
    materialize_tasks,
    normalize_relative,
    representation_id,
    safe_text,
    valid_payload_id,
    validate_material_files,
    validate_material_snapshot,
    wave_concurrency_limit,
    _xls_objects,
    _xlsx_objects,
)
from tools.curador_registro.write_barrido2_w0 import CENSUS_SUFFIX, write_w0
from tools.curador_registro.write_barrido2_material import write_final


def digest(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def write_pdf(path: Path, pages: int) -> None:
    """PDF pequeño y válido, sin depender de librerías externas."""
    font_object = 3 + pages * 2
    objects: list[bytes] = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        ("<< /Type /Pages /Kids [" + " ".join(f"{3 + index * 2} 0 R" for index in range(pages)) + f"] /Count {pages} >>").encode(),
    ]
    for index in range(pages):
        page_object = 3 + index * 2
        stream_object = page_object + 1
        objects.append((f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 {font_object} 0 R >> >> /Contents {stream_object} 0 R >>").encode())
        content = f"BT /F1 12 Tf 72 720 Td (PAGINA {index + 1}) Tj ET".encode()
        objects.append(b"<< /Length " + str(len(content)).encode() + b" >>\nstream\n" + content + b"\nendstream")
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    payload = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for number, body in enumerate(objects, 1):
        offsets.append(len(payload))
        payload.extend(f"{number} 0 obj\n".encode() + body + b"\nendobj\n")
    xref = len(payload)
    payload.extend(f"xref\n0 {len(objects) + 1}\n".encode())
    payload.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        payload.extend(f"{offset:010d} 00000 n \n".encode())
    payload.extend(f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode())
    path.write_bytes(payload)


class Barrido2MaterialTests(unittest.TestCase):
    def setUp(self) -> None:
        self.synthetic_network = mock.patch.dict(os.environ, {"BARRIDO2_SYNTHETIC_TEST_ONLY": "1"})
        self.synthetic_network.start()
        self.network_attestation = mock.patch(
            "tools.curador_registro.barrido2_material.assert_network_disabled", return_value=None
        )
        self.network_attestation.start()
        self.tmp = tempfile.TemporaryDirectory()
        self.base = Path(self.tmp.name)
        self.raw = self.base / "raw"
        self.descargas = self.base / "descargas"
        self.raw.mkdir(); self.descargas.mkdir()
        self.roots = self.base / "raices.yaml"
        self.roots.write_text(yaml.safe_dump({
            "data_raw": str(self.raw),
            "descargas_mx": str(self.descargas),
            "downloads": str(self.base / "prohibida"),
        }, sort_keys=True), encoding="utf-8")
        self.contract = self.base / "contract.json"
        self.contract.write_text('{"network_habilitada":false}\n', encoding="utf-8")

    def tearDown(self) -> None:
        self.tmp.cleanup()
        self.network_attestation.stop()
        self.synthetic_network.stop()

    def _manifest(self, rows: list[dict[str, object]]) -> Path:
        path = self.base / "manifiesto.yaml"
        path.write_text(yaml.safe_dump(rows, allow_unicode=True, sort_keys=False), encoding="utf-8")
        return path

    def _snapshot(self, rows: list[dict[str, object]]) -> tuple[dict[str, object], Path]:
        path = self.base / "snapshot.json"
        snapshot = build_material_snapshot(self._manifest(rows), self.roots, path)
        return snapshot, path

    def test_payload_id_uses_closed_administrative_grammar(self) -> None:
        self.assertTrue(valid_payload_id("nota_metodologica_rotulo_pareada"))
        self.assertTrue(valid_payload_id("20260813130000_export_csv"))
        self.assertFalse(valid_payload_id("persona@example.test"))
        self.assertFalse(valid_payload_id("Alicia Perez"))
        self.assertFalse(valid_payload_id("NO-APLICA"))
        self.assertTrue(valid_payload_id("NO-APLICA", allow_no_aplica=True))

    def test_three_identities_are_not_collapsed_and_root_is_respected(self) -> None:
        raw_payload = b"raw"
        mx_payload = b"mx"
        (self.raw / "same.csv").write_bytes(raw_payload)
        (self.descargas / "same.csv").write_bytes(mx_payload)
        rows = [
            {"id": "P-RAW", "archivo": "same.csv", "sha256": digest(raw_payload), "tamano_bytes": len(raw_payload)},
            {"id": "P-MX", "raiz": "descargas_mx", "archivo": "same.csv", "sha256": digest(mx_payload), "tamano_bytes": len(mx_payload)},
        ]
        snapshot, _ = self._snapshot(rows)
        declarations = {row["payload_id"]: row for row in snapshot["declarations"]}
        self.assertNotEqual(declarations["P-RAW"]["representacion_id"], declarations["P-MX"]["representacion_id"])
        self.assertEqual(digest(raw_payload), declarations["P-RAW"]["sha256_observado"])
        self.assertEqual(digest(mx_payload), declarations["P-MX"]["sha256_observado"])
        self.assertEqual({"data_raw", "descargas_mx"}, {row["root_id"] for row in snapshot["representations"]})
        self.assertEqual(list(AUTHORIZED_ROOTS), snapshot["authorized_roots"])

        snapshot_path = self.base / "snapshot.json"
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        raw_task_path = next(
            path for path in task_root.glob("*.json")
            if json.loads(path.read_text(encoding="utf-8"))["payload_id"] == "P-RAW"
        )
        raw_task = json.loads(raw_task_path.read_text(encoding="utf-8"))
        staging = self.base / "staging" / raw_task_path.stem
        inspect_task(raw_task_path, self.roots, self.contract, staging, verify_network=False)
        record = json.loads((staging / "e2-neutral-index.jsonl").read_text(encoding="utf-8").splitlines()[0])
        self.assertEqual("P-RAW", record["payload_id"])
        self.assertEqual(raw_task["representacion_id"], record["representacion_id"])
        self.assertEqual(digest(raw_payload), record["sha256"])
        self.assertRegex(record["objeto_logico_id"], r"^OBJ-B2-[0-9a-f]{64}$")
        self.assertNotIn(record["payload_id"], {record["representacion_id"], record["sha256"], record["objeto_logico_id"]})

    def test_four_administrative_declarations_are_terminal_and_outside_census(self) -> None:
        rows = [{"id": f"ADMIN-{number}", "hecho": "declaración administrativa"} for number in range(4)]
        snapshot, _ = self._snapshot(rows)
        self.assertEqual(4, snapshot["counts"]["declaraciones_totales"])
        self.assertEqual(0, snapshot["counts"]["declaraciones_con_archivo_sha"])
        self.assertEqual(4, snapshot["counts"]["declaraciones_sin_archivo_sha"])
        self.assertEqual({"DECLARACION-SIN-ARCHIVO-SHA"}, {row["estado_administrativo"] for row in snapshot["declarations"]})
        self.assertEqual({"NO-APLICA"}, {row["representacion_id"] for row in snapshot["declarations"]})

    def test_w0_products_rederive_populations_and_preserve_census_prefix(self) -> None:
        payload = b"present"
        (self.raw / "present.csv").write_bytes(payload)
        self.contract.write_text(json.dumps({
            "base_sha": "a" * 40, "network_habilitada": False,
        }) + "\n", encoding="utf-8")
        rows = [
            {"id": "P-PRESENT", "archivo": "present.csv", "sha256": digest(payload), "tamano_bytes": len(payload)},
            {"id": "P-MISSING", "archivo": "missing.csv", "sha256": "b" * 64, "tamano_bytes": 2},
        ]
        snapshot, snapshot_path = self._snapshot(rows)
        tasks = self.base / "tasks"; task_ledger = self.base / "task-ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, tasks, task_ledger)
        frozen = self.base / "contract-hashes.json"
        frozen.write_text(json.dumps({
            "files": {"data/curacion-universo/contrato-barrido2-v1_0.json": digest(self.contract.read_bytes())},
        }), encoding="utf-8")
        previous = self.base / "previous.tsv"
        prefix = [
            "id_manifiesto", "archivo", "raiz", "tamano_bytes",
            "usado_para_declara_uso", "necesidades_que_lo_citan",
            "tsv_de_apertura_que_lo_toca", "estado", "universo_declarado",
            "consumo_detectado", "consumo_universo_declarado",
        ]
        previous.write_text("\t".join(prefix) + "\n", encoding="utf-8")
        output = self.base / "out"
        result = write_w0(snapshot_path, task_ledger, self.contract, frozen, previous, output, "2026-08-17")
        self.assertEqual({"censo": 2, "fuera_de_disco": 1, "ledger": 1}, {key: result[key] for key in ("censo", "fuera_de_disco", "ledger")})
        with (output / "data/censo-explotacion-2026-08-17.tsv").open(encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t"); census = list(reader)
        self.assertEqual(prefix + CENSUS_SUFFIX, reader.fieldnames)
        self.assertEqual({"P-PRESENT", "P-MISSING"}, {row["id_manifiesto"] for row in census})
        self.assertEqual({"data_raw"}, {row["raiz"] for row in census})
        self.assertEqual({"PRESENTE-INTEGRO", "FUERA-DE-DISCO"}, {row["estado_e0"] for row in census})
        fuera = (output / "data/fuera-de-disco-v1_0.tsv").read_text(encoding="utf-8")
        self.assertNotIn("no existe", fuera.casefold())
        baseline = json.loads((output / "data/curacion-universo/baseline-material-barrido2.json").read_text())
        self.assertFalse(baseline["network_habilitada"])
        self.assertEqual(snapshot["snapshot_sha256"], baseline["reports"]["snapshot_sha256"])

    def test_final_material_writer_keeps_full_private_index_and_compacts_durable_report(self) -> None:
        payload = b"variable,label\nx,Example\ny,Other\n"
        (self.raw / "dictionary.csv").write_bytes(payload)
        self.contract.write_text(json.dumps({
            "base_sha": "a" * 40, "network_habilitada": False,
        }) + "\n", encoding="utf-8")
        _, snapshot_path = self._snapshot([{
            "id": "PAYLOAD-1", "archivo": "dictionary.csv",
            "sha256": digest(payload), "tamano_bytes": len(payload),
        }])
        task_root = self.base / "tasks"; task_ledger = self.base / "task-ledger.tsv"
        staging_root = self.base / "staging"
        materialize_tasks(snapshot_path, self.contract, task_root, task_ledger)
        task_path = next(task_root.glob("*.json"))
        inspect_task(task_path, self.roots, self.contract, staging_root / task_path.stem, verify_network=False)
        materialize_tasks(snapshot_path, self.contract, task_root, task_ledger, staging_root)
        frozen = self.base / "contract-hashes.json"
        frozen.write_text(json.dumps({
            "files": {"data/curacion-universo/contrato-barrido2-v1_0.json": digest(self.contract.read_bytes())},
        }), encoding="utf-8")
        previous = self.base / "previous.tsv"
        prefix = [
            "id_manifiesto", "archivo", "raiz", "tamano_bytes",
            "usado_para_declara_uso", "necesidades_que_lo_citan",
            "tsv_de_apertura_que_lo_toca", "estado", "universo_declarado",
            "consumo_detectado", "consumo_universo_declarado",
        ]
        previous.write_text("\t".join(prefix) + "\n", encoding="utf-8")
        output = self.base / "out"
        write_w0(snapshot_path, task_ledger, self.contract, frozen, previous, output, "2026-08-17")
        private_index = self.base / "private/e2-neutral-index.jsonl"
        result = write_final(
            snapshot_path, task_ledger, task_root, staging_root, self.contract,
            frozen, output, private_index, "2026-08-17",
        )
        source_index = staging_root / task_path.stem / "e2-neutral-index.jsonl"
        self.assertEqual(source_index.read_bytes(), private_index.read_bytes())
        with (output / "data/curacion-universo/reportes-inspeccion-barrido2-v1_0.tsv").open(encoding="utf-8", newline="") as handle:
            reports = list(csv.DictReader(handle, delimiter="\t"))
        self.assertGreater(len(reports), 0)
        self.assertLessEqual(len(reports), result["e2_records"])
        self.assertTrue(all(row["afirmacion_tipo"] == "RESUMEN-NEUTRAL-COMPACTO" for row in reports))
        self.assertTrue(all(value and len(value) <= 160 for row in reports for value in row.values()))
        with (output / "data/curacion-universo/ledger-inspecciones-barrido2.tsv").open(encoding="utf-8", newline="") as handle:
            durable_ledger = list(csv.DictReader(handle, delimiter="\t"))
        self.assertEqual("0", durable_ledger[0]["excepciones"])
        with (output / "data/censo-explotacion-2026-08-17.tsv").open(encoding="utf-8", newline="") as handle:
            census = list(csv.DictReader(handle, delimiter="\t"))
        self.assertEqual("E2", census[0]["grado_inspeccion"])
        baseline = json.loads((output / "data/curacion-universo/baseline-material-barrido2.json").read_text())
        self.assertEqual(digest(private_index.read_bytes()), baseline["e2_index_sha256"])
        self.assertFalse(baseline["network_habilitada"])

        source_report = staging_root / task_path.stem / "reportes-durables.tsv"
        source_report.write_text(source_report.read_text(encoding="utf-8") + "\n", encoding="utf-8")
        rejected_output = self.base / "rejected"
        with self.assertRaisesRegex(ValueError, "GATE_MATERIAL_INVALIDO"):
            write_final(
                snapshot_path, task_ledger, task_root, staging_root, self.contract,
                frozen, rejected_output, self.base / "rejected-index.jsonl", "2026-08-17",
            )
        self.assertFalse(
            (rejected_output / "data/curacion-universo/reportes-inspeccion-barrido2-v1_0.tsv").exists()
        )

    def test_undeclared_physical_representation_keeps_no_payload(self) -> None:
        (self.raw / "oculto.txt").write_text("estructura", encoding="utf-8")
        snapshot, _ = self._snapshot([{"id": "ADMIN", "hecho": "sin payload"}])
        self.assertEqual(1, snapshot["counts"]["representaciones_no_declaradas"])
        representation = snapshot["representations"][0]
        self.assertEqual(["NO-APLICA"], representation["payload_ids"])
        self.assertEqual("NO-DECLARADA", representation["coincidencia_manifiesto"])

    def test_outside_disk_and_invalid_path_are_separate_terminals(self) -> None:
        rows = [
            {"id": "MISSING", "archivo": "missing.csv", "sha256": "0" * 64, "tamano_bytes": 1},
            {"id": "INVALID", "archivo": "../escape.csv", "sha256": "1" * 64, "tamano_bytes": 1},
            {"id": "ROOT", "raiz": "downloads", "archivo": "x.csv", "sha256": "2" * 64, "tamano_bytes": 1},
        ]
        snapshot, _ = self._snapshot(rows)
        states = {row["payload_id"]: row["estado_e0"] for row in snapshot["declarations"]}
        self.assertEqual("FUERA-DE-DISCO", states["MISSING"])
        self.assertEqual("RUTA-INVALIDA", states["INVALID"])
        self.assertEqual("RAIZ-NO-CONFIGURADA", states["ROOT"])

    def test_duplicate_bytes_share_sha_but_keep_representations(self) -> None:
        payload = b"identical"
        (self.raw / "a.bin").write_bytes(payload)
        (self.descargas / "b.bin").write_bytes(payload)
        snapshot, _ = self._snapshot([])
        representations = snapshot["representations"]
        self.assertEqual(2, len(representations))
        self.assertEqual(1, len({row["sha256"] for row in representations}))
        self.assertEqual(2, len({row["representacion_id"] for row in representations}))
        self.assertEqual({2}, {row["duplicate_content_count"] for row in representations})

    def test_identical_sha_reuses_one_e2_inspection_with_two_provenances(self) -> None:
        payload = b"column\nvalue\n"
        (self.raw / "a.csv").write_bytes(payload)
        (self.descargas / "b.csv").write_bytes(payload)
        snapshot, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        tasks = sorted(task_root.glob("*.json"))
        first_task = json.loads(tasks[0].read_text(encoding="utf-8"))
        second_task = json.loads(tasks[1].read_text(encoding="utf-8"))
        staging_root = self.base / "staging"
        first_dir = staging_root / tasks[0].stem
        second_dir = staging_root / tasks[1].stem
        inspect_task(tasks[0], self.roots, self.contract, first_dir, verify_network=False)
        with mock.patch("tools.curador_registro.barrido2_material.inspect_e2", side_effect=AssertionError("no debe reabrir")):
            second_summary = inspect_task(
                tasks[1], self.roots, self.contract, second_dir,
                verify_network=False, reuse_source_dir=first_dir,
            )
        self.assertEqual(first_task["representacion_id"], second_summary["reutilizada_desde_representacion_id"])
        first_records = [json.loads(line) for line in (first_dir / "e2-neutral-index.jsonl").read_text().splitlines()]
        second_records = [json.loads(line) for line in (second_dir / "e2-neutral-index.jsonl").read_text().splitlines()]
        self.assertEqual({row["objeto_logico_id"] for row in first_records}, {row["objeto_logico_id"] for row in second_records})
        self.assertEqual({second_task["representacion_id"]}, {row["representacion_id"] for row in second_records})
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        validation = validate_material_files(snapshot_path, self.contract, task_root, ledger, staging_root, require_complete=True)
        self.assertTrue(validation["ok"], validation["errors"])

    def test_same_path_with_two_hashes_produces_two_stable_ids(self) -> None:
        left = representation_id("data_raw", "folder/a.csv", "a" * 64)
        right = representation_id("data_raw", "folder/a.csv", "b" * 64)
        self.assertNotEqual(left, right)
        self.assertEqual(left, representation_id("data_raw", "folder/a.csv", "a" * 64))
        self.assertNotEqual(logical_object_id("a" * 64, "tabla/uno"), logical_object_id("a" * 64, "tabla/dos"))

        path = self.raw / "folder" / "a.csv"
        path.parent.mkdir()
        path.write_bytes(b"first")
        first_snapshot, _ = self._snapshot([])
        first_observed = first_snapshot["representations"][0]
        path.write_bytes(b"second")
        second_snapshot, _ = self._snapshot([])
        second_observed = second_snapshot["representations"][0]
        self.assertEqual(first_observed["ruta_relativa"], second_observed["ruta_relativa"])
        self.assertNotEqual(first_observed["sha256"], second_observed["sha256"])
        self.assertNotEqual(first_observed["representacion_id"], second_observed["representacion_id"])

    def test_normalized_relative_rejects_absolute_and_traversal(self) -> None:
        self.assertEqual("a/b.csv", normalize_relative("a/b.csv"))
        for value in ("/etc/passwd", "../x", "a/../x", ""):
            with self.assertRaises(ValueError):
                normalize_relative(value)

    def test_external_symlink_and_overlapping_roots_fail_closed(self) -> None:
        outside = self.base / "outside.txt"; outside.write_text("outside", encoding="utf-8")
        (self.raw / "escape.txt").symlink_to(outside)
        with self.assertRaisesRegex(MaterialDriftError, "SYMLINK_FUERA_DE_RAIZ"):
            self._snapshot([])
        (self.raw / "escape.txt").unlink()
        self.roots.write_text(yaml.safe_dump({
            "data_raw": str(self.raw), "descargas_mx": str(self.raw / "nested"),
        }), encoding="utf-8")
        (self.raw / "nested").mkdir()
        with self.assertRaisesRegex(ValueError, "RAICES_SOLAPADAS"):
            self._snapshot([])

    def test_root_self_alias_is_ignored_without_expanding_universe(self) -> None:
        (self.raw / "one.txt").write_text("one", encoding="utf-8")
        (self.raw / "raw").symlink_to(self.raw, target_is_directory=True)
        snapshot, _ = self._snapshot([])
        self.assertEqual(1, snapshot["counts"]["representaciones_fisicas"])
        self.assertEqual("one.txt", snapshot["representations"][0]["ruta_relativa"])

    def test_snapshot_output_cannot_enter_a_material_root(self) -> None:
        manifest = self._manifest([])
        with self.assertRaisesRegex(ValueError, "SNAPSHOT_OUTPUT_DENTRO_DE_RAIZ:data_raw"):
            build_material_snapshot(manifest, self.roots, self.raw / "snapshot.json")

    def test_wave_precedence_is_exclusive(self) -> None:
        self.assertEqual("W4", assign_wave("big.zip", 1024**3, {"max_ratio": 1}))
        self.assertEqual("W4", assign_wave("risk.zip", 1, {"max_ratio": 201}))
        self.assertEqual("W4", assign_wave("expanded.zip", 1, {"uncompressed": 2 * 1024**3}))
        self.assertEqual("W4", assign_wave("large.zip", 512 * 1024**2, {"max_ratio": 1}))
        self.assertEqual("W3", assign_wave("normal.zip", 1, {"max_ratio": 2}))
        self.assertEqual("W2", assign_wave("table.csv", 1))
        self.assertEqual("W1", assign_wave("image.png", 1))
        self.assertEqual(3, wave_concurrency_limit("W1"))
        self.assertEqual(3, wave_concurrency_limit("W2", ["csv"]))
        self.assertEqual(2, wave_concurrency_limit("W2", ["pdf", "csv"]))
        self.assertEqual(2, wave_concurrency_limit("W3"))
        self.assertEqual(1, wave_concurrency_limit("W4"))

    def test_plan_is_disjoint_exhaustive_blind_and_offline(self) -> None:
        (self.raw / "one.txt").write_text("# A", encoding="utf-8")
        (self.raw / "two.csv").write_text("x\n1\n", encoding="utf-8")
        with zipfile.ZipFile(self.raw / "three.zip", "w") as archive:
            archive.writestr("x.txt", "x")
        snapshot, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"
        ledger = self.base / "ledger.tsv"
        result = materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        self.assertEqual(3, result["tasks"])
        validation = validate_material_files(snapshot_path, self.contract, task_root, ledger)
        self.assertTrue(validation["ok"], validation["errors"])
        self.assertEqual(3, sum(validation["waves"].values()))
        for task_path in task_root.glob("*.json"):
            task = json.loads(task_path.read_text(encoding="utf-8"))
            self.assertFalse(task["network_habilitada"])
            self.assertFalse({"necesidad_id", "relacion_id", "clasificacion", "M-APERTURA", "FP-24"} & set(task))
            self.assertNotIn(str(self.raw), task_path.read_text(encoding="utf-8"))

    def test_resume_requires_exact_contract_and_sha(self) -> None:
        (self.raw / "one.txt").write_text("# A", encoding="utf-8")
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"
        ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json"))
        staging_root = self.base / "staging"
        staging = staging_root / task_path.stem
        inspect_task(task_path, self.roots, self.contract, staging, verify_network=False)
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        with ledger.open(encoding="utf-8", newline="") as handle:
            reused = list(csv.DictReader(handle, delimiter="\t"))[0]
        self.assertEqual("SI", reused["estado_terminal"])
        summary_path = staging / "resumen.json"
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        summary["parser_version"] = "OTRO-BUILD"
        summary_path.write_text(json.dumps(summary, sort_keys=True), encoding="utf-8")
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        with ledger.open(encoding="utf-8", newline="") as handle:
            invalidated = list(csv.DictReader(handle, delimiter="\t"))[0]
        self.assertEqual("NO", invalidated["estado_terminal"])
        summary["parser_version"] = MATERIAL_BUILD_VERSION
        summary_path.write_text(json.dumps(summary, sort_keys=True), encoding="utf-8")
        (self.raw / "one.txt").write_text("# B", encoding="utf-8")
        with self.assertRaisesRegex(Exception, "TAREA_RUTA_HASH_NO_RECONCILIA"):
            inspect_task(task_path, self.roots, self.contract, self.base / "changed", verify_network=False)
        summary["parser_version"] = MATERIAL_BUILD_VERSION
        summary_path.write_text(json.dumps(summary, sort_keys=True), encoding="utf-8")
        self.contract.write_text('{"network_habilitada":false,"version":2}\n', encoding="utf-8")
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        with ledger.open(encoding="utf-8", newline="") as handle:
            contract_invalidated = list(csv.DictReader(handle, delimiter="\t"))[0]
        self.assertEqual("NO", contract_invalidated["estado_terminal"])

    def test_manifest_change_invalidates_task_even_with_same_representation(self) -> None:
        payload = b"# heading\n"; (self.raw / "one.txt").write_bytes(payload)
        _, snapshot_path = self._snapshot([{
            "id": "PAYLOAD-A", "archivo": "one.txt", "sha256": digest(payload), "tamano_bytes": len(payload),
        }])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"; staging_root = self.base / "staging"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        old_task = next(task_root.glob("*.json")); staging = staging_root / old_task.stem
        inspect_task(old_task, self.roots, self.contract, staging, verify_network=False)
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        self._snapshot([{
            "id": "PAYLOAD-B", "archivo": "one.txt", "sha256": digest(payload), "tamano_bytes": len(payload),
        }])
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        with ledger.open(encoding="utf-8", newline="") as handle:
            row = next(csv.DictReader(handle, delimiter="\t"))
        self.assertEqual("PAYLOAD-B", row["payload_id"])
        self.assertEqual("NO", row["estado_terminal"])
        self.assertNotEqual(old_task.stem, row["tarea_id"])
        self.assertFalse(old_task.exists())
        validation = validate_material_files(snapshot_path, self.contract, task_root, ledger)
        self.assertTrue(validation["ok"], validation["errors"])

    def test_multiple_payloads_remain_explicit_in_snapshot_and_ledger(self) -> None:
        payload = b"x\n"; (self.raw / "shared.csv").write_bytes(payload)
        snapshot, snapshot_path = self._snapshot([
            {"id": "P-B", "archivo": "shared.csv", "sha256": digest(payload), "tamano_bytes": len(payload)},
            {"id": "P-A", "archivo": "shared.csv", "sha256": digest(payload), "tamano_bytes": len(payload)},
        ])
        self.assertEqual(["P-A", "P-B"], snapshot["representations"][0]["payload_ids"])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        with ledger.open(encoding="utf-8", newline="") as handle:
            row = next(csv.DictReader(handle, delimiter="\t"))
        self.assertEqual(["P-A", "P-B"], json.loads(row["payload_ids_json"]))
        self.assertEqual("P-A", row["payload_id"])
        self.assertTrue(validate_material_files(snapshot_path, self.contract, task_root, ledger)["ok"])

    def test_mutation_during_open_produces_no_expediente(self) -> None:
        payload = b"# before\n"; path = self.raw / "one.txt"; path.write_bytes(payload)
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); staging = self.base / "staging" / task_path.stem

        def mutate(_: Path) -> tuple[list[dict[str, object]], str, str]:
            path.write_bytes(b"# after\n")
            return ([{"locator": "x", "type": "TEXTO", "name": "x"}], "test", "completo")

        with mock.patch("tools.curador_registro.barrido2_material.inspect_e2", side_effect=mutate):
            with self.assertRaisesRegex(MaterialDriftError, "MATERIAL_CAMBIO_DURANTE_APERTURA"):
                inspect_task(task_path, self.roots, self.contract, staging, verify_network=False)
        self.assertFalse(staging.exists())

    def test_internal_symlink_retarget_during_open_is_detected(self) -> None:
        payload = b"# same\n"
        (self.raw / "target-a.txt").write_bytes(payload); (self.raw / "target-b.txt").write_bytes(payload)
        link = self.raw / "link.txt"; link.symlink_to(self.raw / "target-a.txt")
        _, snapshot_path = self._snapshot([{
            "id": "LINK", "archivo": "link.txt", "sha256": digest(payload), "tamano_bytes": len(payload),
        }])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(
            path for path in task_root.glob("*.json")
            if json.loads(path.read_text(encoding="utf-8"))["ruta_relativa"] == "link.txt"
        )

        def retarget(_: Path) -> tuple[list[dict[str, object]], str, str]:
            link.unlink(); link.symlink_to(self.raw / "target-b.txt")
            return ([{"locator": "documento=texto", "type": "TEXTO", "name": "texto"}], "test", "completo")

        with mock.patch("tools.curador_registro.barrido2_material.inspect_e2", side_effect=retarget):
            with self.assertRaisesRegex(MaterialDriftError, "MATERIAL_CAMBIO_DURANTE_APERTURA"):
                inspect_task(task_path, self.roots, self.contract, self.base / "staging", verify_network=False)

    def test_reuse_rejects_same_sha_with_different_format(self) -> None:
        payload = b"a,b\n1,2\n"
        (self.raw / "one.csv").write_bytes(payload); (self.descargas / "one.txt").write_bytes(payload)
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"; staging_root = self.base / "staging"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        tasks = [json.loads(path.read_text(encoding="utf-8")) | {"path": path} for path in task_root.glob("*.json")]
        csv_task = next(row for row in tasks if row["formato"] == ".csv")
        text_task = next(row for row in tasks if row["formato"] == ".txt")
        source = staging_root / csv_task["tarea_id"]
        inspect_task(csv_task["path"], self.roots, self.contract, source, verify_network=False)
        with self.assertRaisesRegex(ValueError, "REUSE_EXPEDIENTE_NO_EXACTO"):
            inspect_task(
                text_task["path"], self.roots, self.contract, staging_root / text_task["tarea_id"],
                verify_network=False, reuse_source_dir=source,
            )

    def test_reuse_rejects_different_sha_with_same_format(self) -> None:
        (self.raw / "one.csv").write_bytes(b"a\n1\n")
        (self.descargas / "two.csv").write_bytes(b"a\n2\n")
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"; staging_root = self.base / "staging"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        tasks = sorted(task_root.glob("*.json")); source = staging_root / tasks[0].stem
        inspect_task(tasks[0], self.roots, self.contract, source, verify_network=False)
        with self.assertRaisesRegex(ValueError, "REUSE_EXPEDIENTE_NO_EXACTO"):
            inspect_task(tasks[1], self.roots, self.contract, staging_root / tasks[1].stem, verify_network=False, reuse_source_dir=source)

    def test_reuse_preserves_privacy_redaction(self) -> None:
        payload = b"variable,label\nx,persona@example.test\n"
        (self.raw / "dictionary-one.csv").write_bytes(payload); (self.descargas / "dictionary-two.csv").write_bytes(payload)
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"; staging_root = self.base / "staging"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        tasks = sorted(task_root.glob("*.json")); first = staging_root / tasks[0].stem; second = staging_root / tasks[1].stem
        inspect_task(tasks[0], self.roots, self.contract, first, verify_network=False)
        inspect_task(tasks[1], self.roots, self.contract, second, verify_network=False, reuse_source_dir=first)
        first_records = [json.loads(line) for line in (first / "e2-neutral-index.jsonl").read_text().splitlines()]
        second_records = [json.loads(line) for line in (second / "e2-neutral-index.jsonl").read_text().splitlines()]
        self.assertIn("[REDACTADO-PRIVACIDAD]", {row["privacidad"] for row in first_records})
        self.assertEqual(
            [row["privacidad"] for row in sorted(first_records, key=lambda row: row["objeto_logico_id"])],
            [row["privacidad"] for row in sorted(second_records, key=lambda row: row["objeto_logico_id"])],
        )

    def test_reuse_matches_fresh_logical_objects_across_physical_names(self) -> None:
        payload = b"a,b\n1,2\n"
        (self.raw / "a.csv").write_bytes(payload); (self.descargas / "b.csv").write_bytes(payload)
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"; staging_root = self.base / "staging"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        tasks = sorted(task_root.glob("*.json")); first = staging_root / "first"; reused = staging_root / "reused"; fresh = staging_root / "fresh"
        inspect_task(tasks[0], self.roots, self.contract, first, verify_network=False)
        inspect_task(tasks[1], self.roots, self.contract, reused, verify_network=False, reuse_source_dir=first)
        inspect_task(tasks[1], self.roots, self.contract, fresh, verify_network=False)
        reused_records = [json.loads(line) for line in (reused / "e2-neutral-index.jsonl").read_text().splitlines()]
        fresh_records = [json.loads(line) for line in (fresh / "e2-neutral-index.jsonl").read_text().splitlines()]
        project = lambda rows: sorted(
            (row["objeto_logico_id"], row["objeto_padre_id"], row["localizador"], row["nombre"], row["definicion"], row["depth"])
            for row in rows
        )
        self.assertEqual(project(fresh_records), project(reused_records))

    def test_e2_csv_has_every_column_and_no_rows_in_outputs(self) -> None:
        secret = "persona@example.test"
        payload = f"variable,etiqueta\na,{secret}\nb,otro\n"
        (self.raw / "dictionary.csv").write_text(payload, encoding="utf-8")
        _, snapshot_path = self._snapshot([{
            "id": "P", "archivo": "dictionary.csv", "sha256": digest(payload.encode()), "tamano_bytes": len(payload.encode()),
        }])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); staging = self.base / "staging" / "worker" / task_path.stem
        result = inspect_task(task_path, self.roots, self.contract, staging, verify_network=False)
        self.assertEqual(5, result["objetos_e1"])
        combined = (staging / "e2-neutral-index.jsonl").read_text(encoding="utf-8") + (staging / "reportes-durables.tsv").read_text(encoding="utf-8")
        self.assertNotIn(secret, combined)
        self.assertNotIn("\na,", combined)
        self.assertIn("filas=2;columnas=2", combined)
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, self.base / "staging")
        validation = validate_material_files(
            snapshot_path, self.contract, task_root, ledger, self.base / "staging",
            require_complete=True,
        )
        self.assertTrue(validation["ok"], validation["errors"])

    def test_e2_json_preserves_schema_not_individual_values(self) -> None:
        payload = {"rows": [{"name": "Alicia", "age": 40}, {"name": "Beatriz", "age": 50, "district": "sur"}]}
        raw = json.dumps(payload).encode()
        (self.raw / "table.json").write_bytes(raw)
        _, snapshot_path = self._snapshot([{"id": "P", "archivo": "table.json", "sha256": digest(raw), "tamano_bytes": len(raw)}])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); staging = self.base / "staging" / task_path.stem
        inspect_task(task_path, self.roots, self.contract, staging, verify_network=False)
        output = (staging / "e2-neutral-index.jsonl").read_text(encoding="utf-8")
        self.assertIn("$.rows[*].name", output)
        self.assertIn("$.rows[*].district", output)
        self.assertNotIn("Alicia", output)
        self.assertNotIn("Beatriz", output)
        self.assertNotIn(":40", output)

        keyed = self.raw / "keyed.json"
        keyed.write_text(json.dumps({"Alicia Pérez": {"age": 40}, "Beatriz López": {"age": 50}}), encoding="utf-8")
        keyed_objects, _, _ = inspect_e2(keyed)
        keyed_output = json.dumps(keyed_objects, ensure_ascii=False)
        self.assertNotIn("Alicia Pérez", keyed_output)
        self.assertNotIn("Beatriz López", keyed_output)
        self.assertIn("$[*]", keyed_output)

    def test_e2_xlsx_enumerates_all_sheets(self) -> None:
        path = self.raw / "book.xlsx"
        workbook = openpyxl.Workbook()
        workbook.active.title = "Uno"; workbook.active.append(["a", "b"])
        workbook.create_sheet("Dos").append(["c"])
        workbook.save(path); workbook.close()
        raw = path.read_bytes()
        _, snapshot_path = self._snapshot([{"id": "P", "archivo": path.name, "sha256": digest(raw), "tamano_bytes": len(raw)}])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); staging = self.base / "staging" / task_path.stem
        inspect_task(task_path, self.roots, self.contract, staging, verify_network=False)
        output = (staging / "e2-neutral-index.jsonl").read_text(encoding="utf-8")
        self.assertIn("hoja=Uno", output); self.assertIn("hoja=Dos", output)

        no_header = self.raw / "no-header.xlsx"
        workbook = openpyxl.Workbook(); workbook.active.append(["PERSONA PRUEBA", 42]); workbook.save(no_header); workbook.close()
        no_header_objects, _, _ = inspect_e2(no_header)
        self.assertNotIn("PERSONA PRUEBA", json.dumps(no_header_objects, ensure_ascii=False))

        empty = self.raw / "empty-sheet.xlsx"
        workbook = openpyxl.Workbook(); workbook.active.title = "Vacia"
        workbook.create_sheet("ConDatos").append(["valor"])
        workbook.save(empty); workbook.close()
        empty_objects, _, _ = inspect_e2(empty)
        sheets = [row for row in empty_objects if row["type"] == "HOJA-XLSX"]
        self.assertEqual(["Vacia", "ConDatos"], [row["name"] for row in sheets])

        class EmptyReadOnlySheet:
            title = "VaciaReal"
            max_row = None
            max_column = None
            tables: dict[str, object] = {}

            def iter_rows(self, **_kwargs):
                raise AssertionError("una hoja sin dimensiones no debe iterarse")

        fake_workbook = mock.Mock(worksheets=[EmptyReadOnlySheet()])
        with mock.patch("tools.curador_registro.barrido2_material.openpyxl.load_workbook", return_value=fake_workbook):
            truly_empty = _xlsx_objects(empty)
        self.assertEqual(1, len(truly_empty))
        self.assertIn("filas=0;columnas=0", truly_empty[0]["definition"])
        fake_workbook.close.assert_called_once()

    def test_e2_pdf_opens_every_page_not_only_first_five(self) -> None:
        path = self.raw / "six.pdf"
        write_pdf(path, 6)
        objects, parser, boundary = inspect_e2(path)
        pages = [row for row in objects if row["type"] == "PAGINA-PDF"]
        self.assertEqual(6, len(pages))
        self.assertEqual({1, 2, 3, 4, 5, 6}, {row["page"] for row in pages})
        self.assertIn("poppler", parser)
        self.assertIn("todas las páginas", boundary)

    def test_e2_zip_enumerates_nested_container_without_extracting(self) -> None:
        inner = self.base / "inner.zip"
        with zipfile.ZipFile(inner, "w") as archive:
            archive.writestr("dictionary.csv", "variable,label\nx,Example\n")
        outer = self.raw / "outer.zip"
        with zipfile.ZipFile(outer, "w") as archive:
            archive.writestr("nested.zip", inner.read_bytes())
            archive.writestr("readme.txt", "# Heading")
        objects, _, boundary = inspect_e2(outer)
        locators = {row["locator"] for row in objects}
        self.assertTrue(any(locator.endswith(":nested.zip") for locator in locators))
        self.assertTrue(any(":nested.zip!/miembro=1:dictionary.csv" in locator for locator in locators))
        self.assertIn("anidados", boundary)
        self.assertFalse(any(path.name == "dictionary.csv" for path in self.base.rglob("dictionary.csv")))

    def test_e2_dta_metadata_and_sav_specific_exception(self) -> None:
        import pandas as pd
        dta = self.raw / "sample.dta"
        pd.DataFrame({"variable": [1, 2], "category": [0, 1]}).to_stata(
            dta, write_index=False, variable_labels={"variable": "Variable label"},
            value_labels={"category": {0: "No", 1: "Yes"}},
        )
        objects, _, boundary = inspect_e2(dta)
        variables = {row["name"]: row for row in objects if row["type"] == "VARIABLE-DTA"}
        self.assertEqual({"variable", "category"}, set(variables))
        self.assertEqual("Variable label", variables["variable"]["label"])
        self.assertIn("observaciones no persistidas", boundary)
        private_names = self.raw / "private-names.dta"
        pd.DataFrame({"first_name": [1], "second_name": [2]}).to_stata(private_names, write_index=False)
        private_objects, _, _ = inspect_e2(private_names)
        private_variables = [row for row in private_objects if row["type"] == "VARIABLE-DTA"]
        self.assertEqual(2, len(private_variables))
        self.assertEqual(2, len({row["locator"] for row in private_variables}))
        self.assertEqual({"[REDACTADO-PRIVACIDAD]"}, {row["name"] for row in private_variables})
        sav = self.raw / "sample.sav"
        header = bytearray(176); header[:4] = b"$FL2"; header[64:68] = struct.pack("<i", 2)
        variable_label = b"Age label"
        variable_record = struct.pack("<iiiiii8s", 2, 0, 1, 0, 0, 0, b"AGE     ")
        variable_record += struct.pack("<i", len(variable_label)) + variable_label + b"\x00" * ((-len(variable_label)) % 4)
        value_label = b"Adult"
        value_record = struct.pack("<ii", 3, 1) + struct.pack("<dB", 1.0, len(value_label)) + value_label
        value_record += b"\x00" * ((-(len(value_label) + 1)) % 8)
        value_record += struct.pack("<iii", 4, 1, 1)
        extension_record = struct.pack("<iiii", 7, 99, 1, 4) + b"ABCD"
        sav.write_bytes(bytes(header) + variable_record + value_record + extension_record + struct.pack("<ii", 999, 0))
        sav_objects, _, sav_boundary = inspect_e2(sav)
        self.assertEqual("Age label", next(row for row in sav_objects if row["type"] == "VARIABLE-SAV")["label"])
        self.assertIn("label=Adult", next(row for row in sav_objects if row["type"] == "VALUE-LABEL-COLLECTION-SAV")["value_labels"][0])
        self.assertEqual("EXCEPCION-ESPECIFICA", next(row for row in sav_objects if row["type"] == "EXTENSION-DICCIONARIO-SAV")["state"])
        self.assertIn("observaciones no leídas", sav_boundary)

    def test_xls_redacted_sheet_names_keep_distinct_private_locators(self) -> None:
        def boundsheet(name: str) -> bytes:
            encoded = name.encode("latin-1")
            data = b"\0" * 6 + bytes((len(encoded), 0)) + encoded
            return struct.pack("<HH", 0x0085, len(data)) + data

        payload = boundsheet("first_name") + boundsheet("second_name")

        class FakeOle:
            def __enter__(self): return self
            def __exit__(self, *_args): return False
            def exists(self, name: str) -> bool: return name == "Workbook"
            def openstream(self, _name: str): return io.BytesIO(payload)

        with mock.patch("tools.curador_registro.barrido2_material.olefile.OleFileIO", return_value=FakeOle()):
            objects = _xls_objects(self.raw / "private.xls")
        self.assertEqual(2, len(objects))
        self.assertEqual(2, len({row["locator"] for row in objects}))
        self.assertEqual({"[REDACTADO-PRIVACIDAD]"}, {row["name"] for row in objects})

    def test_e2_docx_html_xml_and_text_cover_structural_objects(self) -> None:
        docx = self.raw / "sample.docx"
        document_xml = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
        <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body>
        <w:p><w:r><w:t>Question?</w:t></w:r></w:p><w:tbl><w:tr><w:tc><w:p><w:r><w:t>PERSONA PRUEBA</w:t></w:r></w:p></w:tc></w:tr></w:tbl>
        </w:body></w:document>'''
        with zipfile.ZipFile(docx, "w") as archive:
            archive.writestr("word/document.xml", document_xml)
        docx_objects, _, _ = inspect_e2(docx)
        self.assertTrue({"PARTE-DOCX", "PARRAFO-DOCX", "TABLA-DOCX"}.issubset({row["type"] for row in docx_objects}))
        self.assertNotIn("PERSONA PRUEBA", json.dumps(docx_objects, ensure_ascii=False))
        html = self.raw / "sample.html"; html.write_text("<html><body><h1>Title</h1><form><label>Field</label></form><table><th>Column</th></table></body></html>", encoding="utf-8")
        html_objects, _, _ = inspect_e2(html)
        self.assertTrue({"H1", "FORM", "LABEL", "TABLE", "TH"}.issubset({row["type"] for row in html_objects}))
        xml = self.raw / "sample.xml"; xml.write_text("<root><section><item value='secret'>text</item></section></root>", encoding="utf-8")
        xml_objects, _, _ = inspect_e2(xml)
        self.assertEqual(3, len(xml_objects))
        self.assertNotIn("secret", json.dumps(xml_objects))
        text = self.raw / "sample.php"; text.write_text("# Heading\nordinary individual value\n", encoding="utf-8")
        text_objects, _, _ = inspect_e2(text)
        self.assertEqual(2, len(text_objects))
        self.assertNotIn("ordinary individual value", json.dumps(text_objects))

    def test_zip_slip_and_corruption_are_observed_at_e0(self) -> None:
        with zipfile.ZipFile(self.raw / "risk.zip", "w") as archive:
            archive.writestr("../escape.csv", "a\n1\n")
        (self.raw / "bad.zip").write_bytes(b"not a zip")
        snapshot, _ = self._snapshot([])
        rows = {row["ruta_relativa"]: row for row in snapshot["representations"]}
        self.assertTrue(rows["risk.zip"]["zip_slip"])
        self.assertEqual("CORRUPTO", rows["bad.zip"]["estado_e0"])

        corrupt = (self.raw / "bad.zip").read_bytes()
        snapshot, _ = self._snapshot([{
            "id": "BAD", "archivo": "bad.zip", "sha256": "0" * 64, "tamano_bytes": len(corrupt),
        }])
        declaration = snapshot["declarations"][0]
        self.assertEqual("CORRUPTO", declaration["estado_e0"])
        self.assertEqual("NO", declaration["hash_coincide_manifiesto"])

    def test_unknown_zip_compression_becomes_specific_exception(self) -> None:
        path = self.raw / "unsupported.zip"
        with zipfile.ZipFile(path, "w") as archive:
            archive.writestr("member.txt", "payload")
        payload = bytearray(path.read_bytes())
        local = payload.index(b"PK\x03\x04"); central = payload.index(b"PK\x01\x02")
        payload[local + 8:local + 10] = (99).to_bytes(2, "little")
        payload[central + 10:central + 12] = (99).to_bytes(2, "little")
        path.write_bytes(payload)
        objects, _, _ = inspect_e2(path)
        self.assertTrue(any(row.get("state") == "EXCEPCION-ESPECIFICA" for row in objects))

    def test_real_high_ratio_zip_is_assigned_only_to_w4(self) -> None:
        with zipfile.ZipFile(self.raw / "compressed.zip", "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("zeros.bin", b"0" * (2 * 1024 * 1024))
        snapshot, _ = self._snapshot([])
        representation = snapshot["representations"][0]
        self.assertGreater(representation["zip_geometry"]["max_ratio"], 200)
        self.assertEqual("W4", representation["wave_initial"])

    def test_encrypted_zip_flag_is_terminal_cifrado(self) -> None:
        path = self.raw / "encrypted.zip"
        with zipfile.ZipFile(path, "w") as archive:
            archive.writestr("member.txt", "payload")
        payload = bytearray(path.read_bytes())
        local = payload.index(b"PK\x03\x04")
        central = payload.index(b"PK\x01\x02")
        payload[local + 6:local + 8] = (int.from_bytes(payload[local + 6:local + 8], "little") | 1).to_bytes(2, "little")
        payload[central + 8:central + 10] = (int.from_bytes(payload[central + 8:central + 10], "little") | 1).to_bytes(2, "little")
        path.write_bytes(payload)
        snapshot, _ = self._snapshot([])
        representation = snapshot["representations"][0]
        self.assertEqual("CIFRADO", representation["estado_e0"])

    def test_privacy_redaction_preserves_record(self) -> None:
        text, redacted = safe_text("contacto persona@example.test")
        self.assertTrue(redacted)
        self.assertEqual("[REDACTADO-PRIVACIDAD]", text)
        name, name_redacted = safe_text("Alicia Perez")
        self.assertTrue(name_redacted)
        self.assertEqual("[REDACTADO-PRIVACIDAD]", name)
        for variant in ("ALICIA PEREZ", "alicia perez", "ALICIA_PEREZ"):
            self.assertEqual(("[REDACTADO-PRIVACIDAD]", True), safe_text(variant))
        for absolute_path in ("/", "/tmp/private/file.tsv", r"C:\\Users\\Persona\\file.tsv"):
            self.assertEqual(("[REDACTADO-PRIVACIDAD]", True), safe_text(absolute_path))
        compact, compact_redacted = safe_text("x" * 240, durable=True)
        self.assertFalse(compact_redacted)
        self.assertEqual(160, len(compact))

    def test_uncertain_first_row_and_cp1252_never_become_individual_values(self) -> None:
        (self.raw / "no-header.csv").write_bytes("PERSONA PRUEBA,42\nOTRA PERSONA,43\n".encode("cp1252"))
        objects, _, _ = inspect_e2(self.raw / "no-header.csv")
        output = json.dumps(objects, ensure_ascii=False)
        self.assertNotIn("PERSONA PRUEBA", output)
        self.assertNotIn("OTRA PERSONA", output)
        dictionary = self.raw / "dictionary-cp1252.csv"
        dictionary.write_bytes("variable,label\nx,niño\n".encode("cp1252"))
        dictionary_objects, _, _ = inspect_e2(dictionary)
        self.assertIn("niño", json.dumps(dictionary_objects, ensure_ascii=False))

    def test_namespace_verification_is_fail_closed(self) -> None:
        fake_stat = mock.Mock(st_ino=42)
        with mock.patch("os.stat", return_value=fake_stat):
            with self.assertRaises(NetworkIsolationError):
                assert_network_disabled()

        with mock.patch(
            "tools.curador_registro.barrido2_material.assert_network_disabled",
            side_effect=NetworkIsolationError("SIN_ATESTACION"),
        ):
            with self.assertRaisesRegex(NetworkIsolationError, "SIN_ATESTACION"):
                payload = b"x"; (self.raw / "x.txt").write_bytes(payload)
                _, snapshot_path = self._snapshot([])
                task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
                materialize_tasks(snapshot_path, self.contract, task_root, ledger)
                inspect_task(next(task_root.glob("*.json")), self.roots, self.contract, self.base / "staging", verify_network=False)

    def test_manual_task_cannot_inject_private_payload_id(self) -> None:
        (self.raw / "one.txt").write_text("x", encoding="utf-8")
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); task = json.loads(task_path.read_text())
        task["payload_id"] = "persona@example.test"; task_path.write_text(json.dumps(task), encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "TAREA_PAYLOAD_ID_PRIVADO_O_INVALIDO"):
            inspect_task(task_path, self.roots, self.contract, self.base / "staging", verify_network=False)

    def test_administrative_payload_slug_is_not_reinterpreted_as_pii(self) -> None:
        payload = b"# estructura\n"; (self.raw / "one.txt").write_bytes(payload)
        _, snapshot_path = self._snapshot([{
            "id": "nota_metodologica_rotulo_pareada", "archivo": "one.txt",
            "sha256": digest(payload), "tamano_bytes": len(payload),
        }])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        staging_root = self.base / "staging"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json"))
        inspect_task(task_path, self.roots, self.contract, staging_root / task_path.stem, verify_network=False)
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        validation = validate_material_files(
            snapshot_path, self.contract, task_root, ledger, staging_root,
            require_complete=True,
        )
        self.assertTrue(validation["ok"], validation["errors"])

    def test_task_budget_is_enforced_not_merely_documented(self) -> None:
        (self.raw / "one.txt").write_text("x", encoding="utf-8")
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); task = json.loads(task_path.read_text())
        task["presupuesto"]["timeout_segundos"] = 1801
        task_path.write_text(json.dumps(task), encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "TAREA_PRESUPUESTO_INVALIDO"):
            inspect_task(task_path, self.roots, self.contract, self.base / "staging", verify_network=False)

    @unittest.skipUnless(shutil.which("unshare"), "unshare no está instalado")
    def test_namespace_real_blocks_egress(self) -> None:
        environment = os.environ.copy()
        environment["BARRIDO2_OUTER_NET_NS_INODE"] = str(os.stat("/proc/self/ns/net").st_ino)
        result = subprocess.run(
            [
                "unshare", "-Urn", "--", sys.executable, "-c",
                "from tools.curador_registro.barrido2_material import assert_network_disabled; assert_network_disabled()",
            ],
            cwd=Path(__file__).resolve().parents[3],
            env=environment,
            capture_output=True,
            text=True,
            timeout=10,
            check=False,
        )
        self.assertEqual(0, result.returncode, result.stderr)

    def test_complete_gate_rejects_truncated_durable_report(self) -> None:
        payload = b"a,b\n1,2\n"
        (self.raw / "one.csv").write_bytes(payload)
        _, snapshot_path = self._snapshot([{
            "id": "P", "archivo": "one.csv", "sha256": digest(payload), "tamano_bytes": len(payload),
        }])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); staging_root = self.base / "staging"
        staging = staging_root / task_path.stem
        inspect_task(task_path, self.roots, self.contract, staging, verify_network=False)
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        report_path = staging / "reportes-durables.tsv"
        lines = report_path.read_text(encoding="utf-8").splitlines()
        report_path.write_text("\n".join(lines[:-1]) + "\n", encoding="utf-8")
        summary_path = staging / "resumen.json"
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        summary["report_sha256"] = digest(report_path.read_bytes())
        summary_path.write_text(json.dumps(summary, sort_keys=True), encoding="utf-8")
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        with ledger.open(encoding="utf-8", newline="") as handle:
            self.assertEqual("NO", next(csv.DictReader(handle, delimiter="\t"))["estado_terminal"])
        validation = validate_material_files(
            snapshot_path, self.contract, task_root, ledger, staging_root, require_complete=True,
        )
        self.assertFalse(validation["ok"])
        self.assertTrue(any("REPORTE_DURABLE_NO_CUBRE_INDICE_1A1" in error for error in validation["errors"]))

    def test_durable_compaction_is_end_to_end_and_private_index_stays_complete(self) -> None:
        long_label = "x" * 320
        payload = f"variable,label\nx,{long_label}\n".encode()
        (self.raw / "dictionary.csv").write_bytes(payload)
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"; staging_root = self.base / "staging"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); staging = staging_root / task_path.stem
        inspect_task(task_path, self.roots, self.contract, staging, verify_network=False)
        private_index = (staging / "e2-neutral-index.jsonl").read_text(encoding="utf-8")
        self.assertIn(long_label, private_index)
        with (staging / "reportes-durables.tsv").open(encoding="utf-8", newline="") as handle:
            durable = list(csv.DictReader(handle, delimiter="\t"))
        self.assertTrue(durable)
        self.assertTrue(all(len(value) <= 160 for row in durable for value in row.values()))
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        validation = validate_material_files(snapshot_path, self.contract, task_root, ledger, staging_root, require_complete=True)
        self.assertTrue(validation["ok"], validation["errors"])

    def test_validator_rejects_tampered_narrative_duplicate_summary_and_extra_task(self) -> None:
        payload = b"# heading\n"; (self.raw / "one.txt").write_bytes(payload)
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"; staging_root = self.base / "staging"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); staging = staging_root / task_path.stem
        inspect_task(task_path, self.roots, self.contract, staging, verify_network=False)
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        report_path = staging / "reportes-durables.tsv"
        with report_path.open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle, delimiter="\t")); fields = list(rows[0])
        rows[0]["descripcion_neutral"] = "narración adulterada"
        with report_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", lineterminator="\n"); writer.writeheader(); writer.writerows(rows)
        summary_path = staging / "resumen.json"; summary = json.loads(summary_path.read_text())
        summary["report_sha256"] = digest(report_path.read_bytes()); summary["batch_sha256"] = "0" * 64
        summary_path.write_text(json.dumps(summary), encoding="utf-8")
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        duplicate = staging_root / "duplicate"; shutil.copytree(staging, duplicate)
        (task_root / "EXTRA.json").write_text("{}\n", encoding="utf-8")
        validation = validate_material_files(snapshot_path, self.contract, task_root, ledger, staging_root, require_complete=True)
        joined = "\n".join(validation["errors"])
        self.assertIn("TASK_ROOT_NO_CUBRE_LEDGER_1A1", joined)
        self.assertIn("EXPEDIENTE_REPRESENTACION_NO_UNICO", joined)
        self.assertIn("E2_SUMMARY_BATCH_HASH_INVALIDO", joined)
        self.assertIn("REPORTE_DURABLE_NO_DEREFERENCIABLE", joined)

    def test_validator_rejects_joint_task_ledger_mutation_and_foreign_summary(self) -> None:
        (self.raw / "one.txt").write_text("x", encoding="utf-8")
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"; staging_root = self.base / "staging"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); task = json.loads(task_path.read_text())
        task["root_id"] = "descargas_mx"; task_path.write_text(json.dumps(task), encoding="utf-8")
        with ledger.open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle, delimiter="\t")); fields = list(rows[0])
        rows[0]["root_id"] = "descargas_mx"
        with ledger.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", lineterminator="\n"); writer.writeheader(); writer.writerows(rows)
        foreign = staging_root / "foreign"; foreign.mkdir(parents=True)
        (foreign / "resumen.json").write_text(json.dumps({"representacion_id": "REP-" + "f" * 64}), encoding="utf-8")
        validation = validate_material_files(snapshot_path, self.contract, task_root, ledger, staging_root)
        joined = "\n".join(validation["errors"])
        self.assertIn("LEDGER_SNAPSHOT_JOIN_INVALIDO", joined)
        self.assertIn("EXPEDIENTE_REPRESENTACION_AJENA", joined)

    def test_validator_reports_missing_task_instead_of_aborting(self) -> None:
        (self.raw / "one.txt").write_text("x", encoding="utf-8")
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"; staging_root = self.base / "staging"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); staging = staging_root / task_path.stem
        inspect_task(task_path, self.roots, self.contract, staging, verify_network=False)
        materialize_tasks(snapshot_path, self.contract, task_root, ledger, staging_root)
        task_path.unlink()
        validation = validate_material_files(snapshot_path, self.contract, task_root, ledger, staging_root, require_complete=True)
        self.assertFalse(validation["ok"])
        self.assertTrue(any("TAREA_INEXISTENTE" in error for error in validation["errors"]))

    def test_w5_cannot_be_an_initial_executable_task(self) -> None:
        (self.raw / "one.txt").write_text("x", encoding="utf-8")
        _, snapshot_path = self._snapshot([])
        task_root = self.base / "tasks"; ledger = self.base / "ledger.tsv"
        materialize_tasks(snapshot_path, self.contract, task_root, ledger)
        task_path = next(task_root.glob("*.json")); task = json.loads(task_path.read_text())
        task["wave_initial"] = "W5"; task_path.write_text(json.dumps(task), encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "TAREA_OLA_INICIAL_INVALIDA"):
            inspect_task(task_path, self.roots, self.contract, self.base / "staging", verify_network=False)

    def test_nested_zip_magic_and_depth_exception_are_complete(self) -> None:
        current = self.base / "level-6.zip"
        with zipfile.ZipFile(current, "w") as archive:
            archive.writestr("deep.txt", "# deepest")
        for level in range(5, -1, -1):
            parent = self.base / f"level-{level}.zip"
            with zipfile.ZipFile(parent, "w") as archive:
                archive.writestr("inner.bin" if level == 0 else "inner.zip", current.read_bytes())
            current = parent
        objects, _, _ = inspect_e2(current)
        serialized = json.dumps(objects, ensure_ascii=False)
        self.assertIn("inner.bin!/miembro=1:inner.zip", serialized)
        self.assertTrue(any(row.get("definition") == "EXCEPCION-ESPECIFICA:profundidad>4" for row in objects))
        self.assertEqual(len(objects), len({row["locator"] for row in objects}))

    def test_snapshot_validator_recomputes_ids_hash_and_counts(self) -> None:
        (self.raw / "x.txt").write_text("x", encoding="utf-8")
        snapshot, _ = self._snapshot([])
        self.assertEqual([], validate_material_snapshot(snapshot))
        snapshot["representations"][0]["representacion_id"] = "REP-" + "0" * 64
        self.assertTrue(validate_material_snapshot(snapshot))
        snapshot, _ = self._snapshot([])
        snapshot["counts"]["declaraciones_totales"] = 999
        self.assertIn("CONTEOS_SNAPSHOT_NO_RECONCILIAN", validate_material_snapshot(snapshot))
        snapshot, _ = self._snapshot([])
        snapshot["snapshot_sha256"] = "0" * 64
        self.assertIn("SNAPSHOT_HASH_INVALIDO", validate_material_snapshot(snapshot))


if __name__ == "__main__":
    unittest.main()
